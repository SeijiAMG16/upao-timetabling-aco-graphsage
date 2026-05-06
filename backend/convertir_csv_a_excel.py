"""
Convierte el CSV generado por ACO al formato Excel del exportador
Usa el formato de exportar_horarios_un_archivo.py
"""
import sys
import pandas as pd
from app.database import SessionLocal
from app.models import CourseSection, Course, Professor, Classroom, TimeSlot
from sqlalchemy.orm import joinedload
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def crear_hoja_profesor(wb, nombre_hoja, profesor_data, asignaciones_profesor):
    """
    Crea una hoja en el workbook para un profesor específico
    Formato: Grilla Día x Hora
    """
    ws = wb.create_sheet(title=nombre_hoja)
    
    # CONFIGURACIÓN
    dias_semana = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']
    
    # Definir bloques horarios (16 timeslots de 50 minutos)
    bloques_horarios = [
        ('07:00', '07:50'),
        ('07:55', '08:45'),
        ('08:50', '09:40'),
        ('09:45', '10:35'),
        ('10:40', '11:30'),
        ('11:35', '12:25'),
        ('12:30', '13:20'),
        ('13:25', '14:15'),
        ('14:20', '15:10'),
        ('15:15', '16:05'),
        ('16:10', '17:00'),
        ('17:05', '17:55'),
        ('18:00', '18:50'),
        ('18:55', '19:45'),
        ('19:50', '20:40'),
        ('20:45', '21:35'),
    ]
    
    # === HEADER ===
    ws.merge_cells('A1:H1')
    cell_title = ws['A1']
    cell_title.value = f"HORARIO - {profesor_data['nombre']}"
    cell_title.font = Font(size=14, bold=True)
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    cell_title.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell_title.font = Font(size=14, bold=True, color='FFFFFF')
    
    # Info profesor
    ws['A2'] = 'Código:'
    ws['B2'] = profesor_data.get('codigo', 'N/A')
    ws['A3'] = 'Total horas:'
    ws['B3'] = profesor_data.get('total_horas', 0)
    
    # === GRILLA ===
    row_start = 5
    
    # Header de columnas (días)
    ws.merge_cells(f'A{row_start}:A{row_start}')
    ws[f'A{row_start}'] = 'HORA'
    ws[f'A{row_start}'].font = Font(bold=True)
    ws[f'A{row_start}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_start}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for col_idx, dia in enumerate(dias_semana, start=2):
        cell = ws.cell(row=row_start, column=col_idx)
        cell.value = dia
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Crear matriz horario: {(dia_semana, bloque): asignacion}
    # Y también rastrear bloques ocupados
    matriz_horario = {}
    bloques_procesados = set()  # Para evitar duplicados al combinar celdas
    
    for asig in asignaciones_profesor:
        # Organizar timeslots por día
        timeslots_por_dia = {}
        for slot_id in asig['timeslot_ids']:
            ts = asig['timeslots'][slot_id]
            dia = ts['dia_semana']
            bloque = ts['orden'] - 1  # orden es 1-indexed
            if dia not in timeslots_por_dia:
                timeslots_por_dia[dia] = []
            timeslots_por_dia[dia].append(bloque)
        
        # Para cada día, ordenar bloques y detectar consecutivos
        for dia, bloques in timeslots_por_dia.items():
            bloques_ordenados = sorted(bloques)
            # Marcar el primer bloque de cada secuencia consecutiva
            for i, bloque in enumerate(bloques_ordenados):
                key = (dia, bloque)
                if key not in matriz_horario:
                    # Calcular cuántos bloques consecutivos hay
                    bloques_consecutivos = 1
                    for j in range(i + 1, len(bloques_ordenados)):
                        if bloques_ordenados[j] == bloques_ordenados[j-1] + 1:
                            bloques_consecutivos += 1
                        else:
                            break
                    
                    matriz_horario[key] = {
                        'asignacion': asig,
                        'bloques_consecutivos': bloques_consecutivos,
                        'es_primero': True  # Este es el primer bloque
                    }
                    
                    # Marcar los bloques siguientes como "ocupados pero no primeros"
                    for offset in range(1, bloques_consecutivos):
                        next_key = (dia, bloque + offset)
                        matriz_horario[next_key] = {
                            'asignacion': asig,
                            'bloques_consecutivos': 0,  # No es el primero
                            'es_primero': False
                        }
    
    # Llenar grilla CON MERGE DE CELDAS
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for bloque_idx, (hora_inicio, hora_fin) in enumerate(bloques_horarios):
        row = row_start + 1 + bloque_idx
        
        # Columna de hora
        cell_hora = ws.cell(row=row, column=1)
        cell_hora.value = f"{hora_inicio}\n{hora_fin}"
        cell_hora.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_hora.font = Font(size=9)
        cell_hora.border = thin_border
        
        # Columnas de días
        for dia_idx, dia_nombre in enumerate(dias_semana, start=1):
            col = dia_idx + 1
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            key = (dia_idx, bloque_idx)
            if key in matriz_horario:
                info = matriz_horario[key]
                
                # Solo procesar si es el PRIMER bloque de la secuencia
                if info['es_primero'] and info['bloques_consecutivos'] > 0:
                    asig = info['asignacion']
                    bloques_span = info['bloques_consecutivos']
                    
                    # COMBINAR CELDAS si hay más de 1 bloque
                    if bloques_span > 1:
                        start_row = row
                        end_row = row + bloques_span - 1
                        ws.merge_cells(
                            start_row=start_row, start_column=col,
                            end_row=end_row, end_column=col
                        )
                    
                    # Llenar contenido
                    texto = f"{asig['nombre_curso']}\n{asig['tipo_sesion']}-L{asig['liga']}\n{asig['aula']}\nAlum: {asig['alumnos']}"
                    cell.value = texto
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=8)
                    
                    # Color según tipo
                    color_fill = None
                    if asig['tipo_sesion'] == 'T':
                        color_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif asig['tipo_sesion'] == 'P':
                        color_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    elif asig['tipo_sesion'] == 'L':
                        color_fill = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
                    
                    if color_fill:
                        cell.fill = color_fill
                        # Aplicar color a todas las celdas combinadas
                        if bloques_span > 1:
                            for offset in range(1, bloques_span):
                                next_cell = ws.cell(row=row + offset, column=col)
                                next_cell.fill = color_fill
                                next_cell.border = thin_border
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    for col_idx in range(2, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # Ajustar altura de filas
    for row_idx in range(row_start + 1, row_start + 1 + len(bloques_horarios)):
        ws.row_dimensions[row_idx].height = 50


def convertir_csv_a_excel(csv_path):
    """
    Lee el CSV generado por ACO y crea Excel con formato del exportador
    """
    print(f"Leyendo {csv_path}...")
    df = pd.read_csv(csv_path)
    
    db = SessionLocal()
    
    try:
        # Cargar timeslots
        print("Cargando timeslots...")
        timeslots_dict = {}
        for ts in db.query(TimeSlot).all():
            timeslots_dict[ts.id] = {
                'dia_semana': ts.dia_semana,
                'hora_inicio': ts.hora_inicio,
                'hora_fin': ts.hora_fin,
                'orden': ts.orden
            }
        
        # Cargar profesores
        print("Cargando profesores...")
        profesores_dict = {}
        for prof in db.query(Professor).all():
            profesores_dict[prof.id] = {
                'nombre': prof.nombre_completo,
                'codigo': prof.codigo
            }
        
        # Cargar aulas
        print("Cargando aulas...")
        aulas_dict = {}
        for aula in db.query(Classroom).all():
            aulas_dict[aula.id] = aula.codigo
        
        # Cargar cursos para obtener nombres
        print("Cargando cursos...")
        cursos_dict = {}
        for curso in db.query(Course).all():
            cursos_dict[curso.codigo] = curso.nombre
        
        # Organizar asignaciones por profesor
        print("Organizando asignaciones por profesor...")
        asignaciones_por_profesor = {}
        
        for _, row in df.iterrows():
            prof_id = int(row['Profesor ID'])
            aula_id = int(row['Aula ID'])
            
            # Parsear franjas horarias
            franjas_str = str(row['Franjas Horarias'])
            timeslot_ids = [int(x.strip()) for x in franjas_str.split(',')]
            
            asignacion = {
                'section_id': row['Section ID'],
                'codigo_curso': row['Codigo Curso'],
                'nombre_curso': cursos_dict.get(row['Codigo Curso'], row['Codigo Curso']),  # Nombre o código si no se encuentra
                'tipo_sesion': row['Tipo Sesion'],
                'liga': row['Liga'],
                'ciclo': row['Ciclo'],
                'aula': aulas_dict.get(aula_id, f'Aula-{aula_id}'),
                'alumnos': row['Alumnos Proyectados'],
                'timeslot_ids': timeslot_ids,
                'timeslots': {tid: timeslots_dict[tid] for tid in timeslot_ids if tid in timeslots_dict}
            }
            
            if prof_id not in asignaciones_por_profesor:
                asignaciones_por_profesor[prof_id] = []
            asignaciones_por_profesor[prof_id].append(asignacion)
        
        # Crear workbook
        print("Creando archivo Excel...")
        wb = Workbook()
        ws_index = wb.active
        ws_index.title = "ÍNDICE"
        
        # Header índice
        ws_index['A1'] = 'ÍNDICE DE PROFESORES'
        ws_index['A1'].font = Font(size=14, bold=True)
        ws_index.merge_cells('A1:C1')
        
        ws_index['A2'] = 'Profesor'
        ws_index['B2'] = 'Código'
        ws_index['C2'] = 'Horas Totales'
        for cell in ['A2', 'B2', 'C2']:
            ws_index[cell].font = Font(bold=True)
            ws_index[cell].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Crear hojas por profesor
        row_idx = 3
        for prof_id in sorted(asignaciones_por_profesor.keys()):
            asignaciones = asignaciones_por_profesor[prof_id]
            prof_data = profesores_dict.get(prof_id, {'nombre': f'Profesor-{prof_id}', 'codigo': 'N/A'})
            
            # Calcular horas totales
            total_horas = 0
            for asig in asignaciones:
                total_horas += len(asig['timeslot_ids']) * 0.833  # 50 minutos = 0.833 horas
            
            prof_data['total_horas'] = round(total_horas, 1)
            
            # Nombre corto para la hoja (máx 31 caracteres)
            nombre_hoja = prof_data['nombre'][:28]
            if len(prof_data['nombre']) > 28:
                nombre_hoja += '...'
            
            # Crear hoja
            crear_hoja_profesor(wb, nombre_hoja, prof_data, asignaciones)
            
            # Agregar a índice
            ws_index[f'A{row_idx}'] = prof_data['nombre']
            ws_index[f'B{row_idx}'] = prof_data['codigo']
            ws_index[f'C{row_idx}'] = prof_data['total_horas']
            row_idx += 1
            
            print(f"  [OK] {prof_data['nombre']}: {len(asignaciones)} asignaciones, {prof_data['total_horas']} horas")
        
        # Ajustar anchos índice
        ws_index.column_dimensions['A'].width = 40
        ws_index.column_dimensions['B'].width = 15
        ws_index.column_dimensions['C'].width = 15
        
        # Guardar
        output_path = csv_path.replace('.csv', '_formato_profesores.xlsx')
        wb.save(output_path)
        print(f"\n[OK] Archivo Excel creado: {output_path}")
        print(f"   Total profesores: {len(asignaciones_por_profesor)}")
        
        return output_path
        
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    else:
        # Buscar el CSV más reciente
        csv_files = [f for f in os.listdir('.') if f.startswith('horario_generado_') and f.endswith('.csv')]
        if not csv_files:
            print("[ERR] No se encontró ningún archivo CSV de horario generado")
            sys.exit(1)
        csv_file = max(csv_files, key=os.path.getctime)
    
    print(f"Convirtiendo: {csv_file}")
    convertir_csv_a_excel(csv_file)
