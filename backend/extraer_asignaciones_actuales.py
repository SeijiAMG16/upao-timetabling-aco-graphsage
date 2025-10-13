"""
================================================================================
EXTRACTOR DE ASIGNACIONES Y RESTRICCIONES DESDE EXCEL
================================================================================
Lee el archivo Horario_Docentes(2025-20).xlsx para extraer:
1. Asignaciones profesor-curso actuales
2. Restricciones de disponibilidad por día y bloque horario
================================================================================
"""

import pandas as pd
import json
import re
from pathlib import Path
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuración
EXCEL_PATH = Path('../inputs/Horario_Docentes(2025-20).xlsx')
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'sistemas',
    'database': 'upao_timetabling'
}

DIAS_SEMANA = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']

# Mapeo manual de nombres abreviados del Excel a IDs de profesores en BD
MAPEO_PROFESORES = {
    'A. Caballero': 162,      # CABALLERO ALVARADO, ARMANDO
    'C.Cuba': 161,            # CAROLA LIZETH CUBA CASTILLO
    'C.Gay': 163,             # Carlos Gaytan Toledo
    'C. Guijon': None,        # No encontrado
    'C. Julca': 169,          # Carlos Edwin Julca Castillo
    'C.Mend': 174,            # MENDOZA CORPUS CARLOS
    'E.Cieza': 176,           # CIEZA MOSTACERO SEGUNDO EDWIN
    'E. Chav': 152,           # Edilberto Chavez Fernandez
    'E.SantaC': 160,          # SANTA CRUZ, ELIAS
    'Espinola': 157,          # Espinola
    'E.Mir': 179,             # Eddy Miranda Velasquez
    'F.Inf': 150,             # Freddy Infantes Quiroz
    'F.Per': 165,             # Fernando Perez Cueva
    'F.Cas': 175,             # Fernando Castillo Robles
    'H.Aba': 151,             # Heber Abanto Cabrera
    'H. Mendoza': 149,        # Henry Mendoza Puerta
    'H.Sag': 167,             # Hernan Sagastegui Chigne
    'J. Baylon': None,        # No encontrado
    'J.Cal': 154,             # Jose Calderon Sedano
    'J.Cast': 159,            # Jose Castañeda Saldaña
    'J.Dia': 156,             # Jaime Diaz Sanchez
    'J. Gutierrez': None,     # No encontrado
    'J.Hua': 171,             # Jorge Huapaya Escobedo
    'J.Jar': 178,             # Jorge Jara Arenas
    'J.Pim': 173,             # Jorge Piminchumo Flores
    'J.Vasquez': None,        # No encontrado
    'K.Mel': 155,             # Zoraida Vidal Melgarejo (K podría ser Zoraida)
    'L.Vla': 177,             # Luis Vladimir Urrelo
    'L.Llanos': None,         # No encontrado
    'M. Llerena': 158,        # LLERENA FERNANDEZ, MONICA
    'Moises': None,           # No encontrado
    'STAFF': 172,             # CONVOCATORIA
    'S.Rodri': 153,           # Silvia Rodriguez Aguirre
    'Sheyli': 168,            # VALVERDE VELA SHEYLI
    'W.Cue': 166,             # Walter Cueva Chavez
    'W.Lazo': 164,            # Walter Lazo
    'W.Letur': 170,           # Walter Leturia
    'Z.Vidal': 155            # Zoraida Vidal Melgarejo
}

# Bloques horarios UPAO (17 bloques de 50 min)
BLOQUES_TIEMPO = [
    ("07:00", "07:50"),   # Bloque 1
    ("07:55", "08:45"),   # Bloque 2
    ("08:50", "09:40"),   # Bloque 3
    ("09:45", "10:35"),   # Bloque 4
    ("10:40", "11:30"),   # Bloque 5
    ("11:35", "12:25"),   # Bloque 6
    ("12:30", "13:20"),   # Bloque 7
    ("13:25", "14:15"),   # Bloque 8
    ("14:20", "15:10"),   # Bloque 9
    ("15:15", "16:05"),   # Bloque 10
    ("16:10", "17:00"),   # Bloque 11
    ("17:05", "17:55"),   # Bloque 12
    ("18:00", "18:50"),   # Bloque 13
    ("18:55", "19:45"),   # Bloque 14
    ("19:50", "20:40"),   # Bloque 15
    ("20:45", "21:35"),   # Bloque 16
    ("21:40", "22:30")    # Bloque 17
]


def conectar_db():
    """Conecta a la base de datos"""
    return mysql.connector.connect(**DB_CONFIG)


def normalizar_nombre_curso(nombre):
    """Normaliza el nombre del curso para búsqueda"""
    if not nombre or not isinstance(nombre, str):
        return ""
    
    # Remover texto entre paréntesis (T1, P1, L1, etc.)
    nombre = re.sub(r'\([^)]*\)', '', nombre)
    
    # Limpiar espacios y convertir a mayúsculas
    nombre = nombre.strip().upper()
    
    # Remover caracteres especiales
    nombre = re.sub(r'[^\w\s]', '', nombre)
    
    return nombre


def buscar_curso_en_bd(nombre_curso):
    """Busca un curso en la BD por nombre"""
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    nombre_normalizado = normalizar_nombre_curso(nombre_curso)
    
    # Buscar por coincidencia parcial
    cursor.execute("""
        SELECT id, codigo, nombre
        FROM courses
        WHERE UPPER(REPLACE(nombre, '.', '')) LIKE %s
        OR UPPER(codigo) LIKE %s
        LIMIT 1
    """, (f'%{nombre_normalizado}%', f'%{nombre_normalizado}%'))
    
    resultado = cursor.fetchone()
    cursor.close()
    conn.close()
    
    return resultado


def buscar_profesor_en_bd(nombre_apellido):
    """Busca un profesor en la BD por nombre completo"""
    if not nombre_apellido or not isinstance(nombre_apellido, str):
        return None
    
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    # Limpiar el nombre
    nombre_limpio = nombre_apellido.strip().upper()
    
    # Buscar por nombre completo o apellido
    cursor.execute("""
        SELECT id, nombre_completo, nombres, apellidos
        FROM professors
        WHERE UPPER(nombre_completo) LIKE %s
        OR UPPER(apellidos) LIKE %s
        OR UPPER(nombres) LIKE %s
        LIMIT 1
    """, (f'%{nombre_limpio}%', f'%{nombre_limpio}%', f'%{nombre_limpio}%'))
    
    resultado = cursor.fetchone()
    cursor.close()
    conn.close()
    
    return resultado


def extraer_tipo_liga(texto):
    """Extrae el tipo (T/P/L) y liga (1,2,3) de un texto"""
    if not texto or not isinstance(texto, str):
        return None, None
    
    # Buscar patrón (T1), (P2), (L3), etc.
    match = re.search(r'\(([TPL])(\d+)\)', texto.upper())
    if match:
        tipo = match.group(1)
        liga = int(match.group(2))
        return tipo, liga
    
    return None, None


def es_celda_bloqueada(cell):
    """Verifica si una celda está bloqueada (color rojo de fondo)"""
    if not cell.fill or not cell.fill.start_color:
        return False
    
    # Colores rojos/rosados comunes en Excel
    colores_rojo = ['FFFF0000', 'FFFF6B6B', 'FFFF9999', 'FFFFC7CE', 'FFFCD5D5', 'FF0000', 'FFC00000']
    
    color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
    
    return any(cr in color.upper() for cr in colores_rojo)


def extraer_asignaciones_y_restricciones():
    """Extrae asignaciones profesor-curso y restricciones de disponibilidad"""
    
    print("="*80)
    print("📊 EXTRACTOR DE ASIGNACIONES Y RESTRICCIONES")
    print("="*80)
    
    if not EXCEL_PATH.exists():
        print(f"❌ ERROR: No se encontró el archivo {EXCEL_PATH}")
        return None, None
    
    print(f"\n📂 Cargando Excel: {EXCEL_PATH}")
    
    # Cargar workbook con openpyxl para acceder a estilos
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    asignaciones = []
    restricciones = {}
    
    # Procesar cada hoja (cada profesor)
    for sheet_name in wb.sheetnames:
        print(f"\n📄 Procesando hoja: {sheet_name}")
        ws = wb[sheet_name]
        
        # Buscar profesor usando el mapeo
        profesor_id = MAPEO_PROFESORES.get(sheet_name)
        
        if not profesor_id:
            print(f"  ⚠️  Profesor no mapeado: {sheet_name}")
            continue
        
        # Obtener info del profesor
        conn_temp = conectar_db()
        cursor_temp = conn_temp.cursor(dictionary=True)
        cursor_temp.execute("SELECT id, nombre_completo FROM professors WHERE id = %s", (profesor_id,))
        profesor = cursor_temp.fetchone()
        cursor_temp.close()
        conn_temp.close()
        
        if not profesor:
            print(f"  ⚠️  Profesor ID {profesor_id} no encontrado en BD")
            continue
        
        print(f"  ✅ Profesor encontrado: {profesor['nombre_completo']} (ID: {profesor_id})")
        
        # Inicializar restricciones del profesor
        restricciones[profesor_id] = {
            dia: [False] * 17 for dia in DIAS_SEMANA  # False = disponible, True = bloqueado
        }
        
        # Encontrar la fila de encabezados (días de la semana)
        fila_dias = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if 'LUNES' in cell.value.upper():
                        fila_dias = row_idx
                        break
            if fila_dias:
                break
        
        if not fila_dias:
            print(f"  ⚠️  No se encontró fila de días")
            continue
        
        # Mapear columnas a días
        columnas_dias = {}
        for cell in ws[fila_dias]:
            if cell.value and isinstance(cell.value, str):
                dia_upper = cell.value.strip().upper()
                dia_upper = dia_upper.replace('É', 'E').replace('Á', 'A')
                for dia in DIAS_SEMANA:
                    if dia in dia_upper:
                        columnas_dias[cell.column] = dia
                        break
        
        print(f"  📅 Días encontrados: {list(columnas_dias.values())}")
        
        # Procesar filas de horarios
        fila_inicio = fila_dias + 1
        
        for row_idx in range(fila_inicio, min(fila_inicio + 20, ws.max_row + 1)):
            row = list(ws[row_idx])
            
            # Primera columna debería tener el horario
            horario_cell = row[0]
            
            # Procesar cada día
            for col_idx, dia in columnas_dias.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Verificar si está bloqueada (rojo)
                if es_celda_bloqueada(cell):
                    # Mapear a bloque de tiempo
                    bloque_idx = row_idx - fila_inicio
                    if 0 <= bloque_idx < 17:
                        restricciones[profesor_id][dia][bloque_idx] = True
                
                # Verificar si hay asignación de curso
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 3:
                    contenido = cell.value.strip()
                    
                    # Extraer tipo y liga
                    tipo, liga = extraer_tipo_liga(contenido)
                    
                    # Buscar curso
                    nombre_curso = normalizar_nombre_curso(contenido)
                    curso = buscar_curso_en_bd(nombre_curso)
                    
                    if curso and tipo and liga:
                        # Determinar bloque de tiempo
                        bloque_idx = row_idx - fila_inicio
                        if 0 <= bloque_idx < 17:
                            asignacion = {
                                'profesor_id': profesor_id,
                                'profesor_nombre': profesor['nombre_completo'],
                                'curso_id': curso['id'],
                                'curso_nombre': curso['nombre'],
                                'curso_codigo': curso['codigo'],
                                'tipo': tipo,
                                'liga': liga,
                                'session_type': f"{tipo}{liga}",
                                'dia': dia,
                                'bloque_inicio': bloque_idx,
                                'hora_inicio': BLOQUES_TIEMPO[bloque_idx][0],
                                'contenido_original': contenido
                            }
                            asignaciones.append(asignacion)
                            print(f"    ➜ {curso['codigo']} ({tipo}{liga}) - {dia} bloque {bloque_idx+1}")
    
    # Resumen
    print("\n" + "="*80)
    print("📊 RESUMEN DE EXTRACCIÓN")
    print("="*80)
    print(f"✅ Asignaciones encontradas: {len(asignaciones)}")
    print(f"✅ Profesores con restricciones: {len(restricciones)}")
    
    # Contar restricciones
    total_bloques_bloqueados = 0
    for prof_id, dias in restricciones.items():
        for dia, bloques in dias.items():
            total_bloques_bloqueados += sum(bloques)
    
    print(f"🚫 Total de bloques bloqueados: {total_bloques_bloqueados}")
    
    # Guardar en JSON
    output_asignaciones = {
        'asignaciones': asignaciones,
        'total': len(asignaciones)
    }
    
    output_restricciones = {
        'restricciones': {
            str(prof_id): {
                dia: [int(b) for b in bloques]
                for dia, bloques in dias.items()
            }
            for prof_id, dias in restricciones.items()
        },
        'total_profesores': len(restricciones),
        'total_bloques_bloqueados': total_bloques_bloqueados
    }
    
    with open('asignaciones_actuales.json', 'w', encoding='utf-8') as f:
        json.dump(output_asignaciones, f, indent=2, ensure_ascii=False)
    
    with open('restricciones_profesores.json', 'w', encoding='utf-8') as f:
        json.dump(output_restricciones, f, indent=2, ensure_ascii=False)
    
    print(f"\n💾 Archivos guardados:")
    print(f"  - asignaciones_actuales.json")
    print(f"  - restricciones_profesores.json")
    print("="*80)
    
    return asignaciones, restricciones


def insertar_restricciones_en_bd(restricciones):
    """Inserta las restricciones en la tabla professor_restrictions"""
    
    print("\n" + "="*80)
    print("💾 INSERTANDO RESTRICCIONES EN BASE DE DATOS")
    print("="*80)
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Limpiar restricciones existentes
    cursor.execute("DELETE FROM professor_restrictions")
    print("🗑️  Restricciones anteriores eliminadas")
    
    # Mapeo de días en español a la tabla (estructura: day, start_time, end_time)
    dia_tabla = {
        'LUNES': 'LUNES',
        'MARTES': 'MARTES',
        'MIERCOLES': 'MIÉRCOLES',  # Con acento en BD
        'JUEVES': 'JUEVES',
        'VIERNES': 'VIERNES',
        'SABADO': 'SÁBADO'  # Con acento en BD
    }
    
    total_insertadas = 0
    
    for profesor_id, dias in restricciones.items():
        for dia_key, bloques in dias.items():
            for bloque_idx, bloqueado in enumerate(bloques):
                if bloqueado:  # Solo insertar si está bloqueado
                    inicio, fin = BLOQUES_TIEMPO[bloque_idx]
                    dia_bd = dia_tabla.get(dia_key, dia_key)
                    
                    # Insertar restricción (bloques bloqueados = no disponible)
                    cursor.execute("""
                        INSERT INTO professor_restrictions 
                        (professor_id, day, start_time, end_time, duration_blocks, reason)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        profesor_id,
                        dia_bd,
                        inicio,
                        fin,
                        1,  # 1 bloque de 50 min
                        'No disponible según horario actual'
                    ))
                    total_insertadas += 1
    
    conn.commit()
    cursor.close()
    conn.close()
    
    print(f"✅ {total_insertadas} restricciones (bloques NO disponibles) insertadas en la BD")
    print("="*80)


if __name__ == "__main__":
    asignaciones, restricciones = extraer_asignaciones_y_restricciones()
    
    if restricciones:
        insertar_restricciones_en_bd(restricciones)
