"""
EXTRACTOR V4 - DEFINITIVO CON COLOR + PROFESOR COMO IDENTIFICADOR
==================================================================

MEJORAS RESPECTO A V3:
- Usa (color + profesor) como identificador compuesto
- Incorpora mapeo manual de abreviaturas
- Mejor matching de nombres de cursos
- Verifica que cada curso tenga al menos un profesor asignado
"""

import mysql.connector
from openpyxl import load_workbook
from collections import defaultdict
import json
import re
from datetime import time

# ============================================================================
# CONFIGURACIÓN Y CONSTANTES
# ============================================================================

EXCEL_PATH = r'..\inputs\Horario_Docentes(2025-20).xlsx'
COLORES_IGNORAR = ['00000000', 'FFF2F2F2', 'FFFFFFFF', 'FFFF0000']

# MAPEO EXACTO DE HOJAS A PROFESORES (EXTRAÍDO DEL EXCEL REAL)
MAPEO_HOJAS_PROFESORES = {
    'A. Caballero': 'CABALLERO ALVARADO, ARMANDO',
    'C.Cuba': 'CAROLA LIZETH CUBA CASTILLO', 
    'C.Gay': 'Carlos Gaytan Toledo',
    'C. Guijon': 'Carlos Guijon Guerra',
    'C. Julca': 'Carlos Edwin Julca Castillo',
    'C.Mend': 'MENDOZA CORPUS CARLOS',
    'E.Cieza': 'CIEZA MOSTACERO SEGUNDO EDWIN',
    'E. Chav': 'Edilberto Chavez Fernandez',
    'E.SantaC': 'SANTA CRUZ, ELIAS',
    'Espinola': 'Espinola',
    'E.Mir': 'Eddy Miranda Velasquez',
    'F.Inf': 'Freddy Infantes Quiroz',
    'F.Per': 'Fernando Perez Cueva',
    'F.Cas': 'Fernando Castillo Robles',
    'H.Aba': 'Heber Abanto Cabrera',
    'H. Mendoza': 'Henry Mendoza Puerta',
    'H.Sag': 'Hernan Sagastegui Chigne',
    'J. Baylon': 'BAYLÓN CARRANZA JORGE RAMÓN',
    'J.Cal': 'Jose Calderon Sedano',
    'J.Cast': 'Jose Castañeda Saldaña',
    'J.Dia': 'Jaime Diaz Sanchez',
    'J. Gutierrez': 'GUTIERREZ GUTIERREZ JORGE LUIS',
    'J.Hua': 'Jorge Huapaya Escobedo',
    'J.Jar': 'Jorge Jara Arenas',
    'J.Pim': 'Jorge Piminchumo Flores',
    'J.Vasquez': 'VASQUEZ PEREYRA, JOSE',
    'K.Mel': 'Karla Melendez Revilla',
    'L.Vla': 'Luis Vladimir Urrelo',
    'L.Llanos': 'Lenin Llanos Leon',
    'M. Llerena': 'LLERENA FERNANDEZ, MONICA',
    'Moises': 'PEREZ CHAVEZ MOISES',
    'STAFF': 'CONVOCATORIA',
    'S.Rodri': 'Silvia Rodriguez Aguirre',
    'Sheyli': 'VALVERDE VELA SHEYLI',
    'W.Cue': 'Walter Cueva Chavez',
    'W.Lazo': 'Walter Lazo',
    'W.Letur': 'Walter Leturia',
    'Z.Vidal': 'Zoraida Vidal Melgarejo'
}

DIAS_MAPEO = {
    'LUNES': 'LUNES',
    'MARTES': 'MARTES',
    'MIÉRCOLES': 'MIERCOLES',
    'MIERCOLES': 'MIERCOLES',
    'JUEVES': 'JUEVES',
    'VIERNES': 'VIERNES',
    'SÁBADO': 'SABADO',
    'SABADO': 'SABADO'
}

BLOQUES_HORARIOS = [
    ('07:00:00', '07:50:00'), ('07:55:00', '08:45:00'),
    ('08:50:00', '09:40:00'), ('09:45:00', '10:35:00'),
    ('10:40:00', '11:30:00'), ('11:35:00', '12:25:00'),
    ('12:30:00', '13:20:00'), ('13:25:00', '14:15:00'),
    ('14:20:00', '15:10:00'), ('15:15:00', '16:05:00'),
    ('16:10:00', '17:00:00'), ('17:05:00', '17:55:00'),
    ('18:00:00', '18:50:00'), ('18:55:00', '19:45:00'),
    ('19:50:00', '20:40:00'), ('20:45:00', '21:35:00'),
    ('21:40:00', '22:30:00')
]

# ============================================================================
# FUNCIONES DE CONEXIÓN Y CARGA
# ============================================================================

def conectar_db():
    """Conecta a la base de datos"""
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='sistemas',
        database='upao_timetabling'
    )

def cargar_profesores():
    """Carga todos los profesores de la BD"""
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT id, nombre_completo
        FROM professors
        WHERE nombre_completo IS NOT NULL AND nombre_completo != ''
    """)
    profesores = cursor.fetchall()
    cursor.close()
    conn.close()
    print(f"[OK] Profesores cargados: {len(profesores)}")
    return profesores

def cargar_cursos():
    """Carga todos los cursos de la BD"""
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, codigo, nombre FROM courses")
    cursos = cursor.fetchall()
    cursor.close()
    conn.close()
    print(f"[OK] Cursos cargados: {len(cursos)}")
    return cursos

def cargar_mapeo_manual():
    """Carga el mapeo manual de abreviaturas"""
    try:
        with open('mapeo_manual_cursos.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            mapeo = data.get('mapeo_abreviaturas', {})
            print(f"[OK] Mapeo manual cargado: {len(mapeo)} entradas")
            return mapeo
    except FileNotFoundError:
        print("[!] No se encontró mapeo_manual_cursos.json, usando matching automático")
        return {}

# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE EXCEL
# ============================================================================

def obtener_color_celda(cell):
    """Extrae el color RGB de una celda"""
    if not cell.fill or not cell.fill.start_color:
        return '00000000'
    if hasattr(cell.fill.start_color, 'rgb'):
        return str(cell.fill.start_color.rgb).upper()
    return '00000000'

def es_celda_bloqueada(cell):
    """Verifica si una celda está bloqueada (roja)"""
    color = obtener_color_celda(cell)
    return color == 'FFFF0000'

def normalizar_nombre_curso(nombre):
    """Normaliza un nombre de curso para comparación"""
    # Remover paréntesis con tipo de sesión
    nombre = re.sub(r'\([TP][12]\)', '', nombre)
    nombre = re.sub(r'\([LP][12]\)', '', nombre)
    # Remover códigos numéricos
    nombre = re.sub(r'\b\d{4,5}\b', '', nombre)
    # Remover guiones y texto después de guiones (nombres de profesores)
    nombre = re.split(r'\s*-\s*', nombre)[0]
    # Normalizar espacios
    nombre = ' '.join(nombre.split())
    # Remover tildes y convertir a mayúsculas
    nombre = nombre.upper()
    tildes = {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ñ': 'N'}
    for tilde, sin_tilde in tildes.items():
        nombre = nombre.replace(tilde, sin_tilde)
    return nombre.strip()

def extraer_nombre_profesor_de_celda(contenido):
    """Extrae el nombre del profesor de una celda"""
    # Buscar patrón "- NombreProfesor" al final
    match = re.search(r'-\s*([A-Za-zÁÉÍÓÚáéíóúñÑ\s\.]+)$', contenido)
    if match:
        nombre = match.group(1).strip()
        # Limpiar nombres comunes no profesores
        if nombre.upper() not in ['VIRTUAL', 'CRECE', 'CREAR', 'PREGRADO', 'PRESENCIAL']:
            return nombre.upper()
    return None

def encontrar_profesor_por_nombre_parcial(nombre_parcial, profesores):
    """Busca un profesor por nombre parcial (fuzzy matching)"""
    if not nombre_parcial:
        return None
    
    nombre_parcial = nombre_parcial.upper().strip()
    
    # Coincidencia exacta en nombre completo
    for prof in profesores:
        nombre_completo = prof.get('nombre_completo')
        if nombre_completo and nombre_parcial in nombre_completo.upper():
            return prof
    
    # Coincidencia por palabras individuales
    palabras = nombre_parcial.split()
    for prof in profesores:
        nombre_completo = prof.get('nombre_completo')
        if nombre_completo:
            nombre_prof = nombre_completo.upper()
            if any(palabra in nombre_prof for palabra in palabras if len(palabra) > 2):
                return prof
    
    return None

def encontrar_curso_por_nombre(nombre_curso, cursos_bd, mapeo_manual):
    """
    Encuentra un curso en la BD usando:
    1. Mapeo manual
    2. Coincidencia por palabras clave
    """
    nombre_normalizado = normalizar_nombre_curso(nombre_curso)
    
    # 1. Intentar mapeo manual primero
    if nombre_normalizado in mapeo_manual:
        nombre_mapeado = mapeo_manual[nombre_normalizado]
        # Buscar el curso mapeado en BD
        for curso in cursos_bd:
            if nombre_mapeado.upper() in curso['nombre'].upper():
                return curso
    
    # 2. Búsqueda por coincidencia de palabras
    palabras_busqueda = set(nombre_normalizado.split())
    # Filtrar palabras muy cortas o comunes
    palabras_busqueda = {p for p in palabras_busqueda if len(p) > 2 and p not in ['DE', 'LA', 'DEL', 'LAS', 'LOS', 'PARA', 'CON']}
    
    if not palabras_busqueda:
        return None
    
    mejores_matches = []
    for curso in cursos_bd:
        nombre_curso_bd = curso['nombre'].upper()
        palabras_curso = set(nombre_curso_bd.split())
        
        # Contar coincidencias
        coincidencias = len(palabras_busqueda & palabras_curso)
        
        if coincidencias >= 2:  # Al menos 2 palabras coinciden
            mejores_matches.append((curso, coincidencias, len(palabras_curso)))
    
    if mejores_matches:
        # Ordenar por: más coincidencias, menos palabras totales (más específico)
        mejores_matches.sort(key=lambda x: (x[1], -x[2]), reverse=True)
        return mejores_matches[0][0]
    
    return None

def extraer_tipo_liga(contenido):
    """Extrae el tipo de sesión (T/P/L) y la liga (1/2) del contenido"""
    # Buscar patrones como (T1), (P2), (L1)
    match = re.search(r'\(([TPL])([12])\)', contenido, re.IGNORECASE)
    if match:
        tipo = match.group(1).upper()
        liga = int(match.group(2))
        return tipo, liga
    return None, None


def construir_mapa_celdas_unidas(ws):
    """Construye un mapa para identificar rangos de celdas combinadas"""
    merged_lookup = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_lookup[(row, col)] = {
                    'min_row': min_row,
                    'max_row': max_row,
                    'min_col': min_col,
                    'max_col': max_col,
                }
    return merged_lookup

# ============================================================================
# FUNCIÓN PRINCIPAL DE EXTRACCIÓN
# ============================================================================

def extraer_asignaciones_v4():
    """
    Extrae asignaciones usando (color + profesor) como identificador compuesto
    """
    print("\n" + "=" * 80)
    print("EXTRACTOR V4 - COLOR + PROFESOR COMO IDENTIFICADOR")
    print("=" * 80)
    
    # Cargar datos
    print(f"\nCargando Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=False)
    
    profesores_bd = cargar_profesores()
    cursos_bd = cargar_cursos()
    mapeo_manual = cargar_mapeo_manual()
    
    # Estructuras de datos
    asignaciones = []
    restricciones = defaultdict(lambda: defaultdict(lambda: [False] * 17))
    mapeo_profesores = {}
    cursos_asignados = defaultdict(set)  # curso_id -> set de professor_id
    
    estadisticas = {
        'hojas_procesadas': 0,
        'profesores_identificados': 0,
        'cursos_identificados': 0,
        'asignaciones_creadas': 0,
        'restricciones_encontradas': 0
    }
    
    # Procesar cada hoja
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"HOJA: {sheet_name}")
        
        # Buscar fila de encabezado
        fila_encabezado = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and 'LUNES' in str(cell_value).upper():
                    fila_encabezado = row_idx
                    break
            if fila_encabezado:
                break
        
        if not fila_encabezado:
            print(f"  [!] No se encontró encabezado en {sheet_name}")
            continue
        
        # Mapear columnas a días
        columnas_dias = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(fila_encabezado, col).value
            if header:
                dia = str(header).strip().upper()
                if dia in DIAS_MAPEO:
                    columnas_dias[col] = DIAS_MAPEO[dia]
        
        if not columnas_dias:
            print(f"  [!] No se encontraron columnas de días")
            continue
        
        print(f"  Dias: {list(columnas_dias.values())}")

        merged_lookup = construir_mapa_celdas_unidas(ws)
        
        # Identificar profesor de esta hoja USANDO MAPEO MANUAL
        profesor_encontrado = None
        
        # 1. Primero intentar mapeo manual directo
        if sheet_name in MAPEO_HOJAS_PROFESORES:
            nombre_esperado = MAPEO_HOJAS_PROFESORES[sheet_name]
            for prof in profesores_bd:
                if prof.get('nombre_completo') == nombre_esperado:
                    profesor_encontrado = prof
                    print(f"  [MAPEO MANUAL] {sheet_name} -> {nombre_esperado}")
                    break
        
        # 2. Si no hay mapeo manual, usar el método original
        if not profesor_encontrado:
            for row_idx in range(fila_encabezado + 1, min(fila_encabezado + 20, ws.max_row + 1)):
                for col in columnas_dias.keys():
                    cell = ws.cell(row_idx, col)
                    if cell.value:
                        contenido = str(cell.value).strip()
                        nombre_prof = extraer_nombre_profesor_de_celda(contenido)
                        if nombre_prof:
                            profesor_encontrado = encontrar_profesor_por_nombre_parcial(nombre_prof, profesores_bd)
                            if profesor_encontrado:
                                break
                if profesor_encontrado:
                    break
        
        if not profesor_encontrado:
            print(f"  [!] No se pudo identificar profesor para {sheet_name}")
            continue
        
        print(f"  [PROFESOR] {profesor_encontrado['nombre_completo']} (ID: {profesor_encontrado['id']})")
        mapeo_profesores[sheet_name] = profesor_encontrado
        estadisticas['profesores_identificados'] += 1
        
        # Agrupar celdas por COLOR (dentro de este profesor)
        cursos_por_color = defaultdict(list)
        
        for row_idx in range(fila_encabezado + 1, min(fila_encabezado + 20, ws.max_row + 1)):
            bloque_idx = row_idx - fila_encabezado - 1
            
            for col, dia in columnas_dias.items():
                cell = ws.cell(row_idx, col)
                merge_info = merged_lookup.get((row_idx, col))
                is_merge_top_left = True
                if merge_info:
                    is_merge_top_left = row_idx == merge_info['min_row'] and col == merge_info['min_col']
                
                # Verificar si es celda bloqueada (roja)
                if es_celda_bloqueada(cell):
                    if not is_merge_top_left:
                        continue

                    target_rows = [row_idx]
                    target_cols = [col]

                    if merge_info:
                        target_rows = list(range(merge_info['min_row'], merge_info['max_row'] + 1))
                        target_cols = [
                            target_col
                            for target_col in range(merge_info['min_col'], merge_info['max_col'] + 1)
                            if target_col in columnas_dias
                        ] or [col]

                    for target_col in target_cols:
                        dia_target = columnas_dias.get(target_col)
                        if not dia_target:
                            continue
                        for target_row in target_rows:
                            bloque_target = target_row - fila_encabezado - 1
                            if not 0 <= bloque_target < len(BLOQUES_HORARIOS):
                                continue
                            if not restricciones[profesor_encontrado['id']][dia_target][bloque_target]:
                                restricciones[profesor_encontrado['id']][dia_target][bloque_target] = True
                                estadisticas['restricciones_encontradas'] += 1
                    continue
                
                if cell.value:
                    contenido = str(cell.value).strip()
                    if len(contenido) > 3:
                        color = obtener_color_celda(cell)
                        
                        if color not in COLORES_IGNORAR:
                            cursos_por_color[color].append({
                                'contenido': contenido,
                                'dia': dia,
                                'bloque': bloque_idx,
                                'row': row_idx,
                                'col': col
                            })
        
        # Procesar cada color (cada color = un curso para este profesor)
        for color, celdas in cursos_por_color.items():
            if not celdas:
                continue
            
            # Extraer nombre del curso de la primera celda
            nombre_curso_raw = celdas[0]['contenido']
            curso = encontrar_curso_por_nombre(nombre_curso_raw, cursos_bd, mapeo_manual)
            
            if not curso:
                print(f"  [!] No se encontró curso para: {nombre_curso_raw[:50]}")
                continue
            
            print(f"  [CURSO] {curso['codigo']} - {curso['nombre']}")
            cursos_asignados[curso['id']].add(profesor_encontrado['id'])
            
            # Crear asignación para cada celda de este color
            for celda in celdas:
                tipo, liga = extraer_tipo_liga(celda['contenido'])
                
                if not tipo or not liga:
                    continue
                
                hora_inicio, hora_fin = BLOQUES_HORARIOS[celda['bloque']]
                
                asignacion = {
                    'profesor_id': profesor_encontrado['id'],
                    'profesor_nombre': profesor_encontrado['nombre_completo'],
                    'curso_id': curso['id'],
                    'curso_nombre': curso['nombre'],
                    'curso_codigo': curso['codigo'],
                    'tipo': tipo,
                    'liga': liga,
                    'session_type': f"{tipo}{liga}",
                    'dia': celda['dia'],
                    'bloque_inicio': celda['bloque'],
                    'hora_inicio': hora_inicio,
                    'hora_fin': hora_fin,
                    'contenido_original': celda['contenido'],
                    'color': color,
                    'hoja': sheet_name
                }
                
                asignaciones.append(asignacion)
                estadisticas['asignaciones_creadas'] += 1
                print(f"    -> {tipo}{liga} - {celda['dia']} bloque {celda['bloque']+1}")
        
        estadisticas['hojas_procesadas'] += 1
        estadisticas['cursos_identificados'] = len(set(a['curso_id'] for a in asignaciones))
    
    # Crear lista de restricciones
    lista_restricciones = []
    for prof_id, dias in restricciones.items():
        for dia, bloques in dias.items():
            idx = 0
            while idx < len(bloques):
                if not bloques[idx]:
                    idx += 1
                    continue

                start_idx = idx
                while idx + 1 < len(bloques) and bloques[idx + 1]:
                    idx += 1
                end_idx = idx

                hora_inicio = BLOQUES_HORARIOS[start_idx][0]
                hora_fin = BLOQUES_HORARIOS[end_idx][1]
                duration = end_idx - start_idx + 1

                lista_restricciones.append({
                    'professor_id': prof_id,
                    'day': dia,
                    'start_time': hora_inicio,
                    'end_time': hora_fin,
                    'duration_blocks': duration,
                    'reason': 'Extraido de Excel'
                })

                idx += 1
    
    # Guardar resultados
    with open('asignaciones_v4.json', 'w', encoding='utf-8') as f:
        json.dump({
            'asignaciones': asignaciones,
            'total': len(asignaciones)
        }, f, indent=2, ensure_ascii=False)
    
    with open('restricciones_v4.json', 'w', encoding='utf-8') as f:
        json.dump({
            'restricciones': lista_restricciones,
            'total': len(lista_restricciones)
        }, f, indent=2, ensure_ascii=False)
    
    with open('mapeo_profesores_v4.json', 'w', encoding='utf-8') as f:
        json.dump({hoja: {
            'id': prof['id'],
            'nombre': prof['nombre_completo']
        } for hoja, prof in mapeo_profesores.items()}, f, indent=2, ensure_ascii=False)
    
    # Reporte de cursos con/sin profesores
    with open('cobertura_cursos_v4.json', 'w', encoding='utf-8') as f:
        cursos_con_profesor = []
        cursos_sin_profesor = []
        
        for curso in cursos_bd:
            if curso['id'] in cursos_asignados:
                cursos_con_profesor.append({
                    'codigo': curso['codigo'],
                    'nombre': curso['nombre'],
                    'profesores': list(cursos_asignados[curso['id']])
                })
            else:
                cursos_sin_profesor.append({
                    'codigo': curso['codigo'],
                    'nombre': curso['nombre']
                })
        
        json.dump({
            'cursos_con_profesor': cursos_con_profesor,
            'total_con_profesor': len(cursos_con_profesor),
            'cursos_sin_profesor': cursos_sin_profesor,
            'total_sin_profesor': len(cursos_sin_profesor),
            'porcentaje_cobertura': len(cursos_con_profesor) * 100 / len(cursos_bd)
        }, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 80)
    print("RESUMEN FINAL:")
    print("=" * 80)
    for key, value in estadisticas.items():
        print(f"  {key}: {value}")
    print(f"  cursos_con_profesor: {len(cursos_asignados)}")
    print(f"  cursos_sin_profesor: {len(cursos_bd) - len(cursos_asignados)}")
    print(f"  cobertura: {len(cursos_asignados)*100/len(cursos_bd):.1f}%")
    
    return asignaciones, lista_restricciones, estadisticas

# ============================================================================
# INSERCIÓN EN BASE DE DATOS
# ============================================================================

def insertar_en_bd(asignaciones, restricciones):
    """Inserta asignaciones y restricciones en la BD"""
    print("\n" + "=" * 80)
    print("INSERTANDO EN BASE DE DATOS")
    print("=" * 80)
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    # 1. Limpiar tablas
    print("\n[1] Limpiando tablas existentes...")
    cursor.execute("DELETE FROM professor_restrictions WHERE reason = 'Extraido de Excel'")
    cursor.execute("DELETE FROM professor_course_history")
    conn.commit()
    
    # 2. Insertar restricciones
    print(f"\n[2] Insertando {len(restricciones)} restricciones...")
    for rest in restricciones:
        try:
            cursor.execute("""
                INSERT INTO professor_restrictions 
                (professor_id, day, start_time, end_time, duration_blocks, reason)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                rest['professor_id'],
                rest['day'],
                rest['start_time'],
                rest['end_time'],
                rest['duration_blocks'],
                rest['reason']
            ))
        except Exception as e:
            print(f"  [!] Error insertando restricción: {e}")
    conn.commit()
    print(f"  [OK] {len(restricciones)} restricciones insertadas")
    
    # 3. Insertar asignaciones históricas (solo pares únicos)
    print(f"\n[3] Procesando asignaciones históricas...")
    asignaciones_unicas = {}
    for asig in asignaciones:
        key = (asig['profesor_id'], asig['curso_id'])
        if key not in asignaciones_unicas:
            asignaciones_unicas[key] = asig
    
    print(f"  Total asignaciones: {len(asignaciones)}")
    print(f"  Pares únicos profesor-curso: {len(asignaciones_unicas)}")
    
    for asig in asignaciones_unicas.values():
        try:
            cursor.execute("""
                INSERT INTO professor_course_history 
                (professor_id, course_id, semestre, veces_asignado)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE veces_asignado = veces_asignado + 1
            """, (
                asig['profesor_id'],
                asig['curso_id'],
                '2025-20',
                1
            ))
        except Exception as e:
            print(f"  [!] Error insertando asignación: {e}")
    
    conn.commit()
    print(f"  [OK] {len(asignaciones_unicas)} asignaciones históricas insertadas")
    
    cursor.close()
    conn.close()
    
    print("\n" + "=" * 80)
    print("[OK] INSERCIÓN COMPLETADA")
    print("=" * 80)

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    asignaciones, restricciones, stats = extraer_asignaciones_v4()
    
    if asignaciones:
        insertar_en_bd(asignaciones, restricciones)
        print("\n[OK] Proceso completado exitosamente")
        print(f"[OK] Archivos generados:")
        print(f"     - asignaciones_v4.json")
        print(f"     - restricciones_v4.json")
        print(f"     - mapeo_profesores_v4.json")
        print(f"     - cobertura_cursos_v4.json")
    else:
        print("\n[!] No se extrajeron asignaciones")
