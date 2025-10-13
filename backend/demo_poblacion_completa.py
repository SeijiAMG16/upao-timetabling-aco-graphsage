"""
DEMO: WORKFLOW COMPLETO DESDE CERO
===================================

Este script demuestra cómo poblar la base de datos completamente desde cero
usando solo los Excel, de forma que todo esté listo para generación de horarios.

FLUJO PROPUESTO:
1. Extraer CURSOS del Excel de horarios (hojas de profesores)
2. Extraer PROFESORES del Excel de horarios (nombres de hojas)
3. Cargar cursos y profesores a BD
4. Subir Excel para asignaciones horarias
5. Sistema listo para ACO
"""

import pymysql
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime
import re

class CompleteDatabasePopulation:
    """Poblar BD completa desde Excel únicamente"""
    
    def __init__(self):
        self.connection = pymysql.connect(
            host='localhost',
            user='root',
            password='sistemas',
            database='upao_timetabling',
            charset='utf8mb4'
        )
        
    def extract_courses_from_excel(self, excel_path):
        """Extraer todos los cursos únicos del Excel"""
        wb = load_workbook(excel_path, data_only=True)
        cursos_encontrados = set()
        
        print("=== EXTRAYENDO CURSOS DEL EXCEL ===")
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Resumen', 'TOTAL', 'Summary']:
                continue
                
            ws = wb[sheet_name]
            print(f"Analizando hoja: {sheet_name}")
            
            # Buscar celdas con contenido que parezcan cursos
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip()
                        
                        # Patrones típicos de cursos
                        if (len(cell_text) > 5 and 
                            any(word in cell_text.upper() for word in 
                                ['ALGORITMIA', 'PROGRAMACION', 'MATEMATICA', 'CALCULO', 
                                 'FISICA', 'QUIMICA', 'SISTEMAS', 'INGENIERIA', 
                                 'ALGEBRA', 'COMUNICACION', 'INVESTIGACION'])):
                            
                            # Limpiar texto del curso
                            curso_limpio = re.sub(r'[()0-9]', '', cell_text).strip()
                            if len(curso_limpio) > 10:
                                cursos_encontrados.add(curso_limpio)
        
        print(f"Cursos únicos encontrados: {len(cursos_encontrados)}")
        for curso in sorted(cursos_encontrados):
            print(f"  - {curso}")
            
        return list(cursos_encontrados)
    
    def extract_professors_from_excel(self, excel_path):
        """Extraer profesores de nombres de hojas del Excel"""
        wb = load_workbook(excel_path, data_only=True)
        profesores = []
        
        print("\\n=== EXTRAYENDO PROFESORES DEL EXCEL ===")
        
        # Mapeo mejorado basado en el extractor V4
        mapeo_hojas = {
            'A. Caballero': 'CABALLERO ALVARADO, ARMANDO',
            'C.Cuba': 'CAROLA LIZETH CUBA CASTILLO', 
            'C.Gay': 'Carlos Gaytan Toledo',
            'C. Guijon': 'Carlos Guijon Guerra',
            'C. Julca': 'Carlos Edwin Julca Castillo',
            'C.Mend': 'MENDOZA CORPUS CARLOS',
            'E.Cieza': 'CIEZA MOSTACERO SEGUNDO EDWIN',
            'E. Chav': 'Edilberto Chavez Fernandez',
            'F.Inf': 'Freddy Infantes Quiroz',
            'J. Baylon': 'BAYLÓN CARRANZA JORGE RAMÓN',
            'J.Cal': 'Jose Calderon Sedano',
            'H.Aba': 'Heber Abanto Cabrera',
            'W.Lazo': 'William Lazo',
            'Moises': 'Moises Rodriguez'
        }
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Resumen', 'TOTAL', 'Summary']:
                continue
                
            if sheet_name in mapeo_hojas:
                nombre_completo = mapeo_hojas[sheet_name]
                profesores.append({
                    'hoja': sheet_name,
                    'nombre_completo': nombre_completo,
                    'codigo': f'PROF{len(profesores)+1:03d}'
                })
                print(f"  {sheet_name} -> {nombre_completo}")
            else:
                # Generar nombre genérico para hojas no mapeadas
                nombre_completo = sheet_name.replace('.', ' ').title()
                profesores.append({
                    'hoja': sheet_name,
                    'nombre_completo': nombre_completo,
                    'codigo': f'PROF{len(profesores)+1:03d}'
                })
                print(f"  {sheet_name} -> {nombre_completo} (generado)")
        
        print(f"Total profesores: {len(profesores)}")
        return profesores
    
    def populate_courses(self, cursos_list):
        """Poblar tabla de cursos con datos básicos"""
        print("\\n=== POBLANDO TABLA COURSES ===")
        
        cursor = self.connection.cursor()
        
        for i, curso_nombre in enumerate(cursos_list):
            # Generar código de curso
            codigo = f"CURSO-{i+1:03d}"
            
            # Datos básicos del curso
            ciclo = (i % 10) + 1  # Distribuir en ciclos 1-10
            
            # Determinar si requiere lab/práctica basado en nombre
            requiere_lab = any(word in curso_nombre.upper() for word in 
                             ['PROGRAMACION', 'LABORATORIO', 'FISICA', 'QUIMICA'])
            requiere_practica = True  # Asumir que todos tienen práctica
            
            grupos_teoria = 2
            grupos_practica = 2 if requiere_practica else 0
            grupos_laboratorio = 2 if requiere_lab else 0
            
            sql = """
            INSERT INTO courses (
                codigo, nombre, ciclo, creditos, modalidad,
                alumnos_teoria, alumnos_practica, alumnos_laboratorio,
                grupos_teoria, grupos_practica, grupos_laboratorio,
                requiere_laboratorio, requiere_practica,
                active, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            values = (
                codigo, curso_nombre, ciclo, 4, 'PRS',
                80, 40, 40,
                grupos_teoria, grupos_practica, grupos_laboratorio,
                requiere_lab, requiere_practica,
                True, datetime.now()
            )
            
            cursor.execute(sql, values)
            print(f"  Insertado: {codigo} - {curso_nombre}")
        
        self.connection.commit()
        print(f"✓ {len(cursos_list)} cursos insertados")
    
    def populate_professors(self, profesores_list):
        """Poblar tabla de profesores"""
        print("\\n=== POBLANDO TABLA PROFESSORS ===")
        
        cursor = self.connection.cursor()
        
        for profesor in profesores_list:
            # Separar nombres y apellidos
            nombre_completo = profesor['nombre_completo']
            parts = nombre_completo.split(' ')
            
            if len(parts) >= 2:
                nombres = ' '.join(parts[:len(parts)//2])
                apellidos = ' '.join(parts[len(parts)//2:])
            else:
                nombres = nombre_completo
                apellidos = ""
            
            sql = """
            INSERT INTO professors (
                codigo, nombre_completo, nombres, apellidos, email,
                categoria, carga_maxima_horas, active, created_at,
                disponible_lunes, disponible_martes, disponible_miercoles,
                disponible_jueves, disponible_viernes, disponible_sabado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            email = f"{profesor['codigo'].lower()}@upao.edu.pe"  # Usar código único
            
            values = (
                profesor['codigo'], nombre_completo, nombres, apellidos, email,
                'DOCENTE', 40, True, datetime.now(),
                True, True, True, True, True, True  # Disponible todos los días
            )
            
            cursor.execute(sql, values)
            print(f"  Insertado: {profesor['codigo']} - {nombre_completo}")
        
        self.connection.commit()
        print(f"✓ {len(profesores_list)} profesores insertados")
    
    def verify_population(self):
        """Verificar que los datos se insertaron correctamente"""
        print("\\n=== VERIFICANDO DATOS INSERTADOS ===")
        
        cursor = self.connection.cursor()
        
        # Verificar cursos
        cursor.execute("SELECT COUNT(*) FROM courses WHERE active = 1")
        total_courses = cursor.fetchone()[0]
        print(f"Cursos activos: {total_courses}")
        
        # Verificar profesores
        cursor.execute("SELECT COUNT(*) FROM professors WHERE active = 1")
        total_professors = cursor.fetchone()[0]
        print(f"Profesores activos: {total_professors}")
        
        # Mostrar ejemplos
        cursor.execute("SELECT codigo, nombre, grupos_teoria, grupos_practica, grupos_laboratorio FROM courses LIMIT 3")
        print("\\nEjemplos de cursos:")
        for row in cursor.fetchall():
            print(f"  {row[0]}: {row[1]} - T:{row[2]} P:{row[3]} L:{row[4]}")
        
        cursor.execute("SELECT codigo, nombre_completo FROM professors LIMIT 3")
        print("\\nEjemplos de profesores:")
        for row in cursor.fetchall():
            print(f"  {row[0]}: {row[1]}")
            
        return total_courses, total_professors
    
    def run_complete_population(self, excel_path):
        """Ejecutar población completa desde Excel"""
        print("=" * 60)
        print("DEMO: POBLACIÓN COMPLETA DE BD DESDE EXCEL")
        print("=" * 60)
        
        try:
            # 1. Extraer cursos del Excel
            cursos = self.extract_courses_from_excel(excel_path)
            
            # 2. Extraer profesores del Excel
            profesores = self.extract_professors_from_excel(excel_path)
            
            # 3. Poblar cursos
            self.populate_courses(cursos)
            
            # 4. Poblar profesores
            self.populate_professors(profesores)
            
            # 5. Verificar
            total_courses, total_professors = self.verify_population()
            
            print("\\n" + "=" * 60)
            print("✅ POBLACIÓN COMPLETA EXITOSA")
            print(f"📚 {total_courses} cursos creados")
            print(f"👥 {total_professors} profesores creados")
            print("🔄 Ahora el Excel upload debería funcionar correctamente")
            print("=" * 60)
            
            return True
            
        except Exception as e:
            print(f"❌ Error durante población: {e}")
            return False
        finally:
            self.connection.close()

# EJECUTAR DEMO
if __name__ == "__main__":
    excel_path = r"../inputs/Horario_Docentes(2025-20).xlsx"
    
    demo = CompleteDatabasePopulation()
    success = demo.run_complete_population(excel_path)
    
    if success:
        print("\\n🎯 SIGUIENTE PASO: Probar upload de Excel en frontend")
        print("   El sistema ahora debería procesar correctamente el Excel")
    else:
        print("\\n❌ FALLO EN LA DEMO")