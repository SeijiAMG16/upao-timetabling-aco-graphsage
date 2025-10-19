"""
================================================================================
GENERADOR DE HORARIOS POR PROFESOR - FORMATO EXCEL
================================================================================
Lee un experimento ACO con ligas y genera un archivo Excel con:
- Una hoja por profesor
- Formato limpio tipo calendario semanal
- Colores por tipo de sesión (Teoría/Práctica/Laboratorio)
================================================================================
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import mysql.connector
from datetime import datetime, time, timedelta

# Configuración de conexión a BD
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'sistemas',
    'database': 'upao_timetabling'
}

# Colores para tipos de sesión
COLORES = {
    'T': {'fill': 'E3F2FD', 'font': '0D47A1'},  # Azul claro - Teoría
    'P': {'fill': 'F3E5F5', 'font': '6A1B9A'},  # Morado claro - Práctica
    'L': {'fill': 'E8F5E9', 'font': '2E7D32'}   # Verde claro - Laboratorio
}

DIAS_SEMANA = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']

# Bloques de tiempo UPAO (50 minutos cada uno)
BLOQUES_TIEMPO = [
    ("07:00", "07:50"),
    ("07:55", "08:45"),
    ("08:50", "09:40"),
    ("09:45", "10:35"),
    ("10:40", "11:30"),
    ("11:35", "12:25"),
    ("12:30", "13:20"),
    ("13:25", "14:15"),
    ("14:20", "15:10"),
    ("15:15", "16:05"),
    ("16:10", "17:00"),
    ("17:05", "17:55"),
    ("18:00", "18:50"),
    ("18:55", "19:45"),
    ("19:50", "20:40"),
    ("20:45", "21:35"),
    ("21:40", "22:30")
]


def _obtener_columnas(tabla):
    """Obtiene set de columnas disponibles en una tabla."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute(f"SHOW COLUMNS FROM {tabla}")
    columnas = {fila['Field'] for fila in cursor.fetchall()}
    cursor.close()
    conn.close()
    return columnas


def cargar_datos_profesores():
    """Carga información de profesores desde la BD adaptándose al esquema actual."""
    columnas = _obtener_columnas('professors')
    campos = ['id']

    if 'nombre_completo' in columnas:
        campos.append('nombre_completo')
    if 'nombres' in columnas:
        campos.append('nombres')
    if 'apellidos' in columnas:
        campos.append('apellidos')
    if 'especialidad' in columnas:
        campos.append('especialidad')

    orden = []
    if 'apellidos' in columnas:
        orden.append('apellidos')
    if 'nombres' in columnas:
        orden.append('nombres')
    if not orden and 'nombre_completo' in columnas:
        orden.append('nombre_completo')

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    query = f"SELECT {', '.join(campos)} FROM professors"
    if orden:
        query += " ORDER BY " + ", ".join(orden)
    cursor.execute(query)

    profesores = {}
    for fila in cursor.fetchall():
        nombre_completo = fila.get('nombre_completo') or ''
        nombres = fila.get('nombres') or (nombre_completo.split()[0] if nombre_completo else 'N/A')
        apellidos = fila.get('apellidos') or (nombre_completo.split()[-1] if nombre_completo else 'N/A')

        profesores[fila['id']] = {
            'id': fila['id'],
            'nombre': nombres,
            'apellido': apellidos,
            'nombre_completo': nombre_completo or f"{nombres} {apellidos}",
            'especialidad': fila.get('especialidad')
        }

    cursor.close()
    conn.close()

    return profesores


def cargar_datos_cursos():
    """Carga información de cursos desde la BD"""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT id, nombre, codigo, ciclo
        FROM courses
        ORDER BY nombre
    """)
    
    cursos = {c['id']: c for c in cursor.fetchall()}
    cursor.close()
    conn.close()
    
    return cursos


def cargar_datos_aulas():
    """Carga información de aulas desde la BD"""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT id, codigo, edificio, piso, capacidad, tipo
        FROM classrooms
        ORDER BY codigo
    """)
    
    aulas = {a['id']: a for a in cursor.fetchall()}
    cursor.close()
    conn.close()
    
    return aulas


def parsear_hora(hora_str):
    """Convierte string de hora a objeto time"""
    if isinstance(hora_str, time):
        return hora_str
    if isinstance(hora_str, timedelta):
        total_segundos = int(hora_str.total_seconds())
        horas = (total_segundos // 3600) % 24
        minutos = (total_segundos % 3600) // 60
        segundos = total_segundos % 60
        return time(hour=horas, minute=minutos, second=segundos)
    return datetime.strptime(hora_str, "%H:%M:%S").time()


def obtener_bloque_inicio(hora_inicio):
    """Encuentra el índice del bloque de tiempo que contiene la hora de inicio"""
    hora_obj = parsear_hora(hora_inicio)
    
    for i, (inicio, fin) in enumerate(BLOQUES_TIEMPO):
        bloque_inicio = datetime.strptime(inicio, "%H:%M").time()
        bloque_fin = datetime.strptime(fin, "%H:%M").time()
        
        if bloque_inicio <= hora_obj <= bloque_fin:
            return i
    
    return 0


def calcular_duracion_bloques(hora_inicio, hora_fin):
    """Calcula cuántos bloques de 50 min ocupa la sesión"""
    inicio_obj = parsear_hora(hora_inicio)
    fin_obj = parsear_hora(hora_fin)
    
    duracion_minutos = (datetime.combine(datetime.today(), fin_obj) - 
                       datetime.combine(datetime.today(), inicio_obj)).total_seconds() / 60
    
    # Cada sesión de 2 horas ocupa aproximadamente 2-3 bloques
    return max(2, int(duracion_minutos / 50))


def agrupar_por_profesor(asignaciones):
    """Agrupa las asignaciones por profesor"""
    por_profesor = {}
    
    for asig in asignaciones:
        prof_id = asig['professor_id']
        if prof_id not in por_profesor:
            por_profesor[prof_id] = []
        por_profesor[prof_id].append(asig)
    
    return por_profesor


def cargar_asignaciones_desde_db(experimento_id):
    """Obtiene asignaciones desde proposed_schedule_assignments."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            course_id,
            professor_id,
            classroom_id,
            day,
            start_time,
            end_time,
            session_type
        FROM proposed_schedule_assignments
        WHERE algorithm_execution_id = %s
    """, (experimento_id,))

    asignaciones = cursor.fetchall()
    cursor.close()
    conn.close()

    for asig in asignaciones:
        inicio = asig.get('start_time')
        fin = asig.get('end_time')
        if isinstance(inicio, timedelta):
            total_segundos = int(inicio.total_seconds())
            horas = (total_segundos // 3600) % 24
            minutos = (total_segundos % 3600) // 60
            segundos = total_segundos % 60
            inicio = time(hour=horas, minute=minutos, second=segundos)
        if isinstance(inicio, time):
            asig['start_time'] = inicio.strftime('%H:%M:%S')
        elif isinstance(inicio, timedelta):
            asig['start_time'] = str(parsear_hora(inicio))
        
        if isinstance(fin, time):
            asig['end_time'] = fin.strftime('%H:%M:%S')
        elif isinstance(fin, timedelta):
            total_segundos = int(fin.total_seconds())
            horas = (total_segundos // 3600) % 24
            minutos = (total_segundos % 3600) // 60
            segundos = total_segundos % 60
            asig['end_time'] = time(hour=horas, minute=minutos, second=segundos).strftime('%H:%M:%S')
        if asig.get('day'):
            asig['day'] = asig['day'].upper()

    return asignaciones


def cargar_asignaciones_experimento(experimento_id):
    """Carga asignaciones desde JSON si existe, de lo contrario desde la BD."""
    archivo_json = Path(f"experimento_{experimento_id}_ligas.json")

    if archivo_json.exists():
        with open(archivo_json, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        asignaciones = datos.get('asignaciones', [])
        return asignaciones, f"archivo {archivo_json.name}"

    asignaciones_db = cargar_asignaciones_desde_db(experimento_id)
    return asignaciones_db, "tabla proposed_schedule_assignments"


def crear_hoja_profesor(wb, profesor_info, asignaciones, cursos, aulas):
    """Crea una hoja en Excel con el horario del profesor"""
    
    # Nombre seguro para la hoja (max 31 caracteres)
    nombre_hoja = f"{profesor_info['apellido'][:20]}"
    ws = wb.create_sheet(title=nombre_hoja)
    
    # Configurar anchos de columna
    ws.column_dimensions['A'].width = 15  # Columna de horarios
    for col in range(2, 8):  # Columnas de días
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    # ENCABEZADO - Información del profesor
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    nombre_completo = f"{profesor_info['nombre']} {profesor_info['apellido']}"
    cell.value = f"HORARIO - PROFESOR: {nombre_completo.upper()}"
    cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Especialidad
    if profesor_info.get('especialidad'):
        ws.merge_cells('A2:G2')
        cell = ws['A2']
        cell.value = f"Especialidad: {profesor_info['especialidad']}"
        cell.font = Font(name='Arial', size=10, italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    
    # FILA DE DÍAS DE LA SEMANA
    fila_dias = 3 if profesor_info.get('especialidad') else 2
    ws.merge_cells(f'A{fila_dias}:A{fila_dias+1}')
    cell = ws[f'A{fila_dias}']
    cell.value = 'HORARIO'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color='90CAF9', end_color='90CAF9', fill_type='solid')
    
    for i, dia in enumerate(DIAS_SEMANA, start=2):
        ws.merge_cells(f'{get_column_letter(i)}{fila_dias}:{get_column_letter(i)}{fila_dias+1}')
        cell = ws[f'{get_column_letter(i)}{fila_dias}']
        cell.value = dia
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='90CAF9', end_color='90CAF9', fill_type='solid')
    
    # Crear matriz de horario [hora][día]
    matriz_horario = {}
    for bloque_idx in range(len(BLOQUES_TIEMPO)):
        matriz_horario[bloque_idx] = {dia: None for dia in DIAS_SEMANA}
    
    # Llenar matriz con asignaciones
    for asig in asignaciones:
        dia = asig['day']
        bloque_inicio = obtener_bloque_inicio(asig['start_time'])
        duracion = calcular_duracion_bloques(asig['start_time'], asig['end_time'])
        
        curso = cursos.get(asig['course_id'], {})
        aula = aulas.get(asig['classroom_id'], {})
        
        # Extraer tipo de sesión (T/P/L) y liga
        session_type = asig['session_type']
        tipo = session_type[0] if session_type else 'T'
        
        info_sesion = {
            'curso': curso.get('nombre', 'Curso desconocido'),
            'codigo': curso.get('codigo', ''),
            'tipo': tipo,
            'session_type': session_type,
            'aula': aula.get('codigo', 'N/A'),
            'duracion': duracion,
            'inicio': asig['start_time'],
            'fin': asig['end_time']
        }
        
        # Colocar en la matriz
        if matriz_horario.get(bloque_inicio) and dia in matriz_horario[bloque_inicio]:
            matriz_horario[bloque_inicio][dia] = info_sesion
    
    # Generar filas de horario
    fila_actual = fila_dias + 2
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for bloque_idx, (inicio, fin) in enumerate(BLOQUES_TIEMPO):
        # Columna de hora
        cell = ws[f'A{fila_actual}']
        cell.value = f"{inicio}\n{fin}"
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        cell.border = thin_border
        ws.row_dimensions[fila_actual].height = 45
        
        # Columnas de días
        for col_idx, dia in enumerate(DIAS_SEMANA, start=2):
            cell = ws[f'{get_column_letter(col_idx)}{fila_actual}']
            info = matriz_horario[bloque_idx].get(dia)
            
            if info:
                # Formatear texto de la celda
                tipo_letra = {'T': 'TEORÍA', 'P': 'PRÁCTICA', 'L': 'LABORATORIO'}.get(info['tipo'], 'SESIÓN')
                
                cell.value = (f"{info['session_type']}\n"
                             f"{info['curso'][:30]}\n"
                             f"{tipo_letra}\n"
                             f"Aula: {info['aula']}")
                
                # Aplicar color según tipo
                color = COLORES.get(info['tipo'], COLORES['T'])
                cell.fill = PatternFill(start_color=color['fill'], 
                                       end_color=color['fill'], 
                                       fill_type='solid')
                cell.font = Font(name='Arial', size=9, color=color['font'], bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.value = ''
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            cell.border = thin_border
        
        fila_actual += 1
    
    # LEYENDA al final
    fila_leyenda = fila_actual + 1
    ws.merge_cells(f'A{fila_leyenda}:G{fila_leyenda}')
    cell = ws[f'A{fila_leyenda}']
    cell.value = "LEYENDA: 🔵 TEORÍA  |  🟣 PRÁCTICA  |  🟢 LABORATORIO"
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Total de horas
    total_horas = len(asignaciones) * 2  # Cada sesión son 2 horas
    fila_total = fila_leyenda + 1
    ws.merge_cells(f'A{fila_total}:G{fila_total}')
    cell = ws[f'A{fila_total}']
    cell.value = f"TOTAL DE SESIONES: {len(asignaciones)} | HORAS SEMANALES: {total_horas}"
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')


def generar_horarios_excel(experimento_id):
    """Genera archivo Excel con horarios por profesor"""
    
    print("="*80)
    print("📅 GENERADOR DE HORARIOS POR PROFESOR")
    print("="*80)
    
    asignaciones, fuente = cargar_asignaciones_experimento(experimento_id)

    if not asignaciones:
        print("❌ ERROR: No hay asignaciones disponibles para generar el horario")
        return

    print(f"\n📂 Asignaciones obtenidas desde {fuente}")
    print(f"✅ Asignaciones cargadas: {len(asignaciones)}")
    
    # Cargar datos de BD
    print("\n📊 Cargando datos desde la base de datos...")
    profesores = cargar_datos_profesores()
    cursos = cargar_datos_cursos()
    aulas = cargar_datos_aulas()
    
    print(f"✅ Profesores: {len(profesores)}")
    print(f"✅ Cursos: {len(cursos)}")
    print(f"✅ Aulas: {len(aulas)}")
    
    # Agrupar por profesor
    print("\n🔄 Agrupando asignaciones por profesor...")
    por_profesor = agrupar_por_profesor(asignaciones)
    
    print(f"✅ Profesores con asignaciones: {len(por_profesor)}")
    
    # Crear archivo Excel
    print("\n📝 Generando archivo Excel...")
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Crear hoja resumen
    ws_resumen = wb.create_sheet(title="RESUMEN", index=0)
    ws_resumen.column_dimensions['A'].width = 40
    ws_resumen.column_dimensions['B'].width = 15
    ws_resumen.column_dimensions['C'].width = 15
    
    # Encabezado resumen
    ws_resumen['A1'] = "RESUMEN DE HORARIOS - TODOS LOS PROFESORES"
    ws_resumen['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws_resumen['A1'].fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_resumen.merge_cells('A1:C1')
    
    ws_resumen['A2'] = "PROFESOR"
    ws_resumen['B2'] = "SESIONES"
    ws_resumen['C2'] = "HORAS/SEMANA"
    
    for cell in ['A2', 'B2', 'C2']:
        ws_resumen[cell].font = Font(name='Arial', size=11, bold=True)
        ws_resumen[cell].fill = PatternFill(start_color='90CAF9', end_color='90CAF9', fill_type='solid')
        ws_resumen[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    fila_resumen = 3
    
    # Crear hoja para cada profesor
    profesores_procesados = 0
    
    for prof_id in sorted(por_profesor.keys()):
        asigs_profesor = por_profesor[prof_id]
        profesor_info = profesores.get(prof_id)
        
        if not profesor_info:
            continue
        
        print(f"  → {profesor_info['nombre']} {profesor_info['apellido']}: {len(asigs_profesor)} sesiones")
        
        # Crear hoja del profesor
        crear_hoja_profesor(wb, profesor_info, asigs_profesor, cursos, aulas)
        
        # Agregar a resumen
        nombre_completo = f"{profesor_info['nombre']} {profesor_info['apellido']}"
        ws_resumen[f'A{fila_resumen}'] = nombre_completo
        ws_resumen[f'B{fila_resumen}'] = len(asigs_profesor)
        ws_resumen[f'C{fila_resumen}'] = len(asigs_profesor) * 2
        
        for cell in [f'A{fila_resumen}', f'B{fila_resumen}', f'C{fila_resumen}']:
            ws_resumen[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumen[cell].border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        fila_resumen += 1
        profesores_procesados += 1
    
    # Guardar archivo
    nombre_archivo = f"horarios_profesores_exp_{experimento_id}.xlsx"
    wb.save(nombre_archivo)
    
    print("\n" + "="*80)
    print(f"✅ ARCHIVO GENERADO EXITOSAMENTE")
    print("="*80)
    print(f"📁 Archivo: {nombre_archivo}")
    print(f"👥 Profesores: {profesores_procesados}")
    print(f"📊 Total de asignaciones: {len(asignaciones)}")
    print("="*80)
    print(f"\n💡 Abre el archivo para ver los horarios individuales de cada profesor")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        experimento_id = sys.argv[1]
    else:
        # Usar el experimento más reciente
        archivos = list(Path('.').glob('experimento_*_ligas.json'))
        if not archivos:
            print("❌ No se encontraron archivos de experimentos")
            sys.exit(1)
        
        # Ordenar por timestamp en el nombre
        archivos.sort(key=lambda x: int(x.stem.split('_')[1]), reverse=True)
        experimento_id = archivos[0].stem.split('_')[1]
        print(f"📌 Usando experimento más reciente: {experimento_id}")
    
    generar_horarios_excel(experimento_id)
