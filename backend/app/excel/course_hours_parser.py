"""Utilities to extract per-course session hours from Horario_Docentes Excel."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Dict, Iterable, List, Optional, Tuple

import json
import openpyxl
from openpyxl.utils.datetime import from_excel

from pathlib import Path

import re

DAY_HEADERS = {
    "LUNES": "Lunes",
    "MARTES": "Martes",
    "MIÉRCOLES": "Miercoles",
    "MIERCOLES": "Miercoles",
    "JUEVES": "Jueves",
    "VIERNES": "Viernes",
    "SÁBADO": "Sabado",
    "SABADO": "Sabado",
}

SESSION_PATTERN = re.compile(
    r"\s*(?P<name>.+?)\s*\((?P<session>[TPL])\s*(?P<league>\d+)\)\s*(?P<tail>.*)",
    re.IGNORECASE,
)
NRC_PATTERN = re.compile(r"\b(\d{4,6})\b")


@dataclass
class SessionAccumulator:
    total_minutes: int = 0
    entries: int = 0
    leagues: Dict[int, int] = field(default_factory=lambda: defaultdict(int))
    league_occurrences: Dict[int, int] = field(default_factory=lambda: defaultdict(int))
    nrcs: set[str] = field(default_factory=set)
    occurrences: List[Dict[str, object]] = field(default_factory=list)
    occurrence_minutes: List[int] = field(default_factory=list)

    def add(
        self,
        league: int,
        minutes: int,
        nrcs: Iterable[str],
        start_time: time,
        end_time: time,
    ) -> None:
        self.entries += 1
        self.total_minutes += minutes
        self.leagues[league] += minutes
        self.league_occurrences[league] += 1
        self.occurrence_minutes.append(minutes)
        for value in nrcs:
            if value:
                self.nrcs.add(value)
        self.occurrences.append(
            {
                "league": league,
                "start_time": _format_time(start_time),
                "end_time": _format_time(end_time),
                "duration_label": _format_minutes(minutes),
                "duration_hours": round(minutes / 60.0, 2),
            }
        )

    def to_dict(self) -> Dict[str, object]:
        per_occurrence_payload: Optional[Dict[str, object]] = None
        if self.occurrence_minutes:
            unique_minutes = sorted(set(self.occurrence_minutes))
            if len(unique_minutes) == 1:
                minutes_val = unique_minutes[0]
                per_occurrence_payload = {
                    "minutes": minutes_val,
                    "hours": round(minutes_val / 60.0, 2),
                    "duration_label": _format_minutes(minutes_val),
                }
        return {
            "total_hours": round(self.total_minutes / 60.0, 2),
            "total_duration_label": _format_minutes(self.total_minutes),
            "entries": self.entries,
            "per_occurrence": per_occurrence_payload,
            "leagues": {
                league: {
                    "hours": round(minutes / 60.0, 2),
                    "duration_label": _format_minutes(minutes),
                    "occurrences": self.league_occurrences[league],
                    "per_occurrence": _build_per_occurrence_minutes(minutes, self.league_occurrences[league]),
                }
                for league, minutes in sorted(self.leagues.items())
            },
            "nrcs": sorted(self.nrcs),
            "occurrences": self.occurrences,
        }


@dataclass
class ExtractionIssue:
    sheet: str
    row: int
    column: int
    message: str
    value: Optional[str] = None

    def to_dict(self) -> Dict[str, object]:
        payload = {
            "sheet": self.sheet,
            "row": self.row,
            "column": self.column,
            "message": self.message,
        }
        if self.value is not None:
            payload["value"] = self.value
        return payload


def _parse_time(value: object) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).time()
        except Exception:
            return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        raw_upper = raw.upper()
        is_pm = any(token in raw_upper for token in ("PM", "P.M."))
        is_am = any(token in raw_upper for token in ("AM", "A.M.")) and not is_pm
        cleaned = raw_upper
        for suffix in ("AM", "PM", " A.M.", " P.M."):
            cleaned = cleaned.replace(suffix, "")
        cleaned = cleaned.replace(".", ":").replace(" ", "")
        if len(cleaned) == 4 and cleaned.isdigit():
            cleaned = f"{cleaned[:2]}:{cleaned[2:]}"
        if ":" not in cleaned:
            return None
        parts = cleaned.split(":")
        try:
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0
        except ValueError:
            return None
        if is_pm and hour < 12:
            hour += 12
        if is_am and hour == 12:
            hour = 0
        return time(hour, minute)
    return None


def _format_time(value: time) -> str:
    return f"{value.hour}:{value.minute:02d}"


def _format_minutes(total_minutes: int) -> str:
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours}:{minutes:02d}"


def _duration_minutes(start: time, end: time) -> Optional[int]:
    if not start or not end:
        return None
    start_dt = datetime.combine(date.today(), start)
    end_dt = datetime.combine(date.today(), end)
    if end_dt <= start_dt:
        return None
    delta = end_dt - start_dt
    return int(delta.total_seconds() // 60)


def _add_minutes(start: time, minutes: int) -> time:
    start_dt = datetime.combine(date.today(), start)
    return (start_dt + timedelta(minutes=minutes)).time()


def _build_per_occurrence_minutes(total_minutes: int, occurrences: int) -> Optional[Dict[str, object]]:
    if occurrences <= 0:
        return None
    if total_minutes % occurrences != 0:
        return None
    minutes = total_minutes // occurrences
    return {
        "minutes": minutes,
        "hours": round(minutes / 60.0, 2),
        "duration_label": _format_minutes(minutes),
    }


def _minutes_from_label(label: str) -> Optional[int]:
    if not label:
        return None
    cleaned = label.strip()
    if not cleaned:
        return None
    parts = cleaned.split(":")
    if len(parts) != 2:
        return None
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
    except ValueError:
        return None
    return hours * 60 + minutes


def _build_manual_session(minutes: int) -> Dict[str, object]:
    hours = round(minutes / 60.0, 2)
    duration_label = _format_minutes(minutes)
    occurrence = {
        "league": None,
        "start_time": None,
        "end_time": None,
        "duration_label": duration_label,
        "duration_hours": hours,
        "source": "manual",
    }
    return {
        "total_hours": hours,
        "total_duration_label": duration_label,
        "entries": 1,
        "per_occurrence": {
            "minutes": minutes,
            "hours": hours,
            "duration_label": duration_label,
        },
        "leagues": {},
        "nrcs": [],
        "occurrences": [occurrence],
        "source": "manual",
    }


def _load_manual_overrides() -> Dict[str, Dict[str, str]]:
    overrides_path = Path(__file__).resolve().parent.parent.parent / "manual_course_hours_overrides.json"
    if not overrides_path.exists():
        return {}
    try:
        with overrides_path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        return {str(course): {str(k).upper(): str(v) for k, v in sessions.items()} for course, sessions in data.items()}
    except Exception:
        return {}


def _extract_entry(raw_text: str) -> Optional[Tuple[str, str, int, List[str]]]:
    if not raw_text:
        return None
    cleaned = " ".join(raw_text.split())
    match = SESSION_PATTERN.match(cleaned)
    if not match:
        return None
    course_name = match.group("name").strip(" -")
    session_type = match.group("session").upper()
    league = int(match.group("league"))
    tail = match.group("tail") or ""
    nrcs = set(NRC_PATTERN.findall(tail))
    return course_name, session_type, league, sorted(nrcs)


def extract_course_hours(file_path: str) -> Tuple[Dict[str, Dict[str, Dict[str, object]]], List[Dict[str, object]]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    grouped: Dict[str, Dict[str, SessionAccumulator]] = defaultdict(lambda: defaultdict(SessionAccumulator))
    issues: List[ExtractionIssue] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        header_row = None
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            for col_idx in range(1, min(12, sheet.max_column + 1)):
                cell_value = sheet.cell(row_idx, col_idx).value
                if isinstance(cell_value, str) and cell_value.strip().upper() in DAY_HEADERS:
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            issues.append(ExtractionIssue(sheet_name, 1, 1, "No se encontró fila de encabezados con días"))
            continue

        day_columns = {}
        start_col_idx: Optional[int] = None
        end_col_idx: Optional[int] = None
        for col_idx in range(1, sheet.max_column + 1):
            header_value = sheet.cell(header_row, col_idx).value
            if not isinstance(header_value, str):
                continue
            key = header_value.strip().upper()
            if key == "INICIO" and start_col_idx is None:
                start_col_idx = col_idx
            if key in {"TERMINO", "TÉRMINO"} and end_col_idx is None:
                end_col_idx = col_idx
            if key in DAY_HEADERS:
                day_columns[col_idx] = DAY_HEADERS[key]

        if not day_columns:
            issues.append(ExtractionIssue(sheet_name, header_row, 1, "No se identificaron columnas de días"))
            continue

        if start_col_idx is None or end_col_idx is None:
            # fallback a columnas estándar B y C
            start_col_idx = start_col_idx or 2
            end_col_idx = end_col_idx or 3

        merged_lookup: Dict[Tuple[int, int], object] = {}
        for merged_range in sheet.merged_cells.ranges:
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_lookup[(row_idx, col_idx)] = merged_range

        start_cache: Dict[int, Optional[time]] = {}
        end_cache: Dict[int, Optional[time]] = {}
        slot_cache: Dict[int, Optional[int]] = {}

        def _row_start(row: int) -> Optional[time]:
            if row not in start_cache:
                source_row = row
                merged = merged_lookup.get((row, start_col_idx))
                if merged is not None:
                    source_row = merged.min_row
                start_cache[row] = _parse_time(sheet.cell(source_row, start_col_idx).value)
            return start_cache[row]

        def _row_end(row: int) -> Optional[time]:
            if row not in end_cache:
                source_row = row
                merged = merged_lookup.get((row, end_col_idx))
                if merged is not None:
                    source_row = merged.min_row
                end_cache[row] = _parse_time(sheet.cell(source_row, end_col_idx).value)
            return end_cache[row]

        def _row_slot(row: int) -> Optional[int]:
            if row not in slot_cache:
                slot_cache[row] = _duration_minutes(_row_start(row), _row_end(row))
            return slot_cache[row]

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            for col_idx in day_columns.keys():
                cell = sheet.cell(row_idx, col_idx)
                if not cell.value:
                    continue

                merged_range = merged_lookup.get((row_idx, col_idx))
                if merged_range is not None:
                    if row_idx != merged_range.min_row or col_idx != merged_range.min_col:
                        continue
                    start_row = merged_range.min_row
                    end_row = merged_range.max_row
                else:
                    start_row = end_row = row_idx

                cell_text = str(cell.value)
                entry = _extract_entry(cell_text)
                if not entry:
                    if "(" in cell_text and ")" in cell_text:
                        issues.append(
                            ExtractionIssue(
                                sheet_name,
                                row_idx,
                                col_idx,
                                "Formato de celda no reconocido",
                                value=cell_text[:120],
                            )
                        )
                    continue

                start_time = _row_start(start_row)
                if start_time is None:
                    issues.append(
                        ExtractionIssue(
                            sheet_name,
                            start_row,
                            start_col_idx,
                            "No se pudo leer hora de inicio",
                        )
                    )
                    continue

                end_time = _row_end(end_row)
                duration_minutes = _duration_minutes(start_time, end_time)

                if (end_time is None or duration_minutes is None) and merged_range is not None:
                    span = merged_range.max_row - merged_range.min_row + 1
                    slot_minutes = _row_slot(start_row)
                    if slot_minutes:
                        duration_minutes = slot_minutes * span
                        end_time = _add_minutes(start_time, duration_minutes)

                if end_time is None or duration_minutes is None or duration_minutes <= 0:
                    issues.append(
                        ExtractionIssue(
                            sheet_name,
                            row_idx,
                            col_idx,
                            "No se pudo calcular duración a partir de las celdas combinadas",
                        )
                    )
                    continue

                course_name, session_type, league, nrcs = entry
                accumulator = grouped[course_name][session_type]
                accumulator.add(league, duration_minutes, nrcs, start_time, end_time)

    output: Dict[str, Dict[str, Dict[str, object]]] = {}
    for course_name, sessions in grouped.items():
        output[course_name] = {}
        for session_type, accumulator in sessions.items():
            output[course_name][session_type] = accumulator.to_dict()

    overrides = _load_manual_overrides()
    for course_name, sessions in overrides.items():
        course_bucket = output.setdefault(course_name, {})
        for session_type, duration_label in sessions.items():
            if session_type in course_bucket:
                continue
            minutes = _minutes_from_label(duration_label)
            if minutes is None:
                continue
            course_bucket[session_type] = _build_manual_session(minutes)

    return output, [issue.to_dict() for issue in issues]


if __name__ == "__main__":
    data, issues = extract_course_hours("../inputs/Horario_Docentes(2025-20).xlsx")
    print(f"Cursos procesados: {len(data)}")
    sample = sorted(data.items())[:10]
    for course, sessions in sample:
        print(f"\n{course}")
        for session_type, details in sessions.items():
            per_occ = details.get("per_occurrence")
            entries = details.get("entries", 0)
            if per_occ:
                summary = (
                    f"{details['total_duration_label']} total, {entries} ocurrencias, "
                    f"{per_occ['duration_label']} c/u"
                )
            else:
                summary = f"{details['total_duration_label']} total"
            print(f"  {session_type}: {summary}")
            for occurrence in details["occurrences"]:
                league_display = occurrence.get("league")
                if league_display is None:
                    league_display = "-"
                start = occurrence.get("start_time") or "-"
                end = occurrence.get("end_time") or "-"
                duration_label = occurrence.get("duration_label", "-")
                suffix = " (manual)" if occurrence.get("source") == "manual" else ""
                print(f"    L{league_display}: {start} - {end} -> {duration_label} horas{suffix}")
    if issues:
        print(f"\nIncidencias detectadas: {len(issues)}")
        for issue in issues[:10]:
            print(issue)