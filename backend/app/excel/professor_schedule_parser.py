"""
Parser para Horario_Docentes(2025-20).xlsx
Extrae profesores y sus restricciones horarias desde las hojas del Excel
"""
import openpyxl
import re
from typing import Dict, List, Any
from datetime import time


def parse_professor_schedules(file_path: str) -> Dict[str, Any]:
    """
    Parsea el Excel de horarios de docentes y extrae:
    - Lista de profesores encontrados
    - Restricciones horarias (bloques ocupados)
    
    Estructura del Excel:
    - Cada hoja representa un profesor (nombre en el tab)
    - Fila 2: Headers (INICIO, TERMINO, LUNES, MARTES, etc.)
    - Filas 3+: Bloques horarios con asignaciones
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        professors_data = []
        restrictions_data = []
        
        print(f"\n📊 Analizando Excel de Docentes: {file_path}")
        print(f"📄 Número de hojas (profesores): {len(wb.sheetnames)}")
        
        # Mapeo de días en español
        day_mapping = {
            'LUNES': 'Lunes',
            'MARTES': 'Martes',
            'MIÉRCOLES': 'Martes',  # Usar nombre consistente
            'MIERCOLES': 'Miercoles',
            'JUEVES': 'Jueves',
            'VIERNES': 'Viernes',
            'SÁBADO': 'Sabado',
            'SABADO': 'Sabado'
        }
        
        # Procesar cada hoja (cada hoja = un profesor)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Extraer nombre del profesor del nombre de la hoja
            professor_name = sheet_name.strip()
            
            # Generar código único (primeras letras + número)
            name_parts = professor_name.upper().replace('.', '').replace(',', '').split()
            codigo = ''.join([p[0] for p in name_parts[:2]]) if len(name_parts) >= 2 else name_parts[0][:3]
            codigo = f"PROF{codigo}{len(professors_data):03d}"
            
            print(f"\n👤 Profesor: {professor_name} ({codigo})")
            
            # Buscar fila de headers
            header_row = None
            for i in range(1, min(5, ws.max_row + 1)):
                row = [cell.value for cell in ws[i]]
                if any(d in str(cell).upper() for cell in row for d in ['LUNES', 'MARTES', 'INICIO']):
                    header_row = i
                    break
            
            if not header_row:
                print(f"  ⚠️ No se encontraron headers, saltando hoja")
                continue
            
            # Leer headers para identificar columnas de días
            headers = [cell.value for cell in ws[header_row]]
            day_columns = {}
            
            for idx, header in enumerate(headers):
                if header:
                    header_upper = str(header).upper().strip()
                    for key, day in day_mapping.items():
                        if key in header_upper:
                            day_columns[idx] = day
                            break
            
            print(f"  📅 Días detectados: {list(day_columns.values())}")
            
            # Procesar bloques horarios (desde fila después de headers)
            professor_restrictions = []
            
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row = list(ws[row_idx])
                
                # Extraer horarios (columnas 1 y 2: INICIO y TERMINO)
                inicio_cell = row[1].value if len(row) > 1 else None
                termino_cell = row[2].value if len(row) > 2 else None
                
                if not inicio_cell or not termino_cell:
                    continue
                
                # Convertir hora a formato time
                try:
                    inicio_str = str(inicio_cell).strip().replace('am', '').replace('pm', '').replace(' ', '')
                    termino_str = str(termino_cell).strip().replace('am', '').replace('pm', '').replace(' ', '')
                    
                    # Parsear hora (puede venir como "07:00am" o como time object)
                    if ':' in inicio_str:
                        hora_inicio = inicio_str.split(':')[0].zfill(2)
                        min_inicio = inicio_str.split(':')[1][:2].zfill(2)
                        start_time = f"{hora_inicio}:{min_inicio}:00"
                    else:
                        continue
                    
                    if ':' in termino_str:
                        hora_termino = termino_str.split(':')[0].zfill(2)
                        min_termino = termino_str.split(':')[1][:2].zfill(2)
                        end_time = f"{hora_termino}:{min_termino}:00"
                    else:
                        continue
                    
                except Exception as e:
                    continue
                
                # Revisar cada día de la semana
                for col_idx, day in day_columns.items():
                    if col_idx < len(row):
                        cell_value = row[col_idx].value
                        
                        # Si hay contenido en la celda, es una restricción (bloque ocupado)
                        if cell_value and str(cell_value).strip():
                            restriction = {
                                'professor_codigo': codigo,
                                'professor_name': professor_name,
                                'day': day,
                                'start_time': start_time,
                                'end_time': end_time,
                                'reason': f"Clase asignada: {str(cell_value)[:50]}",
                                'duration_blocks': 1
                            }
                            professor_restrictions.append(restriction)
            
            print(f"  🚫 Restricciones encontradas: {len(professor_restrictions)}")
            
            # Agregar profesor a la lista
            professors_data.append({
                'codigo': codigo,
                'nombres': professor_name.split()[0] if ' ' in professor_name else professor_name,
                'apellidos': ' '.join(professor_name.split()[1:]) if ' ' in professor_name else '',
                'nombre_completo': professor_name,
                'email': f"{codigo.lower()}@upao.edu.pe",
                'categoria': 'TC',  # Tiempo Completo por defecto
                'carga_maxima_horas': 40,
                'especialidad': 'Por definir',
                'restrictions_count': len(professor_restrictions)
            })
            
            restrictions_data.extend(professor_restrictions)
        
        wb.close()
        
        print(f"\n✅ Total profesores extraídos: {len(professors_data)}")
        print(f"✅ Total restricciones extraídas: {len(restrictions_data)}")
        
        return {
            'success': True,
            'professors': professors_data,
            'restrictions': restrictions_data,
            'total_professors': len(professors_data),
            'total_restrictions': len(restrictions_data)
        }
    
    except Exception as e:
        print(f"❌ Error al parsear Excel de profesores: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'professors': [],
            'restrictions': []
        }
