"""
Parser para procesar el archivo Libro1.xlsx (Proyecciones)
Extrae: cursos, alumnos proyectados por tipo de sesión
"""

import openpyxl
from typing import Dict, List, Any
import re

def parse_libro1_projections(file_path: str) -> Dict[str, Any]:
    """
    Procesa Libro1.xlsx y extrae proyecciones de cursos
    
    Returns:
        {
            'courses': [
                {
                    'codigo': 'ISII01',
                    'nombre': 'PROGRAMACION ORIENTADA A OBJETOS',
                    'ciclo': 4,
                    'alumnos_teoria': 120,
                    'alumnos_practica': 120,
                    'alumnos_laboratorio': 120,
                    'creditos': 4
                }
            ]
        }
    """
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        courses = []
        
        print(f"📊 Analizando Excel: {file_path}")
        print(f"📄 Sheet: {ws.title}, Dimensiones: {ws.max_row} filas")
        
        # ESTRUCTURA LIBRO1.xlsx:
        # Col A: N° | B: PROGRAMA | C: CICLO | D: COD | E: NUM | F: ASIGNATURA 
        # G: CRED | H: HT | I: HP | J: HL | N: PRESENCIAL/NO PRESENCIAL
        # O: N° alumnos Teoría | P: N° alumnos Práctica | Q: N° alumnos Laboratorio
        # R: N° Grupos Teoría | S: N° Grupos Práctica | T: N° Grupos Laboratorio
        
        # Procesar filas desde la 2
        for row_idx in range(2, ws.max_row + 1):
            row = list(ws[row_idx])
            
            # Extraer valores por índice (0-based)
            ciclo_str = row[2].value if len(row) > 2 else None  # Col C: CICLO
            codigo_prefix = row[3].value if len(row) > 3 else None  # Col D: COD (ej: CIEN, ICSI)
            codigo_num = row[4].value if len(row) > 4 else None  # Col E: NUM (ej: 752, 506)
            nombre = row[5].value if len(row) > 5 else None  # Col F: ASIGNATURA
            creditos = row[6].value if len(row) > 6 else 3  # Col G: CRED
            modalidad_str = row[13].value if len(row) > 13 else 'PRS'  # Col N: PRESENCIAL/NO PRESENCIAL
            
            # Columnas de alumnos (índices 14, 15, 16 = cols O, P, Q)
            alumnos_teoria = row[14].value if len(row) > 14 else 0  # Col O
            alumnos_practica = row[15].value if len(row) > 15 else 0  # Col P
            alumnos_laboratorio = row[16].value if len(row) > 16 else 0  # Col Q
            
            # ⚠️ COLUMNAS CRÍTICAS - N° de GRUPOS (índices 17, 18, 19 = cols R, S, T)
            grupos_teoria = row[17].value if len(row) > 17 else 0  # Col R
            grupos_practica = row[18].value if len(row) > 18 else 0  # Col S
            grupos_laboratorio = row[19].value if len(row) > 19 else 0  # Col T
            
            # Validar datos mínimos
            if not codigo_prefix or not codigo_num or not nombre:
                continue
            
            # Construir código completo (ej: CIEN752, ICSI506)
            codigo = f"{str(codigo_prefix).strip()}{str(codigo_num).strip()}"
            nombre = str(nombre).strip().upper()
            
            if len(codigo) < 3:
                continue
            
            # Procesar ciclo (puede ser "C1", "C2", etc.)
            ciclo = 1
            if ciclo_str:
                ciclo_match = re.search(r'\d+', str(ciclo_str))
                if ciclo_match:
                    ciclo = int(ciclo_match.group())
            
            # Convertir valores numéricos
            try:
                creditos = int(creditos) if creditos else 3
                alumnos_teoria = int(alumnos_teoria) if alumnos_teoria else 0
                alumnos_practica = int(alumnos_practica) if alumnos_practica else 0
                alumnos_laboratorio = int(alumnos_laboratorio) if alumnos_laboratorio else 0
                grupos_teoria = int(grupos_teoria) if grupos_teoria else 0
                grupos_practica = int(grupos_practica) if grupos_practica else 0
                grupos_laboratorio = int(grupos_laboratorio) if grupos_laboratorio else 0
            except (ValueError, TypeError):
                creditos = 3
                alumnos_teoria = 0
                alumnos_practica = 0
                alumnos_laboratorio = 0
                grupos_teoria = 0
                grupos_practica = 0
                grupos_laboratorio = 0
            
            # Determinar modalidad (PRS = Presencial, NPR = No Presencial)
            modalidad = 'PRESENCIAL'
            if modalidad_str and 'NPR' in str(modalidad_str).upper():
                modalidad = 'NO_PRESENCIAL'
            
            course_data = {
                'codigo': codigo,
                'nombre': nombre,
                'ciclo': ciclo,
                'modalidad': modalidad,
                'alumnos_teoria': alumnos_teoria,
                'alumnos_practica': alumnos_practica,
                'alumnos_laboratorio': alumnos_laboratorio,
                'grupos_teoria': grupos_teoria,
                'grupos_practica': grupos_practica,
                'grupos_laboratorio': grupos_laboratorio,
                'creditos': creditos,
                'requiere_laboratorio': alumnos_laboratorio > 0,
                'requiere_practica': alumnos_practica > 0,
            }
            
            courses.append(course_data)
            if len(courses) <= 5:  # Mostrar primeros 5 para debug
                print(f"✅ {codigo} - {nombre[:30]} | Grupos T:{grupos_teoria} P:{grupos_practica} L:{grupos_laboratorio} | Alumnos T:{alumnos_teoria} P:{alumnos_practica} L:{alumnos_laboratorio}")
        
        wb.close()
        
        print(f"🎉 Total extraído: {len(courses)} cursos")
        
        return {
            'success': True,
            'courses': courses,
            'total': len(courses)
        }
    
    except Exception as e:
        print(f"❌ Error al parsear Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'courses': []
        }
