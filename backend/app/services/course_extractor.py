"""
Servicio de extracción de cursos desde Excel
Integra el ExtractorCursosV2 al backend
"""
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
import math
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Course, CourseSection
from app.database import SessionLocal
from app.services.section_normalizer import NRCGenerator

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class CourseExtractorService:
    """
    Servicio para extraer cursos desde Excel
    Integra la funcionalidad de ExtractorCursosV2 al backend
    """
    
    def __init__(self):
        self.excel_path = None
        self.cursos_extraidos = []
        self.codigo_counts: Dict[str, int] = {}
        self.codigos_generados: set[str] = set()
        self._nrc_generator: Optional[NRCGenerator] = None
        
        # Mapeo de modalidades
        self.modalidad_mapping = {
            'PRS': 'presencial',
            'NPR': 'no_presencial'
        }
        
        # Mapeo de ciclos
        self.ciclo_mapping = {
            'C1': 1, 'C2': 2, 'C3': 3, 'C4': 4, 'C5': 5,
            'C6': 6, 'C7': 7, 'C8': 8, 'C9': 9, 'C10': 10
        }
    
    def extract_courses_from_excel(self, excel_path: str) -> Dict[str, Any]:
        """
        Extrae cursos desde archivo Excel
        """
        try:
            self.excel_path = excel_path
            self.cursos_extraidos = []
            self.codigo_counts = {}
            self.codigos_generados = set()
            
            logger.info(f"🔍 Iniciando extracción desde: {excel_path}")
            
            # Cargar workbook
            workbook = load_workbook(excel_path, data_only=True)
            
            # Procesar cada hoja
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet_name.startswith('C') and len(sheet_name) <= 3:
                    self._procesar_hoja_por_ciclo(sheet, sheet_name)
                else:
                    self._procesar_hoja_general(sheet, sheet_name)
            
            logger.info(f"✅ Extracción completada: {len(self.cursos_extraidos)} cursos")
            
            return {
                "success": True,
                "total_courses": len(self.cursos_extraidos),
                "courses": self.cursos_extraidos,
                "summary": self._generate_summary()
            }
            
        except Exception as e:
            logger.error(f"❌ Error en extracción: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "total_courses": 0,
                "courses": []
            }
    
    def _procesar_hoja_por_ciclo(self, sheet, ciclo_nombre: str):
        """Procesa una hoja de ciclo específica"""
        try:
            ciclo_numero = self.ciclo_mapping.get(ciclo_nombre)
            if not ciclo_numero:
                logger.warning(f"⚠️  Ciclo no reconocido: {ciclo_nombre}")
                return
            
            logger.info(f"📚 Procesando {ciclo_nombre} (Ciclo {ciclo_numero})")
            
            # Iterar filas desde la 2 (omitir header)
            for row_num in range(2, sheet.max_row + 1):
                curso_data = self._extraer_curso_fila(sheet, row_num, ciclo_numero)
                if curso_data:
                    self.cursos_extraidos.append(curso_data)
                    
        except Exception as e:
            logger.error(f"❌ Error procesando {ciclo_nombre}: {str(e)}")
    
    def _procesar_hoja_general(self, sheet, sheet_name: str):
        """Procesa una hoja genérica con estructura completa de Libro1.xlsx"""
        try:
            logger.info(f"📄 Procesando hoja general: {sheet_name}")
            max_row = sheet.max_row
            for row_num in range(2, max_row + 1):
                curso_data = self._extraer_curso_desde_hoja_general(sheet, row_num)
                if curso_data:
                    self.cursos_extraidos.append(curso_data)
        except Exception as e:
            logger.error(f"❌ Error procesando hoja {sheet_name}: {str(e)}")

    def _extraer_curso_desde_hoja_general(self, sheet, row_num: int) -> Optional[Dict]:
        """Extrae datos de una fila para estructura general"""
        try:
            ciclo_raw = sheet.cell(row=row_num, column=3).value  # CICLO (C)
            asignatura_raw = sheet.cell(row=row_num, column=6).value  # ASIGNATURA (F)

            if not ciclo_raw or not asignatura_raw:
                return None

            ciclo = self._normalizar_ciclo(ciclo_raw)
            if not ciclo:
                return None

            nombre = str(asignatura_raw).strip().upper()
            if not nombre or nombre in {"TOTAL", "SUBTOTAL"}:
                return None

            modalidad_raw = sheet.cell(row=row_num, column=14).value  # Modalidad
            creditos_raw = sheet.cell(row=row_num, column=7).value  # Créditos
            horas_teoria = self._procesar_horas(sheet.cell(row=row_num, column=8).value)
            horas_practica = self._procesar_horas(sheet.cell(row=row_num, column=9).value)
            horas_laboratorio = self._procesar_horas(sheet.cell(row=row_num, column=10).value)

            alumnos_teoria = self._safe_int(sheet.cell(row=row_num, column=15).value)
            alumnos_practica = self._safe_int(sheet.cell(row=row_num, column=16).value)
            alumnos_laboratorio = self._safe_int(sheet.cell(row=row_num, column=17).value)

            grupos_teoria = self._safe_int(sheet.cell(row=row_num, column=18).value)
            grupos_practica = self._safe_int(sheet.cell(row=row_num, column=19).value)
            grupos_laboratorio = self._safe_int(sheet.cell(row=row_num, column=20).value)

            codigo_excel = sheet.cell(row=row_num, column=4).value  # COD
            numero_excel = sheet.cell(row=row_num, column=5).value  # NUM
            codigo = self._construir_codigo_curso(codigo_excel, numero_excel, nombre, ciclo)

            modalidad_final = self.modalidad_mapping.get(str(modalidad_raw or '').strip().upper(), 'presencial')
            creditos = self._safe_int(creditos_raw)
            if creditos <= 0:
                creditos = max(2, (horas_teoria + horas_practica + horas_laboratorio) // 2 or 2)

            curso_data = {
                'codigo': codigo,
                'nombre': nombre,
                'ciclo': ciclo,
                'modalidad': modalidad_final,
                'creditos': creditos,
                'horas_teoria': horas_teoria,
                'horas_practica': horas_practica,
                'horas_laboratorio': horas_laboratorio,
                'alumnos_teoria': alumnos_teoria,
                'alumnos_practica': alumnos_practica,
                'alumnos_laboratorio': alumnos_laboratorio,
                'grupos_teoria': grupos_teoria,
                'grupos_practica': grupos_practica,
                'grupos_laboratorio': grupos_laboratorio,
                'requiere_laboratorio': grupos_laboratorio > 0 or horas_laboratorio > 0,
                'requiere_practica': grupos_practica > 0 or horas_practica > 0,
                'fila_excel': row_num
            }

            return curso_data
        except Exception as e:
            logger.warning(f"⚠️  Error en fila {row_num}: {str(e)}")
            return None

    def _extraer_curso_fila(self, sheet, row_num: int, ciclo: int) -> Optional[Dict]:
        """Extrae datos de curso de una fila específica"""
        try:
            # Leer celdas clave (columnas C, F, N, H, I, J)
            ciclo_cell = sheet.cell(row=row_num, column=3).value  # C
            asignatura = sheet.cell(row=row_num, column=6).value  # F
            modalidad = sheet.cell(row=row_num, column=14).value  # N
            horas_teoria = sheet.cell(row=row_num, column=8).value  # H
            horas_practica = sheet.cell(row=row_num, column=9).value  # I
            horas_laboratorio = sheet.cell(row=row_num, column=10).value  # J
            
            # Validar datos mínimos
            if not asignatura or not isinstance(asignatura, str):
                return None
            
            if not ciclo_cell:
                return None
            ciclo_detectado = self._normalizar_ciclo(ciclo_cell)
            if ciclo_detectado != ciclo:
                return None
            
            # Limpiar y procesar
            asignatura = str(asignatura).strip().upper()
            modalidad = str(modalidad or 'PRS').strip().upper()
            
            # Procesar horas
            horas_teoria = self._procesar_horas(horas_teoria)
            horas_practica = self._procesar_horas(horas_practica)
            horas_laboratorio = self._procesar_horas(horas_laboratorio)
            
            # Calcular créditos (simplificado)
            creditos = max(2, min(6, (horas_teoria + horas_practica + horas_laboratorio) // 2))
            
            # Generar código único
            codigo = f"C{ciclo}_{self._generar_codigo_curso(asignatura)}"
            
            # Mapear modalidad
            modalidad_final = self.modalidad_mapping.get(modalidad, 'presencial')
            
            curso_data = {
                'codigo': codigo,
                'nombre': asignatura,
                'ciclo': ciclo,
                'modalidad': modalidad_final,
                'creditos': creditos,
                'horas_teoria': horas_teoria,
                'horas_practica': horas_practica,
                'horas_laboratorio': horas_laboratorio,
                'requiere_laboratorio': horas_laboratorio > 0,
                'requiere_practica': horas_practica > 0,
                'alumnos_teoria': 40 if horas_teoria > 0 else 0,
                'alumnos_practica': 20 if horas_practica > 0 else 0,
                'alumnos_laboratorio': 15 if horas_laboratorio > 0 else 0,
                'grupos_teoria': 1 if horas_teoria > 0 else 0,
                'grupos_practica': 1 if horas_practica > 0 else 0,
                'grupos_laboratorio': 1 if horas_laboratorio > 0 else 0,
            }
            
            return curso_data
            
        except Exception as e:
            logger.warning(f"⚠️  Error en fila {row_num}: {str(e)}")
            return None
    
    def _safe_int(self, valor) -> int:
        """Convierte valores numéricos (incluyendo None/NaN) a int seguro"""
        if valor is None:
            return 0
        if isinstance(valor, (int, bool)):
            return int(valor)
        if isinstance(valor, float):
            if math.isnan(valor):
                return 0
            return int(round(valor))
        valor_str = str(valor).strip()
        if not valor_str:
            return 0
        try:
            return int(float(valor_str.replace(',', '.')))
        except ValueError:
            return 0

    def _construir_codigo_curso(self, codigo_raw, numero_raw, nombre: str, ciclo: int) -> str:
        """Construye un código de curso estable usando columnas COD/NUM"""
        codigo_part = str(codigo_raw).strip().upper() if codigo_raw else ''
        numero_part = None
        if numero_raw not in (None, ''):
            try:
                numero_part = int(float(str(numero_raw).strip()))
            except ValueError:
                numero_part = None
        if codigo_part and numero_part is not None:
            base_code = f"{codigo_part}{numero_part:03d}"
            return self._asegurar_codigo_unico(base_code, nombre, ciclo)
        if codigo_part:
            base_code = f"{codigo_part}_{ciclo}"
            return self._asegurar_codigo_unico(base_code, nombre, ciclo)
        base_code = f"C{ciclo}_{self._generar_codigo_curso(nombre)}"
        return self._asegurar_codigo_unico(base_code, nombre, ciclo)

    def _asegurar_codigo_unico(self, base_code: str, nombre: str, ciclo: int) -> str:
        """Evita colisiones de código agregando sufijos cuando es necesario"""
        normalized_code = base_code.strip() if base_code else ''
        if not normalized_code:
            normalized_code = f"C{ciclo}_{self._generar_codigo_curso(nombre)}"

        if normalized_code not in self.codigo_counts and normalized_code not in self.codigos_generados:
            self.codigo_counts[normalized_code] = 1
            self.codigos_generados.add(normalized_code)
            return normalized_code

        count = self.codigo_counts.get(normalized_code, 1)
        while True:
            suffix_index = count - 1  # segunda ocurrencia -> 1 -> 'A'
            if suffix_index < 26:
                suffix = chr(ord('A') + suffix_index)
            else:
                suffix = f"X{suffix_index - 25}"
            candidate = f"{normalized_code}{suffix}"
            if candidate not in self.codigos_generados:
                self.codigo_counts[normalized_code] = count + 1
                self.codigos_generados.add(candidate)
                return candidate
            count += 1

    def _normalizar_ciclo(self, ciclo_raw) -> Optional[int]:
        """Normaliza el valor de ciclo desde la celda"""
        if ciclo_raw is None:
            return None
        ciclo_str = str(ciclo_raw).strip().upper()
        if not ciclo_str:
            return None
        if ciclo_str in self.ciclo_mapping:
            return self.ciclo_mapping[ciclo_str]
        if ciclo_str.startswith('C') and ciclo_str[1:].isdigit():
            return int(ciclo_str[1:])
        if ciclo_str.isdigit():
            return int(ciclo_str)
        try:
            # Manejar valores numéricos como 1.0, 2.0
            return int(float(ciclo_str))
        except ValueError:
            logger.warning(f"⚠️  No se pudo normalizar ciclo: {ciclo_raw}")
            return None

    def _procesar_horas(self, valor) -> int:
        """Procesa valor de horas desde Excel"""
        if valor is None:
            return 0
        if isinstance(valor, (int, float)):
            return max(0, int(valor))
        if isinstance(valor, str):
            try:
                return max(0, int(float(valor.strip())))
            except (ValueError, AttributeError):
                return 0
        return 0
    
    def _generar_codigo_curso(self, nombre: str) -> str:
        """Genera código corto para el curso"""
        palabras = nombre.split()
        if len(palabras) == 1:
            return palabras[0][:6]
        
        # Tomar primeras letras de cada palabra importante
        codigo_partes = []
        for palabra in palabras[:3]:  # Max 3 palabras
            if len(palabra) > 2:  # Omitir palabras muy cortas
                codigo_partes.append(palabra[:2])
        
        codigo = ''.join(codigo_partes).upper()
        return codigo[:6] if codigo else nombre[:6].upper()
    
    def _generate_summary(self) -> Dict[str, Any]:
        """Genera resumen de la extracción"""
        if not self.cursos_extraidos:
            return {}
        
        # Estadísticas
        total_creditos = sum(curso['creditos'] for curso in self.cursos_extraidos)
        cursos_por_ciclo = {}
        modalidades = {}
        
        for curso in self.cursos_extraidos:
            ciclo = curso['ciclo']
            modalidad = curso['modalidad']
            
            cursos_por_ciclo[ciclo] = cursos_por_ciclo.get(ciclo, 0) + 1
            modalidades[modalidad] = modalidades.get(modalidad, 0) + 1
        
        con_laboratorio = sum(1 for c in self.cursos_extraidos if c['requiere_laboratorio'])
        con_practica = sum(1 for c in self.cursos_extraidos if c['requiere_practica'])
        
        return {
            'total_cursos': len(self.cursos_extraidos),
            'total_creditos': total_creditos,
            'cursos_por_ciclo': cursos_por_ciclo,
            'modalidades': modalidades,
            'con_laboratorio': con_laboratorio,
            'con_practica': con_practica,
            'promedio_creditos': round(total_creditos / len(self.cursos_extraidos), 2)
        }

class CourseIntegratorService:
    """
    Servicio para integrar cursos extraídos a la base de datos
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._nrc_generator: Optional[NRCGenerator] = None
    
    def replace_all_courses(self, courses_data: List[Dict[str, Any]], db: Session = None) -> Dict[str, Any]:
        """
        Reemplaza todos los cursos en la base de datos
        """
        if db is None:
            db = SessionLocal()
            close_db = True
        else:
            close_db = False
        
        try:
            self.logger.info("🗑️  Eliminando cursos existentes...")
            
            # Eliminar secciones primero (FK constraint)
            db.query(CourseSection).delete()
            
            # Eliminar cursos
            deleted_courses = db.query(Course).delete()
            self.logger.info(f"🗑️  Eliminados {deleted_courses} cursos existentes")
            
            # Insertar nuevos cursos
            self.logger.info("📚 Insertando cursos nuevos...")
            inserted_courses = 0
            inserted_sections = 0
            self._nrc_generator = None
            
            for i, curso_data in enumerate(courses_data):
                # Crear curso
                curso = Course(
                    codigo=curso_data['codigo'],
                    nombre=curso_data['nombre'],
                    ciclo=curso_data['ciclo'],
                    modalidad=curso_data['modalidad'],
                    creditos=curso_data['creditos'],
                    alumnos_teoria=curso_data['alumnos_teoria'],
                    alumnos_practica=curso_data['alumnos_practica'],
                    alumnos_laboratorio=curso_data['alumnos_laboratorio'],
                    grupos_teoria=curso_data['grupos_teoria'],
                    grupos_practica=curso_data['grupos_practica'],
                    grupos_laboratorio=curso_data['grupos_laboratorio'],
                    requiere_laboratorio=curso_data['requiere_laboratorio'],
                    requiere_practica=curso_data['requiere_practica'],
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                    active=1
                )
                
                db.add(curso)
                db.flush()  # Para obtener el ID
                
                # Crear secciones
                sections_created = self._create_course_sections(curso, db)
                inserted_sections += sections_created
                inserted_courses += 1
                
                # Log progreso cada 10 cursos
                if (i + 1) % 10 == 0:
                    self.logger.info(f"📚 Insertados {i + 1}/{len(courses_data)} cursos...")
            
            db.commit()
            
            # Verificar estado final
            final_courses = db.query(Course).count()
            final_sections = db.query(CourseSection).count()
            
            result = {
                "success": True,
                "courses_inserted": inserted_courses,
                "sections_created": inserted_sections,
                "final_courses_count": final_courses,
                "final_sections_count": final_sections
            }
            
            self.logger.info(f"✅ Integración completada: {inserted_courses} cursos, {inserted_sections} secciones")
            return result
            
        except Exception as e:
            db.rollback()
            self.logger.error(f"❌ Error en integración: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "courses_inserted": 0,
                "sections_created": 0
            }
        finally:
            if close_db:
                db.close()
    
    def _ensure_nrc_generator(self, db: Session) -> None:
        if self._nrc_generator is None:
            self._nrc_generator = NRCGenerator(db)

    def _next_nrc(self, db: Session) -> str:
        self._ensure_nrc_generator(db)
        return self._nrc_generator.next()

    def _create_course_sections(self, curso: Course, db: Session) -> int:
        """Crea secciones para un curso respetando las ligas T→P→L"""

        self._ensure_nrc_generator(db)

        def _balanced_distribution(total: int, buckets: int) -> list[int]:
            if buckets <= 0:
                return []
            base = total // buckets
            remainder = total % buckets
            return [base + (1 if idx < remainder else 0) for idx in range(buckets)]

        def _students_per_group(total_groups: int, total_students: int) -> list[int]:
            if total_groups <= 0:
                return []
            base = total_students // total_groups
            remainder = total_students % total_groups
            return [base + (1 if idx < remainder else 0) for idx in range(total_groups)]

        def _section_label(prefix: str, league: int) -> str:
            return f"{prefix}{league}"

        sections_created = 0

        total_leagues = curso.grupos_teoria
        if total_leagues <= 0:
            if curso.grupos_practica > 0 or curso.grupos_laboratorio > 0:
                total_leagues = max(curso.grupos_practica, curso.grupos_laboratorio, 1)
            else:
                total_leagues = 0

        # Secciones de teoría (una liga por teoría)
        teoria_students = _students_per_group(curso.grupos_teoria, curso.alumnos_teoria)
        for idx in range(curso.grupos_teoria):
            league = idx + 1
            alumnos_seccion = teoria_students[idx] if idx < len(teoria_students) else 0
            section = CourseSection(
                course_id=curso.id,
                tipo='teoria',
                seccion=_section_label('T', league),
                league=league,
                nrc=self._next_nrc(db),
                alumnos_proyectados=alumnos_seccion,
                alumnos_reales=0,
                activa=1,
                created_at=datetime.now()
            )
            db.add(section)
            sections_created += 1

        if total_leagues == 0:
            return sections_created

        # Distribución equilibrada para prácticas y laboratorios
        practica_distribution = _balanced_distribution(curso.grupos_practica, total_leagues)
        practica_students = _students_per_group(curso.grupos_practica, curso.alumnos_practica)
        practica_pointer = 0

        for league, count in enumerate(practica_distribution, start=1):
            for _ in range(count):
                alumnos_seccion = (
                    practica_students[practica_pointer] if practica_pointer < len(practica_students) else 0
                )
                practica_pointer += 1
                section = CourseSection(
                    course_id=curso.id,
                    tipo='practica',
                    seccion=_section_label('P', league),
                    league=league,
                    nrc=self._next_nrc(db),
                    alumnos_proyectados=alumnos_seccion,
                    alumnos_reales=0,
                    activa=1,
                    created_at=datetime.now()
                )
                db.add(section)
                sections_created += 1

        laboratorio_distribution = _balanced_distribution(curso.grupos_laboratorio, total_leagues)
        laboratorio_students = _students_per_group(curso.grupos_laboratorio, curso.alumnos_laboratorio)
        laboratorio_pointer = 0

        for league, count in enumerate(laboratorio_distribution, start=1):
            for _ in range(count):
                alumnos_seccion = (
                    laboratorio_students[laboratorio_pointer] if laboratorio_pointer < len(laboratorio_students) else 0
                )
                laboratorio_pointer += 1
                section = CourseSection(
                    course_id=curso.id,
                    tipo='laboratorio',
                    seccion=_section_label('L', league),
                    league=league,
                    nrc=self._next_nrc(db),
                    alumnos_proyectados=alumnos_seccion,
                    alumnos_reales=0,
                    activa=1,
                    created_at=datetime.now()
                )
                db.add(section)
                sections_created += 1

        return sections_created