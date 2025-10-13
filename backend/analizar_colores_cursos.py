"""
Análisis de colores para identificar cursos correctamente
Cada color de fondo = Un curso diferente
"""

from openpyxl import load_workbook
from collections import defaultdict
import json

EXCEL_PATH = '../inputs/Horario_Docentes(2025-20).xlsx'

wb = load_workbook(EXCEL_PATH, data_only=True)

# Analizar la primera hoja en detalle
sheet_name = 'A. Caballero'
ws = wb[sheet_name]

print("="*80)
print(f"ANALISIS DE COLORES - HOJA: {sheet_name}")
print("="*80)

# Encontrar fila de encabezados
fila_encabezado = None
for row_idx in range(1, 10):
    for col_idx in range(1, 10):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str) and 'LUNES' in cell.value.upper():
            fila_encabezado = row_idx
            break
    if fila_encabezado:
        break

print(f"\nFila encabezado: {fila_encabezado}")

# Analizar todas las celdas con contenido
colores_cursos = defaultdict(list)

for row_idx in range(fila_encabezado + 1, min(fila_encabezado + 20, ws.max_row + 1)):
    for col_idx in range(4, 8):  # Columnas de días
        cell = ws.cell(row=row_idx, column=col_idx)
        
        if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 5:
            # Obtener color
            color = "SIN_COLOR"
            if cell.fill and cell.fill.start_color:
                if hasattr(cell.fill.start_color, 'rgb'):
                    color = str(cell.fill.start_color.rgb)
                elif hasattr(cell.fill.start_color, 'index'):
                    color = f"Index:{cell.fill.start_color.index}"
            
            contenido = cell.value.strip()[:50]
            
            colores_cursos[color].append({
                'fila': row_idx,
                'col': col_idx,
                'contenido': contenido
            })

print("\n" + "="*80)
print("CURSOS AGRUPADOS POR COLOR:")
print("="*80)

for color, celdas in colores_cursos.items():
    print(f"\n>>> COLOR: {color}")
    print(f"    Total celdas: {len(celdas)}")
    # Mostrar primera celda como ejemplo
    if celdas:
        print(f"    Ejemplo: {celdas[0]['contenido']}")
        # Mostrar todas las primeras 3
        for i, celda in enumerate(celdas[:3]):
            print(f"      [{i+1}] Fila {celda['fila']}: {celda['contenido']}")

print("\n" + "="*80)
print("CONCLUSIÓN:")
print("="*80)
print(f"Total de colores diferentes (cursos): {len(colores_cursos)}")
