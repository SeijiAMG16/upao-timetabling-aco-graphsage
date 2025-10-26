"""
Exporta horarios individuales por profesor en UN SOLO archivo Excel
URGENTE - Para presentación de tesis
Formato: Una hoja por profesor + hoja índice
"""
import sys
sys.path.insert(0, 'c:\\Users\\amaya\\Downloads\\10mo Ciclo\\TESIS\\upao-timetabling-aco-graphsage\\backend')

import pandas as pd
from app.database import SessionLocal
from app.models import CourseSection, Course, Professor, Classroom, TimeSlot
from sqlalchemy.orm import joinedload
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import glob

def crear_hoja_profesor(wb, nombre_hoja, profesor_data, asignaciones_profesor):
    """
    Crea una hoja en el workbook para un profesor específico
    
    Args:
        wb: Workbook de openpyxl
        nombre_hoja: nombre de la hoja
        profesor_data: dict con info del profesor
        asignaciones_profesor: lista de asignaciones
    """
    ws = wb.create_sheet(title=nombre_hoja)
    
    # CONFIGURACIÓN
    dias_semana = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']
    
    # Definir bloques horarios REALES (16 timeslots de 50 minutos cada uno)
    bloques_horarios = [
        ('07:00', '07:50'),
        ('07:55', '08:45'),
        ('08:50', '09:40'),
        ('09:45', '10:35'),
        ('10:40', '11:30'),
        ('11:35', '12:25'),
        ('12:30', '13:20'),
        ('13:25', '14:15'),
        ('14:20', '15:10'),
        ('15:15', '16:05'),
        ('16:10', '17:00'),
        ('17:05', '17:55'),
        ('18:00', '18:50'),
        ('18:55', '19:45'),
        ('19:50', '20:40'),
        ('20:45', '21:35'),
    ]
    
    # ESTILOS
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    clase_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    clase_font = Font(name='Calibri', size=10)
    
    # ENCABEZADO DEL PROFESOR
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = f"HORARIO DE CLASES - {profesor_data['nombre'].upper()}"
    cell.font = Font(name='Calibri', size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = f"Código: {profesor_data['codigo']}"
    cell.font = Font(name='Calibri', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ENCABEZADOS DE COLUMNAS (Días)
    ws['A4'] = 'HORARIO'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].border = cell_border
    
    for idx, dia in enumerate(dias_semana, start=2):
        col_letter = get_column_letter(idx)
        cell = ws[f'{col_letter}4']
        cell.value = dia
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    
    # Crear estructura de horario vacía
    horario_grid = {}
    for bloque in bloques_horarios:
        horario_grid[bloque] = {dia: [] for dia in dias_semana}
    
    # Llenar con las asignaciones
    for asig in asignaciones_profesor:
        dia = asig['dia']
        hora_inicio = asig['hora_inicio']
        hora_fin = asig['hora_fin']
        
        # Encontrar el bloque correspondiente
        for bloque_inicio, bloque_fin in bloques_horarios:
            if hora_inicio >= bloque_inicio and hora_inicio < bloque_fin:
                info_clase = (
                    f"{asig['nombre_curso']}\n"  # Usar nombre completo en lugar de código
                    f"{asig['tipo']}-{asig['seccion']} (L{asig['league']})\n"
                    f"{asig['aula']}\n"
                    f"{hora_inicio}-{hora_fin}\n"
                    f"{asig['estudiantes']} est."
                )
                horario_grid[(bloque_inicio, bloque_fin)][dia].append(info_clase)
    
    # ESCRIBIR FILAS DE HORARIO
    row_idx = 5
    for bloque_inicio, bloque_fin in bloques_horarios:
        # Columna de hora
        ws[f'A{row_idx}'] = f"{bloque_inicio}\n{bloque_fin}"
        ws[f'A{row_idx}'].font = Font(name='Calibri', size=10, bold=True)
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'A{row_idx}'].border = cell_border
        ws[f'A{row_idx}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Columnas de días
        for idx, dia in enumerate(dias_semana, start=2):
            col_letter = get_column_letter(idx)
            clases = horario_grid[(bloque_inicio, bloque_fin)][dia]
            
            cell = ws[f'{col_letter}{row_idx}']
            if clases:
                cell.value = '\n---\n'.join(clases)
                cell.fill = clase_fill
                cell.font = clase_font
            else:
                cell.value = ''
            
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
        
        row_idx += 1
    
    # AJUSTAR TAMAÑOS
    ws.column_dimensions['A'].width = 12
    for idx in range(2, 8):  # Columnas B-G (días)
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    # Con 16 bloques, usar altura más pequeña para que entre mejor
    for row in range(5, row_idx):
        ws.row_dimensions[row].height = 60
    
    # ESTADÍSTICAS AL FINAL
    ws[f'A{row_idx + 2}'] = 'RESUMEN:'
    ws[f'A{row_idx + 2}'].font = Font(name='Calibri', size=11, bold=True)
    
    total_clases = len(asignaciones_profesor)
    cursos_unicos = len(set(a['curso'] for a in asignaciones_profesor))
    total_estudiantes = sum(a['estudiantes'] for a in asignaciones_profesor)
    
    ws[f'A{row_idx + 3}'] = f"Total de clases: {total_clases}"
    ws[f'A{row_idx + 4}'] = f"Cursos diferentes: {cursos_unicos}"
    ws[f'A{row_idx + 5}'] = f"Total estudiantes: {total_estudiantes}"

def exportar_horarios_un_solo_archivo():
    """Genera UN SOLO archivo Excel con una hoja por profesor"""
    
    print("="*80)
    print("EXPORTANDO HORARIOS DE PROFESORES EN UN SOLO ARCHIVO")
    print("="*80)
    
    # Buscar el archivo JSON más reciente
    json_files = glob.glob("horario_generado_*.json")
    if not json_files:
        print("❌ ERROR: No se encontró ningún horario generado")
        return None
    
    json_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = json_files[0]
    
    print(f"📄 Archivo encontrado: {latest_file}")
    
    # Cargar el JSON
    with open(latest_file, 'r', encoding='utf-8') as f:
        horario_data = json.load(f)
    
    asignaciones = horario_data.get('asignaciones', [])
    print(f"✅ Horario cargado: {len(asignaciones)} asignaciones")
    
    if not asignaciones:
        print("❌ ERROR: El archivo JSON no contiene asignaciones")
        return None
    
    # Conectar a BD
    db = SessionLocal()
    
    # Organizar asignaciones por profesor
    asignaciones_por_profesor = {}
    profesores_info = {}
    
    print("\n📊 Procesando asignaciones...")
    
    for asig in asignaciones:
        section_id = asig['section_id']
        professor_id = asig['professor_id']
        classroom_id = asig['classroom_id']
        timeslot_ids = asig['timeslot_ids']
        
        # Obtener información de BD
        section = db.query(CourseSection).options(
            joinedload(CourseSection.course)
        ).filter(CourseSection.id == section_id).first()
        
        if not section:
            continue
        
        professor = db.query(Professor).filter(Professor.id == professor_id).first()
        if not professor:
            continue
        
        # Guardar info del profesor
        if professor_id not in profesores_info:
            profesores_info[professor_id] = {
                'id': professor_id,
                'codigo': professor.codigo,
                'nombre': professor.nombre_completo
            }
        
        classroom = db.query(Classroom).filter(Classroom.id == classroom_id).first()
        aula_codigo = classroom.codigo if classroom else 'VIRTUAL'
        
        # Obtener timeslots
        timeslots = db.query(TimeSlot).filter(
            TimeSlot.id.in_(timeslot_ids)
        ).order_by(TimeSlot.dia_semana, TimeSlot.hora_inicio).all()
        
        if not timeslots:
            continue
        
        # Mapeo de días
        dias_map = {
            1: "LUNES",
            2: "MARTES", 
            3: "MIÉRCOLES",
            4: "JUEVES",
            5: "VIERNES",
            6: "SÁBADO"
        }
        
        dia = dias_map.get(timeslots[0].dia_semana, f"DÍA {timeslots[0].dia_semana}")
        hora_inicio = str(timeslots[0].hora_inicio)[:5]
        hora_fin = str(timeslots[-1].hora_fin)[:5]
        
        # Crear registro de asignación
        asig_info = {
            'curso': section.course.codigo if section.course else 'N/A',
            'nombre_curso': section.course.nombre if section.course else 'N/A',
            'tipo': section.tipo,
            'seccion': section.seccion,
            'league': section.league,
            'nrc': section.nrc,
            'estudiantes': section.alumnos_proyectados,
            'aula': aula_codigo,
            'dia': dia,
            'hora_inicio': hora_inicio,
            'hora_fin': hora_fin,
        }
        
        # Agregar a lista del profesor
        if professor_id not in asignaciones_por_profesor:
            asignaciones_por_profesor[professor_id] = []
        
        asignaciones_por_profesor[professor_id].append(asig_info)
    
    db.close()
    
    print(f"✅ {len(asignaciones_por_profesor)} profesores con asignaciones")
    
    # Crear UN SOLO archivo Excel
    print(f"\n📝 Creando archivo Excel único...")
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # CREAR HOJA ÍNDICE PRIMERO
    print("   📋 Creando hoja de índice...")
    ws_index = wb.create_sheet(title="ÍNDICE")
    
    # Encabezados del índice
    headers = ['#', 'Código', 'Nombre Completo', 'Total Clases', 'Cursos Diferentes', 'Total Estudiantes']
    for idx, header in enumerate(headers, start=1):
        cell = ws_index.cell(row=1, column=idx)
        cell.value = header
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ordenar profesores por código
    profesores_ordenados = sorted(asignaciones_por_profesor.items(), 
                                  key=lambda x: profesores_info[x[0]]['codigo'])
    
    # Llenar datos del índice
    row_idx = 2
    for numero, (professor_id, asignaciones_prof) in enumerate(profesores_ordenados, start=1):
        profesor_info = profesores_info[professor_id]
        total_clases = len(asignaciones_prof)
        cursos_unicos = len(set(a['curso'] for a in asignaciones_prof))
        total_estudiantes = sum(a['estudiantes'] for a in asignaciones_prof)
        
        ws_index.cell(row=row_idx, column=1, value=numero)
        ws_index.cell(row=row_idx, column=2, value=profesor_info['codigo'])
        ws_index.cell(row=row_idx, column=3, value=profesor_info['nombre'])
        ws_index.cell(row=row_idx, column=4, value=total_clases)
        ws_index.cell(row=row_idx, column=5, value=cursos_unicos)
        ws_index.cell(row=row_idx, column=6, value=total_estudiantes)
        
        row_idx += 1
    
    # Ajustar anchos del índice
    ws_index.column_dimensions['A'].width = 5
    ws_index.column_dimensions['B'].width = 12
    ws_index.column_dimensions['C'].width = 40
    ws_index.column_dimensions['D'].width = 15
    ws_index.column_dimensions['E'].width = 18
    ws_index.column_dimensions['F'].width = 18
    
    # CREAR HOJAS DE PROFESORES
    print("\n📊 Creando hojas de profesores:")
    print("-" * 80)
    
    for numero, (professor_id, asignaciones_prof) in enumerate(profesores_ordenados, start=1):
        profesor_info = profesores_info[professor_id]
        
        # Nombre de hoja seguro (máximo 31 caracteres para Excel)
        codigo = profesor_info['codigo']
        nombre_corto = profesor_info['nombre'].split()[0]  # Primer nombre/apellido
        nombre_hoja = f"{numero:02d}_{codigo}_{nombre_corto}"[:31]
        
        print(f"   {numero:2d}. {nombre_hoja}")
        
        crear_hoja_profesor(wb, nombre_hoja, profesor_info, asignaciones_prof)
    
    print("-" * 80)
    
    # Guardar archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"HORARIOS_PROFESORES_UPAO_{timestamp}.xlsx"
    wb.save(output_filename)
    
    print(f"\n✅ COMPLETADO: Archivo generado con {len(asignaciones_por_profesor)} hojas de profesores + índice")
    print(f"📂 Archivo: {os.path.abspath(output_filename)}")
    
    return output_filename

if __name__ == "__main__":
    try:
        output_file = exportar_horarios_un_solo_archivo()
        if output_file:
            print("\n" + "="*80)
            print("✅ EXPORTACIÓN COMPLETADA")
            print("="*80)
            print(f"\n📄 Archivo Excel único: {output_file}")
            print("\n💡 Contenido:")
            print("   • Hoja 'ÍNDICE' con resumen de todos los profesores")
            print("   • Una hoja por cada profesor (44 hojas)")
            print("   • Formato de cuadrícula (días × horas)")
            print("   • Total: 45 hojas en un solo archivo Excel")
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
