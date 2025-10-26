"""
Genera Excel con horario por profesor (una hoja por cada profesor)
Lee del JSON generado por el ACO
"""
import json
import pandas as pd
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

sys.path.insert(0, os.path.dirname(__file__))
from app.database import SessionLocal
from sqlalchemy import text

# Colores para tipos de sesión
COLORES = {
    'T': {'fill': 'E3F2FD', 'font': '0D47A1'},  # Azul - Teoría
    'P': {'fill': 'F3E5F5', 'font': '6A1B9A'},  # Morado - Práctica
    'L': {'fill': 'E8F5E9', 'font': '2E7D32'}   # Verde - Laboratorio
}

DAY_NAMES = {
    1: "Lunes", 2: "Martes", 3: "Miércoles",
    4: "Jueves", 5: "Viernes", 6: "Sábado"
}

def get_timeslot_info(timeslot_id):
    """Convierte timeslot_id a día y hora"""
    # 96 timeslots: 6 días × 16 bloques (8am-4pm, 30min cada uno)
    dia_idx = (timeslot_id - 1) // 16
    bloque_idx = (timeslot_id - 1) % 16
    
    dia = dia_idx + 1  # 1=Lunes, 2=Martes, etc.
    
    # Hora de inicio (8:00 AM = bloque 0)
    hora_inicio = 8.0 + (bloque_idx * 0.5)
    hora_fin = hora_inicio + 0.5
    
    def format_time(h):
        horas = int(h)
        minutos = int((h - horas) * 60)
        return f"{horas:02d}:{minutos:02d}"
    
    return dia, format_time(hora_inicio), format_time(hora_fin)

def clean_name(name: str) -> str:
    """Limpia nombre para Excel sheet"""
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name[:31]

def write_schedule_sheet(ws, prof_name, prof_assignments, cursos_dict, aulas_dict):
    """Escribe hoja de horario para un profesor"""
    
    # Título
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"HORARIO - {prof_name}"
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Encabezados de días
    ws['A2'] = 'Hora'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    for i, (day_num, day_name) in enumerate(DAY_NAMES.items(), start=2):
        cell = ws.cell(row=2, column=i)
        cell.value = day_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Crear grid de horario: schedule_grid[hora][dia] = (contenido, tipo)
    schedule_grid = defaultdict(lambda: defaultdict(lambda: None))
    
    for assign in prof_assignments:
        timeslot_ids = assign['timeslot_ids']
        course_code = assign['course_code']
        session_type = assign['session_type']
        classroom_id = assign['classroom_id']
        league_id = assign['league_id']
        
        curso_nombre = cursos_dict.get(course_code, course_code)
        aula_codigo = aulas_dict.get(classroom_id, {}).get('codigo', f'Aula {classroom_id}')
        
        # Procesar cada timeslot
        for ts_id in timeslot_ids:
            dia, hora_inicio, hora_fin = get_timeslot_info(ts_id)
            time_str = f"{hora_inicio}-{hora_fin}"
            
            # Contenido de la celda
            content = f"{course_code}\n{curso_nombre[:30]}\nLiga {league_id}\n{aula_codigo}"
            
            schedule_grid[time_str][dia] = (content, session_type)
    
    # Escribir horario
    row = 3
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Obtener todas las horas únicas y ordenarlas
    all_times = sorted(schedule_grid.keys())
    
    for time_str in all_times:
        # Hora
        hour_cell = ws.cell(row=row, column=1, value=time_str)
        hour_cell.alignment = Alignment(horizontal='center', vertical='center')
        hour_cell.font = Font(bold=True)
        hour_cell.border = border
        
        # Días
        for day in range(1, 7):
            col = day + 1
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            if day in schedule_grid[time_str]:
                content, session_type = schedule_grid[time_str][day]
                cell.value = content
                
                # Aplicar color según tipo
                tipo = session_type.upper()
                if tipo in COLORES:
                    cell.fill = PatternFill(
                        start_color=COLORES[tipo]['fill'],
                        end_color=COLORES[tipo]['fill'],
                        fill_type='solid'
                    )
                    cell.font = Font(color=COLORES[tipo]['font'], size=9)
                
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        ws.row_dimensions[row].height = 60
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 25

def main():
    print("="*80)
    print("GENERANDO EXCEL CON HORARIOS POR PROFESOR")
    print("="*80)
    
    # Cargar JSON
    horario_file = 'horario_generado_20251022_015751.json'
    print(f"\n📂 Leyendo: {horario_file}")
    
    with open(horario_file, 'r', encoding='utf-8') as f:
        horario = json.load(f)
    
    asignaciones = horario['asignaciones']
    print(f"   Total asignaciones: {len(asignaciones)}")
    
    # Conectar a BD
    session = SessionLocal()
    
    # Cargar datos
    print("\n📚 Cargando datos de profesores, cursos y aulas...")
    
    # Profesores
    profesores = {}
    result = session.execute(text("SELECT id, codigo, nombre_completo FROM professors")).fetchall()
    for prof_id, codigo, nombre in result:
        profesores[prof_id] = {'codigo': codigo, 'nombre': nombre}
    
    # Cursos
    cursos = {}
    result = session.execute(text("SELECT codigo, nombre FROM courses")).fetchall()
    for codigo, nombre in result:
        cursos[codigo] = nombre
    
    # Aulas
    aulas = {}
    result = session.execute(text("SELECT id, codigo, edificio FROM classrooms")).fetchall()
    for aula_id, codigo, edificio in result:
        aulas[aula_id] = {'codigo': codigo, 'edificio': edificio}
    
    session.close()
    
    # Agrupar por profesor
    print("\n👨‍🏫 Agrupando asignaciones por profesor...")
    by_professor = defaultdict(list)
    
    for assign in asignaciones:
        prof_id = assign['professor_id']
        by_professor[prof_id].append(assign)
    
    print(f"   Total profesores con asignaciones: {len(by_professor)}")
    
    # Crear Excel
    print("\n📊 Creando archivo Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Crear hoja por profesor
    for prof_id, prof_assignments in sorted(by_professor.items()):
        prof_info = profesores.get(prof_id, {'codigo': f'PROF_{prof_id}', 'nombre': f'Profesor {prof_id}'})
        prof_name = f"{prof_info['codigo']} - {prof_info['nombre']}"
        
        # Crear hoja
        sheet_name = clean_name(prof_name)
        ws = wb.create_sheet(title=sheet_name)
        
        # Escribir horario
        write_schedule_sheet(ws, prof_name, prof_assignments, cursos, aulas)
        
        print(f"   ✅ {prof_name}: {len(prof_assignments)} asignaciones")
    
    # Guardar archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'horarios_profesores_{timestamp}.xlsx'
    wb.save(output_file)
    
    print(f"\n{'='*80}")
    print(f"✅ ARCHIVO EXCEL GENERADO EXITOSAMENTE")
    print(f"{'='*80}")
    print(f"\n📁 Archivo: {output_file}")
    print(f"📊 Hojas creadas: {len(by_professor)} (una por cada profesor)")
    print(f"🎨 Formato: Horario tipo tabla con días y horas")
    print(f"🌈 Colores: Azul (Teoría), Morado (Práctica), Verde (Lab)")
    print(f"\n🎉 Listo para abrir en Excel")

if __name__ == "__main__":
    main()
