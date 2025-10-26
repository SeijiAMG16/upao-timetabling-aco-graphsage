"""
Exportador de Horarios ACO+GraphSAGE a Excel

Genera archivos Excel con los horarios generados por el algoritmo.
Actualizado para trabajar con la tabla schedule_assignments.
"""

import pandas as pd
from sqlalchemy.orm import Session
from typing import Dict, List
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import (
    ScheduleAssignment,
    Professor,
    Classroom,
    TimeSlot,
    CourseSection,
    Course,
    AlgorithmExecution,
)


# Colores para tipos de sesión
COLORES = {
    'T': {'fill': 'E3F2FD', 'font': '0D47A1'},  # Azul - Teoría
    'P': {'fill': 'F3E5F5', 'font': '6A1B9A'},  # Morado - Práctica
    'L': {'fill': 'E8F5E9', 'font': '2E7D32'}   # Verde - Laboratorio
}

DAY_NAMES = {
    1: "Lunes", 2: "Martes", 3: "Miércoles",
    4: "Jueves", 5: "Viernes", 6: "Sábado"
}


class ExcelScheduleExporter:
    """Exportador de horarios a Excel con formato"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_professor_schedules(self, execution_id: int, output_path: str):
        """Exporta horarios por profesor con formato"""
        
        print(f"📊 Generando Excel por profesor...")
        
        # Obtener asignaciones
        assignments = (
            self.db.query(ScheduleAssignment)
            .filter(ScheduleAssignment.algorithm_execution_id == execution_id)
            .all()
        )
        
        if not assignments:
            print("⚠️  No hay asignaciones")
            return
        
        # Agrupar por profesor
        by_professor = defaultdict(list)
        for assign in assignments:
            by_professor[assign.professor_id].append(assign)
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        for prof_id, prof_assignments in by_professor.items():
            professor = self.db.query(Professor).get(prof_id)
            
            # Crear hoja
            ws = wb.create_sheet(title=self._clean_name(professor.nombre_completo))
            
            # Escribir horario
            self._write_schedule_sheet(ws, prof_assignments, f"Profesor: {professor.nombre_completo}")
            
            print(f"  ✅ {professor.nombre_completo}: {len(prof_assignments)} bloques")
        
        # Guardar
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        
        print(f"\n✅ Archivo generado: {output_file}")
    
    def _write_schedule_sheet(self, ws, assignments: List[ScheduleAssignment], title: str):
        """Escribe una hoja de horario con formato"""
        
        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezados de días
        ws['A2'] = 'Hora'
        for i, day_name in enumerate(DAY_NAMES.values(), start=2):
            cell = ws.cell(row=2, column=i)
            cell.value = day_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Crear estructura de horario
        schedule_grid = defaultdict(lambda: defaultdict(str))
        
        # Llenar grid
        for assign in assignments:
            timeslot = self.db.query(TimeSlot).get(assign.timeslot_id)
            section = self.db.query(CourseSection).get(assign.section_id)
            classroom = self.db.query(Classroom).get(assign.classroom_id)
            
            if timeslot and section and classroom:
                course = section.course
                day = timeslot.dia_semana
                time_str = f"{timeslot.hora_inicio.strftime('%H:%M')} - {timeslot.hora_fin.strftime('%H:%M')}"
                
                content = f"{course.nombre if course else ''}\n{section.seccion}\n{classroom.codigo}"
                schedule_grid[time_str][day] = (content, section.tipo)
        
        # Escribir horario
        row = 3
        for time_str in sorted(schedule_grid.keys()):
            ws.cell(row=row, column=1, value=time_str)
            
            for day in range(1, 7):
                col = day + 1
                if day in schedule_grid[time_str]:
                    content, session_type = schedule_grid[time_str][day]
                    cell = ws.cell(row=row, column=col, value=content)
                    
                    # Aplicar color
                    tipo = session_type.upper()
                    if tipo in COLORES:
                        cell.fill = PatternFill(start_color=COLORES[tipo]['fill'],
                                              end_color=COLORES[tipo]['fill'],
                                              fill_type='solid')
                        cell.font = Font(color=COLORES[tipo]['font'])
                    
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 20
    
    def _clean_name(self, name: str) -> str:
        """Limpia nombre para usar como nombre de hoja"""
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name[:31]


def export_all_formats(db: Session, execution_id: int, output_dir: str = "resultados"):
    """Exporta todos los formatos de horarios"""
    
    exporter = ExcelScheduleExporter(db)
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    print(f"\n{'='*80}")
    print(f"Exportando horarios - Ejecución #{execution_id}")
    print(f"{'='*80}\n")
    
    # Exportar por profesor
    exporter.export_professor_schedules(
        execution_id=execution_id,
        output_path=str(output_path / f"horarios_profesores_{timestamp}.xlsx"),
    )
    
    print(f"\n{'='*80}")
    print(f"✅ Exportación completada: {output_path}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    """Ejecutar desde CLI"""
    import sys
    from app.database import SessionLocal
    
    if len(sys.argv) < 2:
        print("Uso: python export_schedules_excel.py <execution_id>")
        sys.exit(1)
    
    execution_id = int(sys.argv[1])
    
    db = SessionLocal()
    try:
        export_all_formats(db, execution_id)
    finally:
        db.close()
