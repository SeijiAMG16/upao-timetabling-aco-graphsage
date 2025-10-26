"""
================================================================================
EXTRACTOR CONFIABLE DE ASIGNACIONES Y RESTRICCIONES V2
================================================================================
Extracción 100% confiable basada en:
1. Nombres de profesores dentro de cada celda (ej: "- Armando")
2. Color EXACTO de celdas bloqueadas: FFFF0000 (rojo)
3. Mapeo automático a BD usando coincidencia de nombres
================================================================================
"""

import pandas as pd
import json
import re
from pathlib import Path
import mysql.connector
from openpyxl import load_workbook
from collections import defaultdict

# Configuración
EXCEL_PATH = Path('../inputs/Horario_Docentes(2025-20).xlsx')
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'sistemas',
    'database': 'upao_timetabling'
}

DIAS_SEMANA = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']

# Bloques horarios UPAO (17 bloques de 50 min)
BLOQUES_TIEMPO = [
    ("07:00:00", "07:50:00"),   # Bloque 1
    ("07:55:00", "08:45:00"),   # Bloque 2
    ("08:50:00", "09:40:00"),   # Bloque 3
    ("09:45:00", "10:35:00"),   # Bloque 4
    ("10:40:00", "11:30:00"),   # Bloque 5
    ("11:35:00", "12:25:00"),   # Bloque 6
    ("12:30:00", "13:20:00"),   # Bloque 7
    ("13:25:00", "14:15:00"),   # Bloque 8
    ("14:20:00", "15:10:00"),   # Bloque 9
    ("15:15:00", "16:05:00"),   # Bloque 10
    ("16:10:00", "17:00:00"),   # Bloque 11
    ("17:05:00", "17:55:00"),   # Bloque 12
    ("18:00:00", "18:50:00"),   # Bloque 13
    ("18:55:00", "19:45:00"),   # Bloque 14
    ("19:50:00", "20:40:00"),   # Bloque 15
    ("20:45:00", "21:35:00"),   # Bloque 16
    ("21:40:00", "22:30:00")    # Bloque 17
]


def conectar_db():
    """Conecta a la base de datos"""
    return mysql.connector.connect(**DB_CONFIG)


def cargar_todos_profesores():
    """Carga todos los profesores de la BD para mapeo"""
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT id, nombre_completo, nombres, apellidos
        FROM professors
        ORDER BY id
    """)
    
    profesores = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return profesores


def cargar_todos_cursos():
    """Carga todos los cursos de la BD"""
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT id, codigo, nombre
        FROM courses
        ORDER BY id
    """)
    
    cursos = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return cursos


def encontrar_profesor_por_nombre_parcial(nombre_parcial, profesores):
    """Encuentra un profesor por coincidencia parcial del nombre"""
    if not nombre_parcial:
        return None
    
    nombre_limpio = nombre_parcial.strip().upper()
    
    # Buscar coincidencia exacta primero
    for prof in profesores:
        nombre_completo = (prof['nombre_completo'] or '').upper()
        nombres = (prof['nombres'] or '').upper()
        apellidos = (prof['apellidos'] or '').upper()
        
        if nombre_limpio in nombre_completo or nombre_limpio in apellidos or nombre_limpio in nombres:
            return prof
    
    # Buscar por palabras
    palabras = nombre_limpio.split()
    for palabra in palabras:
        if len(palabra) > 3:  # Palabras significativas
            for prof in profesores:
                nombre_completo = (prof['nombre_completo'] or '').upper()
                if palabra in nombre_completo:
                    return prof
    
    return None


def normalizar_nombre_curso(nombre):
    """Normaliza el nombre del curso"""
    if not nombre or not isinstance(nombre, str):
        return ""
    
    # Remover texto entre paréntesis
    nombre = re.sub(r'\([^)]*\)', '', nombre)
    
    # Limpiar espacios y convertir a mayúsculas
    nombre = nombre.strip().upper()
    
    # Remover caracteres especiales y números de sección
    nombre = re.sub(r'\d{4,}', '', nombre)  # Remover códigos de 4+ dígitos
    nombre = re.sub(r'[^\w\s]', ' ', nombre)
    nombre = re.sub(r'\s+', ' ', nombre)
    
    return nombre.strip()


def encontrar_curso_por_nombre(nombre_curso, cursos):
    """Encuentra un curso por coincidencia de nombre"""
    if not nombre_curso:
        return None
    
    nombre_normalizado = normalizar_nombre_curso(nombre_curso)
    
    # Buscar por coincidencia
    mejores_matches = []
    for curso in cursos:
        nombre_bd = curso['nombre'].upper()
        codigo_bd = curso['codigo'].upper()
        
        # Calcular similitud simple
        palabras_busqueda = set(nombre_normalizado.split())
        palabras_curso = set(nombre_bd.split())
        
        coincidencias = len(palabras_busqueda & palabras_curso)
        
        if coincidencias > 0:
            mejores_matches.append((curso, coincidencias))
    
    if mejores_matches:
        # Ordenar por número de coincidencias
        mejores_matches.sort(key=lambda x: x[1], reverse=True)
        return mejores_matches[0][0]
    
    return None


def extraer_nombre_profesor_de_celda(contenido):
    """Extrae el nombre del profesor de una celda (ej: '- Armando' -> 'Armando')"""
    if not contenido or not isinstance(contenido, str):
        return None
    
    # Buscar patrón "- NombreProfesor" al final
    match = re.search(r'-\s*([A-Za-zÁÉÍÓÚáéíóúÑñ\s]+?)(?:\s|$)', contenido)
    if match:
        nombre = match.group(1).strip()
        # Filtrar palabras comunes que no son nombres
        if nombre.upper() not in ['VIRTUAL', 'PRESENCIAL', 'NPR', 'PRS']:
            return nombre
    
    return None


def extraer_tipo_liga(texto):
    """Extrae el tipo (T/P/L) y liga (1,2,3) de un texto"""
    if not texto or not isinstance(texto, str):
        return None, None
    
    # Buscar patrón (T1), (P2), (L3), etc.
    match = re.search(r'\(([TPL])(\d+)\)', texto.upper())
    if match:
        tipo = match.group(1)
        liga = int(match.group(2))
        return tipo, liga
    
    return None, None


def es_celda_bloqueada(cell):
    """Verifica si una celda está bloqueada (color rojo EXACTO)"""
    if not cell.fill or not cell.fill.start_color:
        return False
    
    # Color rojo exacto
    if hasattr(cell.fill.start_color, 'rgb'):
        color = str(cell.fill.start_color.rgb).upper()
        return color == 'FFFF0000'
    
    return False


def extraer_asignaciones_y_restricciones_v2():
    """Extracción confiable de asignaciones y restricciones"""
    
    print("="*80)
    print("EXTRACTOR CONFIABLE V2")
    print("="*80)
    
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encontro el archivo {EXCEL_PATH}")
        return None, None, None
    
    print(f"\nCargando Excel: {EXCEL_PATH}")
    print("Cargando datos de la BD...")
    
    # Cargar datos de BD
    profesores_bd = cargar_todos_profesores()
    cursos_bd = cargar_todos_cursos()
    
    print(f"[OK] Profesores en BD: {len(profesores_bd)}")
    print(f"[OK] Cursos en BD: {len(cursos_bd)}")
    
    # Cargar workbook
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    asignaciones = []
    restricciones = {}
    mapeo_profesores = {}  # hoja -> profesor_id
    nombres_encontrados = defaultdict(set)  # Para estadísticas
    
    # Procesar cada hoja
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"📄 HOJA: {sheet_name}")
        print("="*80)
        
        ws = wb[sheet_name]
        
        # Encontrar fila de encabezados (LUNES, MARTES, etc.)
        fila_encabezado = None
        columnas_dias = {}
        
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    valor_upper = cell.value.strip().upper()
                    if 'LUNES' in valor_upper:
                        fila_encabezado = row_idx
                        break
            if fila_encabezado:
                break
        
        if not fila_encabezado:
            print("  ⚠️  No se encontró fila de encabezados")
            continue
        
        # Mapear columnas a días
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=fila_encabezado, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                valor = cell.value.strip().upper()
                for dia in DIAS_SEMANA:
                    if dia in valor:
                        columnas_dias[col_idx] = dia
                        break
        
        print(f"  📅 Días encontrados: {list(columnas_dias.values())}")
        
        # Analizar celdas para encontrar profesor y asignaciones
        profesor_id_hoja = None
        
        # Procesar filas de horarios
        fila_inicio = fila_encabezado + 1
        restricciones_hoja = {dia: [False] * 17 for dia in DIAS_SEMANA}
        
        for row_idx in range(fila_inicio, min(fila_inicio + 20, ws.max_row + 1)):
            bloque_idx = row_idx - fila_inicio
            
            if bloque_idx >= 17:
                break
            
            for col_idx, dia in columnas_dias.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Verificar si está bloqueada (ROJO)
                if es_celda_bloqueada(cell):
                    restricciones_hoja[dia][bloque_idx] = True
                    print(f"  🚫 Restricción: {dia} bloque {bloque_idx+1} ({BLOQUES_TIEMPO[bloque_idx][0]})")
                
                # Verificar si hay asignación
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 5:
                    contenido = cell.value.strip()
                    
                    # Extraer nombre del profesor de la celda
                    nombre_prof = extraer_nombre_profesor_de_celda(contenido)
                    if nombre_prof:
                        nombres_encontrados[sheet_name].add(nombre_prof)
                        
                        # Si aún no hemos identificado el profesor de esta hoja
                        if not profesor_id_hoja:
                            prof_encontrado = encontrar_profesor_por_nombre_parcial(nombre_prof, profesores_bd)
                            if prof_encontrado:
                                profesor_id_hoja = prof_encontrado['id']
                                mapeo_profesores[sheet_name] = prof_encontrado
                                print(f"  ✅ PROFESOR IDENTIFICADO: {prof_encontrado['nombre_completo']} (ID: {profesor_id_hoja})")
                    
                    # Extraer información del curso
                    tipo, liga = extraer_tipo_liga(contenido)
                    if tipo and liga and profesor_id_hoja:
                        curso_encontrado = encontrar_curso_por_nombre(contenido, cursos_bd)
                        
                        if curso_encontrado:
                            asignacion = {
                                'profesor_id': profesor_id_hoja,
                                'profesor_nombre': mapeo_profesores[sheet_name]['nombre_completo'],
                                'curso_id': curso_encontrado['id'],
                                'curso_nombre': curso_encontrado['nombre'],
                                'curso_codigo': curso_encontrado['codigo'],
                                'tipo': tipo,
                                'liga': liga,
                                'session_type': f"{tipo}{liga}",
                                'dia': dia,
                                'bloque_inicio': bloque_idx,
                                'hora_inicio': BLOQUES_TIEMPO[bloque_idx][0],
                                'contenido_original': contenido
                            }
                            asignaciones.append(asignacion)
                            print(f"    ➜ {curso_encontrado['codigo']} ({tipo}{liga}) - {dia} bloque {bloque_idx+1}")
        
        # Guardar restricciones si se identificó el profesor
        if profesor_id_hoja:
            restricciones[profesor_id_hoja] = restricciones_hoja
    
    # Resumen
    print("\n" + "="*80)
    print("📊 RESUMEN FINAL")
    print("="*80)
    print(f"✅ Hojas procesadas: {len(wb.sheetnames)}")
    print(f"✅ Profesores identificados: {len(mapeo_profesores)}")
    print(f"✅ Asignaciones encontradas: {len(asignaciones)}")
    print(f"✅ Profesores con restricciones: {len(restricciones)}")
    
    # Contar restricciones
    total_bloques_bloqueados = sum(
        sum(sum(bloques) for bloques in dias.values())
        for dias in restricciones.values()
    )
    print(f"🚫 Total bloques bloqueados: {total_bloques_bloqueados}")
    
    # Guardar en JSON
    with open('asignaciones_actuales_v2.json', 'w', encoding='utf-8') as f:
        json.dump({
            'asignaciones': asignaciones,
            'total': len(asignaciones)
        }, f, indent=2, ensure_ascii=False)
    
    with open('restricciones_profesores_v2.json', 'w', encoding='utf-8') as f:
        json.dump({
            'restricciones': {
                str(prof_id): {
                    dia: [int(b) for b in bloques]
                    for dia, bloques in dias.items()
                }
                for prof_id, dias in restricciones.items()
            },
            'total_profesores': len(restricciones),
            'total_bloques_bloqueados': total_bloques_bloqueados
        }, f, indent=2, ensure_ascii=False)
    
    with open('mapeo_profesores_v2.json', 'w', encoding='utf-8') as f:
        json.dump({
            'mapeo': {
                hoja: {
                    'id': prof['id'],
                    'nombre_completo': prof['nombre_completo']
                }
                for hoja, prof in mapeo_profesores.items()
            }
        }, f, indent=2, ensure_ascii=False)
    
    print(f"\n💾 Archivos guardados:")
    print(f"  - asignaciones_actuales_v2.json")
    print(f"  - restricciones_profesores_v2.json")
    print(f"  - mapeo_profesores_v2.json")
    print("="*80)
    
    return asignaciones, restricciones, mapeo_profesores


def insertar_en_bd(asignaciones, restricciones):
    """Inserta asignaciones y restricciones en la BD"""
    
    conn = conectar_db()
    cursor = conn.cursor()
    
    # 1. Limpiar e insertar restricciones
    print("\n" + "="*80)
    print("💾 INSERTANDO RESTRICCIONES EN BD")
    print("="*80)
    
    cursor.execute("DELETE FROM professor_restrictions")
    print("🗑️  Restricciones anteriores eliminadas")
    
    dia_tabla = {
        'LUNES': 'LUNES',
        'MARTES': 'MARTES',
        'MIÉRCOLES': 'MIÉRCOLES',
        'JUEVES': 'JUEVES',
        'VIERNES': 'VIERNES',
        'SÁBADO': 'SÁBADO'
    }
    
    total_rest = 0
    for profesor_id, dias in restricciones.items():
        for dia_key, bloques in dias.items():
            for bloque_idx, bloqueado in enumerate(bloques):
                if bloqueado:
                    inicio, fin = BLOQUES_TIEMPO[bloque_idx]
                    dia_bd = dia_tabla.get(dia_key, dia_key)
                    
                    cursor.execute("""
                        INSERT INTO professor_restrictions 
                        (professor_id, day, start_time, end_time, duration_blocks, reason)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        profesor_id,
                        dia_bd,
                        inicio,
                        fin,
                        1,
                        'No disponible - horario actual 2025-20'
                    ))
                    total_rest += 1
    
    print(f"✅ {total_rest} restricciones insertadas")
    
    # 2. Limpiar e insertar asignaciones históricas
    print("\n" + "="*80)
    print("💾 INSERTANDO ASIGNACIONES HISTÓRICAS EN BD")
    print("="*80)
    
    cursor.execute("DELETE FROM professor_course_history")
    print("🗑️  Asignaciones anteriores eliminadas")
    
    total_asig = 0
    for asig in asignaciones:
        try:
            cursor.execute("""
                INSERT INTO professor_course_history 
                (professor_id, course_id, semestre, veces_asignado)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE veces_asignado = veces_asignado + 1
            """, (asig['profesor_id'], asig['curso_id'], '2025-20', 1))
            total_asig += 1
        except Exception as e:
            print(f"  ⚠️  Error: {e}")
    
    print(f"✅ {total_asig} asignaciones históricas insertadas")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    print("="*80)


if __name__ == "__main__":
    asignaciones, restricciones, mapeo = extraer_asignaciones_y_restricciones_v2()
    
    if asignaciones and restricciones:
        insertar_en_bd(asignaciones, restricciones)
