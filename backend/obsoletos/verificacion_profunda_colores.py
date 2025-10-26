"""
VERIFICACIÓN PROFUNDA DE COLORES Y ASIGNACIONES
===============================================

Este script hace un análisis exhaustivo para verificar:
1. Si los colores realmente identifican únicamente a un curso
2. Casos donde colores se repiten pero son diferentes cursos
3. Matching de nombres abreviados vs nombres completos en BD
4. Cobertura de cursos: ¿todos los cursos tienen profesores asignados?
"""

import mysql.connector
from openpyxl import load_workbook
from collections import defaultdict
import json

def conectar_db():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='sistemas',
        database='upao_timetabling'
    )

def obtener_color_celda(cell):
    """Extrae el color de una celda"""
    if not cell.fill or not cell.fill.start_color:
        return '00000000'
    if hasattr(cell.fill.start_color, 'rgb'):
        return str(cell.fill.start_color.rgb).upper()
    return '00000000'

def limpiar_nombre_curso(nombre):
    """Limpia y normaliza nombre de curso"""
    # Remover paréntesis con tipo de sesión
    import re
    nombre = re.sub(r'\([TP][12]\)', '', nombre)
    nombre = re.sub(r'\([LP][12]\)', '', nombre)
    nombre = re.sub(r'\d+', '', nombre)  # Remover números
    nombre = nombre.replace('\n', ' ')
    nombre = ' '.join(nombre.split())  # Normalizar espacios
    return nombre.strip().upper()

def analizar_colores_global():
    """Analiza TODOS los colores en TODAS las hojas para detectar conflictos"""
    
    print("=" * 100)
    print("ANÁLISIS GLOBAL DE COLORES EN EL EXCEL")
    print("=" * 100)
    
    excel_path = r'..\inputs\Horario_Docentes(2025-20).xlsx'
    print(f"\nCargando: {excel_path}")
    wb = load_workbook(excel_path, data_only=False)
    print(f"Total hojas: {len(wb.sheetnames)}")
    
    # Estructura: color -> [(hoja, curso_nombre)]
    colores_global = defaultdict(list)
    
    COLORES_IGNORAR = ['00000000', 'FFF2F2F2', 'FFFFFFFF', 'FFFF0000']
    
    hojas_procesadas = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Buscar fila de encabezado (buscar "LUNES" en cualquier parte)
        fila_encabezado = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and 'LUNES' in str(cell_value).upper():
                    fila_encabezado = row_idx
                    break
            if fila_encabezado:
                break
        
        if not fila_encabezado:
            continue
        
        hojas_procesadas += 1
        
        # Mapear columnas a días
        columnas_dias = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(fila_encabezado, col).value
            if header:
                dia = str(header).strip().upper()
                if dia in ['LUNES', 'MARTES', 'MIÉRCOLES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'SABADO']:
                    columnas_dias[col] = dia
        
        if not columnas_dias:
            continue
        
        # Escanear todas las celdas con contenido
        max_row_scan = min(fila_encabezado + 20, ws.max_row + 1)
        for row_idx in range(fila_encabezado + 1, max_row_scan):
            for col in columnas_dias.keys():
                cell = ws.cell(row_idx, col)
                
                if cell.value:
                    contenido = str(cell.value).strip()
                    if contenido and len(contenido) > 3:
                        color = obtener_color_celda(cell)
                        
                        if color not in COLORES_IGNORAR:
                            nombre_curso = limpiar_nombre_curso(contenido)
                            colores_global[color].append((sheet_name, nombre_curso))
    
    print(f"Hojas procesadas: {hojas_procesadas}")
    
    # Analizar conflictos
    print("\n[1] COLORES Y SUS CURSOS ASOCIADOS:")
    print("-" * 100)
    
    conflictos = []
    colores_unicos = {}
    
    for color, apariciones in sorted(colores_global.items()):
        # Obtener nombres únicos de cursos para este color
        cursos_unicos = set(curso for _, curso in apariciones)
        
        print(f"\nColor: {color}")
        print(f"  Total apariciones: {len(apariciones)}")
        print(f"  Hojas: {len(set(hoja for hoja, _ in apariciones))}")
        print(f"  Cursos únicos: {len(cursos_unicos)}")
        
        if len(cursos_unicos) > 1:
            print(f"  [!] CONFLICTO: Este color tiene {len(cursos_unicos)} cursos diferentes:")
            for curso in cursos_unicos:
                hojas_con_curso = [hoja for hoja, c in apariciones if c == curso]
                print(f"    - {curso[:60]:60} (en {len(hojas_con_curso)} hojas)")
            conflictos.append({
                'color': color,
                'cursos': list(cursos_unicos),
                'apariciones': len(apariciones)
            })
        else:
            curso = list(cursos_unicos)[0]
            print(f"  [OK] Curso unico: {curso[:70]}")
            colores_unicos[color] = curso
    
    print("\n" + "=" * 100)
    print(f"RESUMEN DE CONFLICTOS:")
    print(f"  Colores totales analizados: {len(colores_global)}")
    print(f"  Colores con curso unico: {len(colores_unicos)}")
    print(f"  Colores con conflicto: {len(conflictos)}")
    
    if conflictos:
        print("\n[!] COLORES CONFLICTIVOS:")
        for conf in conflictos:
            print(f"  {conf['color']}: {conf['cursos']}")
    
    return colores_global, conflictos, colores_unicos

def analizar_cobertura_cursos():
    """Verifica qué cursos de la BD tienen profesores asignados"""
    
    print("\n" + "=" * 100)
    print("ANÁLISIS DE COBERTURA: CURSOS CON/SIN PROFESORES ASIGNADOS")
    print("=" * 100)
    
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    
    # Obtener todos los cursos
    cursor.execute("""
        SELECT id, codigo, nombre 
        FROM courses 
        ORDER BY codigo
    """)
    todos_cursos = cursor.fetchall()
    
    # Obtener cursos con profesores asignados
    cursor.execute("""
        SELECT DISTINCT c.id, c.codigo, c.nombre
        FROM courses c
        JOIN professor_course_history h ON c.id = h.course_id
        ORDER BY c.codigo
    """)
    cursos_con_profesor = cursor.fetchall()
    
    cursos_con_prof_ids = set(c['id'] for c in cursos_con_profesor)
    cursos_sin_profesor = [c for c in todos_cursos if c['id'] not in cursos_con_prof_ids]
    
    print(f"\n[1] ESTADÍSTICAS:")
    print(f"  Total cursos en BD: {len(todos_cursos)}")
    print(f"  Cursos CON profesor asignado: {len(cursos_con_profesor)} ({len(cursos_con_profesor)*100/len(todos_cursos):.1f}%)")
    print(f"  Cursos SIN profesor asignado: {len(cursos_sin_profesor)} ({len(cursos_sin_profesor)*100/len(todos_cursos):.1f}%)")
    
    print(f"\n[2] CURSOS CON PROFESORES ASIGNADOS:")
    print("-" * 100)
    for curso in cursos_con_profesor:
        # Contar profesores
        cursor.execute("""
            SELECT COUNT(DISTINCT professor_id) as num_profesores
            FROM professor_course_history
            WHERE course_id = %s
        """, (curso['id'],))
        num_prof = cursor.fetchone()['num_profesores']
        print(f"  {curso['codigo']:15} {curso['nombre'][:60]:60} ({num_prof} prof)")
    
    print(f"\n[3] CURSOS SIN PROFESORES ASIGNADOS:")
    print("-" * 100)
    for curso in cursos_sin_profesor:
        print(f"  {curso['codigo']:15} {curso['nombre'][:60]}")
    
    cursor.close()
    conn.close()
    
    return cursos_sin_profesor

def analizar_nombres_abreviados():
    """Analiza casos donde nombres en Excel no coinciden con BD"""
    
    print("\n" + "=" * 100)
    print("ANÁLISIS DE NOMBRES ABREVIADOS VS BD")
    print("=" * 100)
    
    # Cargar asignaciones V3
    try:
        with open('asignaciones_v3.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            asignaciones = data['asignaciones']
    except FileNotFoundError:
        print("⚠️  No se encontró asignaciones_v3.json")
        return
    
    # Agrupar por contenido original
    nombres_excel = defaultdict(set)
    for asig in asignaciones:
        contenido = asig['contenido_original']
        nombre_limpio = limpiar_nombre_curso(contenido)
        nombres_excel[nombre_limpio].add((asig['curso_codigo'], asig['curso_nombre']))
    
    print("\n[1] MAPEO DE NOMBRES ABREVIADOS:")
    print("-" * 100)
    
    casos_especiales = []
    
    for nombre_excel, cursos_bd in sorted(nombres_excel.items()):
        if len(nombre_excel) < 10:  # Nombre muy corto, probablemente abreviado
            print(f"\nExcel: '{nombre_excel}'")
            for codigo, nombre_bd in cursos_bd:
                print(f"  -> BD: {codigo:15} {nombre_bd}")
                if nombre_excel != nombre_bd[:len(nombre_excel)]:
                    casos_especiales.append({
                        'excel': nombre_excel,
                        'bd_codigo': codigo,
                        'bd_nombre': nombre_bd
                    })
    
    print(f"\n[2] CASOS ESPECIALES DETECTADOS: {len(casos_especiales)}")
    print("-" * 100)
    for caso in casos_especiales:
        print(f"  Excel: '{caso['excel']}'")
        print(f"    BD: {caso['bd_codigo']} - {caso['bd_nombre']}")
        print()
    
    return casos_especiales

def crear_mapeo_manual_sugerido(casos_especiales):
    """Crea un archivo con mapeo manual sugerido"""
    
    print("\n" + "=" * 100)
    print("GENERANDO MAPEO MANUAL SUGERIDO")
    print("=" * 100)
    
    # Conectar a BD para obtener todos los cursos
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT codigo, nombre FROM courses ORDER BY codigo")
    todos_cursos = cursor.fetchall()
    cursor.close()
    conn.close()
    
    # Crear mapeo sugerido
    mapeo_sugerido = {
        "POO": "PROGRAM ORIENTADA A OBJETOS",
        "BD": "BASE DE DATOS",
        "SIST. GESTION BASE DATOS": "SISTEM GESTION BASE DE DATOS",
        "ARQUITECTURA DE SISTEMAS": "ARQUITECTURA DE SISTEMAS",
        # Agregar más según análisis
    }
    
    with open('mapeo_nombres_cursos.json', 'w', encoding='utf-8') as f:
        json.dump({
            'mapeo_manual': mapeo_sugerido,
            'casos_especiales': casos_especiales,
            'todos_cursos_bd': [{'codigo': c['codigo'], 'nombre': c['nombre']} for c in todos_cursos]
        }, f, indent=2, ensure_ascii=False)
    
    print("[OK] Archivo 'mapeo_nombres_cursos.json' creado")
    print("  Contiene:")
    print(f"    - {len(mapeo_sugerido)} mapeos manuales sugeridos")
    print(f"    - {len(casos_especiales)} casos especiales detectados")
    print(f"    - {len(todos_cursos)} cursos de la BD para referencia")

if __name__ == "__main__":
    print("\n")
    print("=" * 100)
    print("=" + " " * 98 + "=")
    print("=" + " " * 20 + "VERIFICACION PROFUNDA DE COLORES Y ASIGNACIONES" + " " * 20 + "=")
    print("=" + " " * 98 + "=")
    print("=" * 100)
    
    # 1. Analizar colores globalmente
    colores_global, conflictos, colores_unicos = analizar_colores_global()
    
    # 2. Verificar cobertura de cursos
    cursos_sin_profesor = analizar_cobertura_cursos()
    
    # 3. Analizar nombres abreviados
    casos_especiales = analizar_nombres_abreviados()
    
    # 4. Crear mapeo manual sugerido
    crear_mapeo_manual_sugerido(casos_especiales)
    
    print("\n" + "=" * 100)
    print("CONCLUSIONES Y RECOMENDACIONES:")
    print("=" * 100)
    
    if conflictos:
        print("\n[!] COLORES CONFLICTIVOS DETECTADOS:")
        print("  Hay colores que se usan para diferentes cursos.")
        print("  Solucion: Usar color + hoja del profesor como identificador compuesto")
    else:
        print("\n[OK] No se detectaron conflictos de colores")
        print("  Cada color identifica unicamente a un curso")
    
    if cursos_sin_profesor:
        print(f"\n[!] HAY {len(cursos_sin_profesor)} CURSOS SIN PROFESOR:")
        print("  Estos cursos no tienen asignaciones historicas en el Excel")
        print("  Solucion: ACO debera asignar profesores basandose en otros criterios")
    else:
        print("\n[OK] Todos los cursos tienen al menos un profesor asignado")
    
    if casos_especiales:
        print(f"\n[!] HAY {len(casos_especiales)} CASOS DE NOMBRES ABREVIADOS:")
        print("  Algunos nombres en Excel no coinciden exactamente con BD")
        print("  Solucion: Usar mapeo manual en 'mapeo_nombres_cursos.json'")
    else:
        print("\n[OK] Todos los nombres coinciden correctamente")
    
    print("\n" + "=" * 100)
    print("FIN DEL ANÁLISIS")
    print("=" * 100)
