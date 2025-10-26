"""
================================================================================
EXTRACTOR DEFINITIVO V3 - USANDO COLORES PARA IDENTIFICAR CURSOS
================================================================================
Estrategia correcta:
1. Agrupar celdas por COLOR DE FONDO (cada color = un curso)
2. Extraer nombre del curso de la primera celda de ese color
3. Extraer tipo (T/P/L) y liga de cada celda
4. Asignar profesor identificado en las celdas
================================================================================
"""

import json
import re
from pathlib import Path
import mysql.connector
from openpyxl import load_workbook
from collections import defaultdict

# Configuracion
EXCEL_PATH = Path('../inputs/Horario_Docentes(2025-20).xlsx')
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'sistemas',
    'database': 'upao_timetabling'
}

DIAS_SEMANA = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']

# Bloques horarios UPAO
BLOQUES_TIEMPO = [
    ("07:00:00", "07:50:00"), ("07:55:00", "08:45:00"), ("08:50:00", "09:40:00"),
    ("09:45:00", "10:35:00"), ("10:40:00", "11:30:00"), ("11:35:00", "12:25:00"),
    ("12:30:00", "13:20:00"), ("13:25:00", "14:15:00"), ("14:20:00", "15:10:00"),
    ("15:15:00", "16:05:00"), ("16:10:00", "17:00:00"), ("17:05:00", "17:55:00"),
    ("18:00:00", "18:50:00"), ("18:55:00", "19:45:00"), ("19:50:00", "20:40:00"),
    ("20:45:00", "21:35:00"), ("21:40:00", "22:30:00")
]

# Colores que NO son cursos (grises, blancos, rojos)
COLORES_IGNORAR = [
    '00000000',  # Negro/sin color
    'FFF2F2F2',  # Gris claro (encabezados)
    'FFFFFFFF',  # Blanco
    'FFFF0000',  # Rojo (bloqueado)
]


def conectar_db():
    return mysql.connector.connect(**DB_CONFIG)


def cargar_todos_profesores():
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, nombre_completo, nombres, apellidos FROM professors ORDER BY id")
    profesores = cursor.fetchall()
    cursor.close()
    conn.close()
    return profesores


def cargar_todos_cursos():
    conn = conectar_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, codigo, nombre FROM courses ORDER BY id")
    cursos = cursor.fetchall()
    cursor.close()
    conn.close()
    return cursos


def obtener_color_celda(cell):
    """Obtiene el color RGB de una celda"""
    if not cell.fill or not cell.fill.start_color:
        return '00000000'
    
    if hasattr(cell.fill.start_color, 'rgb'):
        return str(cell.fill.start_color.rgb).upper()
    
    return '00000000'


def es_color_valido(color):
    """Verifica si el color representa un curso (no es gris, blanco o rojo)"""
    return color not in COLORES_IGNORAR


def es_celda_bloqueada(cell):
    """Verifica si es una celda bloqueada (roja)"""
    return obtener_color_celda(cell) == 'FFFF0000'


def extraer_nombre_profesor_de_celda(contenido):
    """Extrae nombre del profesor: '... - Armando' -> 'Armando'"""
    if not contenido or not isinstance(contenido, str):
        return None
    
    match = re.search(r'-\s*([A-Za-zÁÉÍÓÚáéíóúÑñ\s]+?)(?:\s|$)', contenido)
    if match:
        nombre = match.group(1).strip()
        if nombre.upper() not in ['VIRTUAL', 'PRESENCIAL', 'NPR', 'PRS']:
            return nombre
    return None


def extraer_tipo_liga(texto):
    """Extrae tipo (T/P/L) y liga (1,2,3) de un texto"""
    if not texto:
        return None, None
    
    match = re.search(r'\(([TPL])(\d+)\)', str(texto).upper())
    if match:
        return match.group(1), int(match.group(2))
    return None, None


def normalizar_nombre_curso(nombre):
    """Normaliza nombre de curso para búsqueda"""
    if not nombre:
        return ""
    
    # Remover paréntesis y su contenido
    nombre = re.sub(r'\([^)]*\)', '', nombre)
    # Remover números de 4+ dígitos
    nombre = re.sub(r'\d{4,}', '', nombre)
    # Remover guiones y símbolos
    nombre = re.sub(r'[-_]', ' ', nombre)
    # Limpiar espacios múltiples
    nombre = re.sub(r'\s+', ' ', nombre)
    
    return nombre.strip().upper()


def encontrar_curso_por_nombre(nombre_curso, cursos):
    """Encuentra un curso en BD por coincidencia de palabras"""
    if not nombre_curso:
        return None
    
    nombre_normalizado = normalizar_nombre_curso(nombre_curso)
    palabras_busqueda = set(nombre_normalizado.split())
    
    # Filtrar palabras muy cortas
    palabras_busqueda = {p for p in palabras_busqueda if len(p) > 2}
    
    mejores_matches = []
    for curso in cursos:
        nombre_bd = curso['nombre'].upper()
        codigo_bd = curso['codigo'].upper()
        
        # Palabras del curso en BD
        palabras_curso = set(nombre_bd.split())
        
        # Calcular coincidencias
        coincidencias = len(palabras_busqueda & palabras_curso)
        
        if coincidencias >= 2:  # Al menos 2 palabras coinciden
            mejores_matches.append((curso, coincidencias, len(palabras_curso)))
    
    if mejores_matches:
        # Ordenar por: más coincidencias, menos palabras totales (más específico)
        mejores_matches.sort(key=lambda x: (x[1], -x[2]), reverse=True)
        return mejores_matches[0][0]
    
    return None


def encontrar_profesor_por_nombre_parcial(nombre_parcial, profesores):
    """Encuentra profesor por coincidencia parcial"""
    if not nombre_parcial:
        return None
    
    nombre_limpio = nombre_parcial.strip().upper()
    
    # Búsqueda en nombre completo
    for prof in profesores:
        nombre_completo = (prof['nombre_completo'] or '').upper()
        apellidos = (prof['apellidos'] or '').upper()
        
        if nombre_limpio in nombre_completo or nombre_limpio in apellidos:
            return prof
    
    # Búsqueda por palabras
    palabras = nombre_limpio.split()
    for palabra in palabras:
        if len(palabra) > 3:
            for prof in profesores:
                nombre_completo = (prof['nombre_completo'] or '').upper()
                if palabra in nombre_completo:
                    return prof
    
    return None


def extraer_asignaciones_por_colores():
    """Extracción definitiva usando colores como identificador de cursos"""
    
    print("="*80)
    print("EXTRACTOR DEFINITIVO V3 - POR COLORES")
    print("="*80)
    
    if not EXCEL_PATH.exists():
        print(f"ERROR: No se encontro {EXCEL_PATH}")
        return None, None, None
    
    print(f"\nCargando Excel: {EXCEL_PATH}")
    print("Cargando datos de BD...")
    
    profesores_bd = cargar_todos_profesores()
    cursos_bd = cargar_todos_cursos()
    
    print(f"[OK] Profesores: {len(profesores_bd)}")
    print(f"[OK] Cursos: {len(cursos_bd)}")
    
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    asignaciones = []
    restricciones = {}
    mapeo_profesores = {}
    estadisticas = {
        'hojas_procesadas': 0,
        'profesores_identificados': 0,
        'cursos_identificados': 0,
        'asignaciones_creadas': 0,
        'restricciones_encontradas': 0
    }
    
    # Procesar cada hoja
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"HOJA: {sheet_name}")
        print("="*80)
        
        ws = wb[sheet_name]
        estadisticas['hojas_procesadas'] += 1
        
        # Encontrar fila de encabezados
        fila_encabezado = None
        columnas_dias = {}
        
        for row_idx in range(1, 10):
            for col_idx in range(1, 15):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and 'LUNES' in str(cell.value).upper():
                    fila_encabezado = row_idx
                    break
            if fila_encabezado:
                break
        
        if not fila_encabezado:
            print("  [WARN] No se encontro fila de encabezados")
            continue
        
        # Mapear columnas a días
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=fila_encabezado, column=col_idx)
            if cell.value:
                valor = str(cell.value).strip().upper()
                for dia in DIAS_SEMANA:
                    if dia in valor:
                        columnas_dias[col_idx] = dia
                        break
        
        print(f"  Dias: {list(columnas_dias.values())}")
        
        # Agrupar celdas por color
        cursos_por_color = defaultdict(list)
        profesor_id_hoja = None
        restricciones_hoja = {dia: [False] * 17 for dia in DIAS_SEMANA}
        
        for row_idx in range(fila_encabezado + 1, min(fila_encabezado + 20, ws.max_row + 1)):
            bloque_idx = row_idx - fila_encabezado - 1
            
            if bloque_idx >= 17:
                break
            
            for col_idx, dia in columnas_dias.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                color = obtener_color_celda(cell)
                
                # Verificar si está bloqueada
                if es_celda_bloqueada(cell):
                    restricciones_hoja[dia][bloque_idx] = True
                    estadisticas['restricciones_encontradas'] += 1
                    print(f"  [BLOCK] {dia} bloque {bloque_idx+1}")
                    continue
                
                # Si tiene contenido y color válido
                if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 5:
                    if es_color_valido(color):
                        contenido = cell.value.strip()
                        
                        # Identificar profesor si aún no se ha hecho
                        if not profesor_id_hoja:
                            nombre_prof = extraer_nombre_profesor_de_celda(contenido)
                            if nombre_prof:
                                prof_encontrado = encontrar_profesor_por_nombre_parcial(nombre_prof, profesores_bd)
                                if prof_encontrado:
                                    profesor_id_hoja = prof_encontrado['id']
                                    mapeo_profesores[sheet_name] = prof_encontrado
                                    estadisticas['profesores_identificados'] += 1
                                    print(f"  [PROFESOR] {prof_encontrado['nombre_completo']} (ID: {profesor_id_hoja})")
                        
                        # Guardar celda agrupada por color
                        cursos_por_color[color].append({
                            'contenido': contenido,
                            'dia': dia,
                            'bloque': bloque_idx,
                            'row': row_idx,
                            'col': col_idx
                        })
        
        # Procesar cada color (cada curso)
        if profesor_id_hoja:
            restricciones[profesor_id_hoja] = restricciones_hoja
            
            for color, celdas in cursos_por_color.items():
                # Primera celda tiene el nombre del curso
                nombre_curso_raw = celdas[0]['contenido']
                curso_encontrado = encontrar_curso_por_nombre(nombre_curso_raw, cursos_bd)
                
                if curso_encontrado:
                    estadisticas['cursos_identificados'] += 1
                    print(f"  [CURSO] {curso_encontrado['codigo']} - {curso_encontrado['nombre'][:40]}")
                    
                    # Crear asignación para cada celda de este color
                    for celda in celdas:
                        tipo, liga = extraer_tipo_liga(celda['contenido'])
                        
                        if tipo and liga:
                            asignacion = {
                                'profesor_id': profesor_id_hoja,
                                'profesor_nombre': mapeo_profesores[sheet_name]['nombre_completo'],
                                'curso_id': curso_encontrado['id'],
                                'curso_nombre': curso_encontrado['nombre'],
                                'curso_codigo': curso_encontrado['codigo'],
                                'tipo': tipo,
                                'liga': liga,
                                'session_type': f"{tipo}{liga}",
                                'dia': celda['dia'],
                                'bloque_inicio': celda['bloque'],
                                'hora_inicio': BLOQUES_TIEMPO[celda['bloque']][0],
                                'contenido_original': celda['contenido'],
                                'color': color
                            }
                            asignaciones.append(asignacion)
                            estadisticas['asignaciones_creadas'] += 1
                            print(f"    -> {tipo}{liga} - {celda['dia']} bloque {celda['bloque']+1}")
                else:
                    print(f"  [WARN] Curso no encontrado para color {color}: {nombre_curso_raw[:50]}")
    
    # Resumen
    print("\n" + "="*80)
    print("RESUMEN FINAL")
    print("="*80)
    for k, v in estadisticas.items():
        print(f"  {k}: {v}")
    
    # Guardar JSONs
    with open('asignaciones_v3.json', 'w', encoding='utf-8') as f:
        json.dump({'asignaciones': asignaciones, 'total': len(asignaciones)}, f, indent=2, ensure_ascii=False)
    
    with open('restricciones_v3.json', 'w', encoding='utf-8') as f:
        json.dump({
            'restricciones': {
                str(prof_id): {dia: [int(b) for b in bloques] for dia, bloques in dias.items()}
                for prof_id, dias in restricciones.items()
            },
            'total_profesores': len(restricciones),
            'total_bloques': estadisticas['restricciones_encontradas']
        }, f, indent=2, ensure_ascii=False)
    
    with open('mapeo_profesores_v3.json', 'w', encoding='utf-8') as f:
        json.dump({
            'mapeo': {
                hoja: {'id': prof['id'], 'nombre_completo': prof['nombre_completo']}
                for hoja, prof in mapeo_profesores.items()
            }
        }, f, indent=2, ensure_ascii=False)
    
    print("\n[SAVE] Archivos guardados:")
    print("  - asignaciones_v3.json")
    print("  - restricciones_v3.json")
    print("  - mapeo_profesores_v3.json")
    print("="*80)
    
    return asignaciones, restricciones, mapeo_profesores


def insertar_en_bd(asignaciones, restricciones):
    """Inserta en BD"""
    conn = conectar_db()
    cursor = conn.cursor()
    
    # Restricciones
    print("\n" + "="*80)
    print("INSERTANDO EN BD")
    print("="*80)
    
    cursor.execute("DELETE FROM professor_restrictions")
    cursor.execute("DELETE FROM professor_course_history")
    
    dia_tabla = {
        'LUNES': 'LUNES', 'MARTES': 'MARTES', 'MIERCOLES': 'MIÉRCOLES',
        'JUEVES': 'JUEVES', 'VIERNES': 'VIERNES', 'SABADO': 'SÁBADO'
    }
    
    total_rest = 0
    for profesor_id, dias in restricciones.items():
        for dia_key, bloques in dias.items():
            for bloque_idx, bloqueado in enumerate(bloques):
                if bloqueado:
                    inicio, fin = BLOQUES_TIEMPO[bloque_idx]
                    cursor.execute("""
                        INSERT INTO professor_restrictions 
                        (professor_id, day, start_time, end_time, duration_blocks, reason)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (profesor_id, dia_tabla.get(dia_key, dia_key), inicio, fin, 1, 'No disponible 2025-20'))
                    total_rest += 1
    
    print(f"[OK] {total_rest} restricciones insertadas")
    
    # Asignaciones
    # Agrupar por profesor+curso para evitar duplicados
    asignaciones_unicas = {}
    for asig in asignaciones:
        key = (asig['profesor_id'], asig['curso_id'])
        if key not in asignaciones_unicas:
            asignaciones_unicas[key] = asig
    
    total_asig = 0
    for asig in asignaciones_unicas.values():
        try:
            cursor.execute("""
                INSERT INTO professor_course_history 
                (professor_id, course_id, semestre, veces_asignado)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE veces_asignado = veces_asignado + 1
            """, (asig['profesor_id'], asig['curso_id'], '2025-20', 1))
            total_asig += 1
        except Exception as e:
            print(f"[ERROR] {e}")
    
    print(f"[OK] {total_asig} asignaciones historicas insertadas")
    
    conn.commit()
    cursor.close()
    conn.close()
    print("="*80)


if __name__ == "__main__":
    asignaciones, restricciones, mapeo = extraer_asignaciones_por_colores()
    
    if asignaciones and restricciones:
        insertar_en_bd(asignaciones, restricciones)
