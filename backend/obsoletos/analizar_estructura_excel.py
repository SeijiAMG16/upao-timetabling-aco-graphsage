"""
Análisis detallado de la estructura del Excel para entender:
1. Dónde está el nombre del profesor dentro de cada hoja
2. Cómo identificar exactamente las celdas bloqueadas (rojas)
3. Estructura de las filas de horarios
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

EXCEL_PATH = '../inputs/Horario_Docentes(2025-20).xlsx'

wb = load_workbook(EXCEL_PATH, data_only=False)

print("="*80)
print("ANÁLISIS DETALLADO DEL EXCEL")
print("="*80)

# Analizar las primeras 3 hojas en detalle
for sheet_name in wb.sheetnames[:3]:
    print(f"\n{'='*80}")
    print(f"HOJA: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    
    # Analizar las primeras 15 filas
    print("\n📋 PRIMERAS 15 FILAS:")
    for row_idx in range(1, min(16, ws.max_row + 1)):
        print(f"\nFila {row_idx}:")
        for col_idx in range(1, min(8, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            valor = cell.value
            
            # Analizar color de fondo
            color_info = "Sin color"
            if cell.fill and cell.fill.start_color:
                if hasattr(cell.fill.start_color, 'rgb'):
                    color = cell.fill.start_color.rgb
                elif hasattr(cell.fill.start_color, 'index'):
                    color = f"Index:{cell.fill.start_color.index}"
                else:
                    color = str(cell.fill.start_color)
                
                if color and color != '00000000' and color != 'Index:64':
                    color_info = f"Color:{color}"
            
            if valor or color_info != "Sin color":
                print(f"  Col {col_idx}: '{valor}' | {color_info}")
    
    print(f"\n{'='*80}")
    input("Presiona Enter para continuar con la siguiente hoja...")
