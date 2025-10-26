#!/usr/bin/env python3
"""
EXTRACTOR DE CURSOS V1 - UPAO TIMETABLING
=============================================

Este script extrae la información REAL de cursos del Excel:
- Ciclo (C1->1, C2->2, etc.)
- Asignatura (nombre del curso)
- Modalidad (PRS/NPR)
- Alumnos por tipo (Teoría, Práctica, Laboratorio)
- Grupos por tipo (Teoría, Práctica, Laboratorio)

Autor: Sistema UPAO
Fecha: 2025-10-10
"""

import openpyxl
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ExtractorCursosV1:
    """Extractor de cursos desde Excel UPAO"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.cursos_extraidos = []
        
        # Mapeo de ciclos
        self.mapeo_ciclos = {
            'C1': 1, 'C2': 2, 'C3': 3, 'C4': 4, 'C5': 5,
            'C6': 6, 'C7': 7, 'C8': 8, 'C9': 9, 'C10': 10
        }
        
        # Mapeo de modalidades
        self.mapeo_modalidades = {
            'PRS': 'presencial',
            'NPR': 'no_presencial',
            'PRESENCIAL': 'presencial',
            'NO PRESENCIAL': 'no_presencial'
        }
    
    def abrir_excel(self):
        """Abrir el archivo Excel"""
        try:
            logger.info(f"Abriendo Excel: {self.excel_path}")
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            logger.info(f"Excel abierto correctamente. Hojas: {self.workbook.sheetnames}")
            return True
        except Exception as e:
            logger.error(f"Error al abrir Excel: {e}")
            return False
    
    def buscar_hoja_cursos(self) -> Optional[str]:
        """Buscar la hoja que contenga información de cursos"""
        if not self.workbook:
            return None
            
        # Posibles nombres de hojas con cursos
        nombres_posibles = [
            'CURSOS', 'Cursos', 'cursos',
            'PROYECCIONES', 'Proyecciones', 'proyecciones',
            'PLAN', 'Plan', 'plan',
            'CURRICULUM', 'Curriculum', 'curriculum'
        ]
        
        for nombre in nombres_posibles:
            if nombre in self.workbook.sheetnames:
                logger.info(f"Hoja de cursos encontrada: {nombre}")
                return nombre
        
        # Si no encuentra por nombre, buscar por contenido
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Buscar palabras clave en las primeras filas
            for row in range(1, 10):
                for col in range(1, 20):
                    try:
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_upper = cell_value.upper()
                            if any(keyword in cell_upper for keyword in ['ASIGNATURA', 'CURSO', 'CICLO', 'MODALIDAD']):
                                logger.info(f"Hoja de cursos detectada por contenido: {sheet_name}")
                                return sheet_name
                    except:
                        continue
        
        logger.warning("No se encontró hoja de cursos específica")
        return None
    
    def detectar_estructura_cursos(self, sheet_name: str) -> Dict:
        """Detectar la estructura de la hoja de cursos"""
        sheet = self.workbook[sheet_name]
        estructura = {
            'fila_headers': None,
            'columnas': {},
            'rango_datos': None
        }
        
        # Buscar fila de headers
        for row in range(1, 15):
            headers_encontrados = 0
            columnas_temp = {}
            
            for col in range(1, 30):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_upper = cell_value.upper().strip()
                        
                        # Mapear columnas importantes
                        if 'CICLO' in cell_upper:
                            columnas_temp['ciclo'] = col
                            headers_encontrados += 1
                        elif 'ASIGNATURA' in cell_upper or 'CURSO' in cell_upper:
                            columnas_temp['nombre'] = col
                            headers_encontrados += 1
                        elif 'MODALIDAD' in cell_upper or 'PRESENCIAL' in cell_upper:
                            columnas_temp['modalidad'] = col
                            headers_encontrados += 1
                        elif 'ALUMNOS' in cell_upper and 'TEORIA' in cell_upper:
                            columnas_temp['alumnos_teoria'] = col
                            headers_encontrados += 1
                        elif 'ALUMNOS' in cell_upper and 'PRACTICA' in cell_upper:
                            columnas_temp['alumnos_practica'] = col
                            headers_encontrados += 1
                        elif 'ALUMNOS' in cell_upper and 'LABORATORIO' in cell_upper:
                            columnas_temp['alumnos_laboratorio'] = col
                            headers_encontrados += 1
                        elif 'GRUPOS' in cell_upper and 'TEORIA' in cell_upper:
                            columnas_temp['grupos_teoria'] = col
                            headers_encontrados += 1
                        elif 'GRUPOS' in cell_upper and 'PRACTICA' in cell_upper:
                            columnas_temp['grupos_practica'] = col
                            headers_encontrados += 1
                        elif 'GRUPOS' in cell_upper and 'LABORATORIO' in cell_upper:
                            columnas_temp['grupos_laboratorio'] = col
                            headers_encontrados += 1
                except:
                    continue
            
            # Si encontramos suficientes headers, esta es la fila
            if headers_encontrados >= 4:
                estructura['fila_headers'] = row
                estructura['columnas'] = columnas_temp
                logger.info(f"Headers encontrados en fila {row}: {list(columnas_temp.keys())}")
                break
        
        # Determinar rango de datos
        if estructura['fila_headers']:
            inicio_datos = estructura['fila_headers'] + 1
            fin_datos = sheet.max_row
            
            # Buscar última fila con datos
            for row in range(inicio_datos, sheet.max_row + 1):
                ciclo_val = sheet.cell(row=row, column=estructura['columnas'].get('ciclo', 1)).value
                nombre_val = sheet.cell(row=row, column=estructura['columnas'].get('nombre', 1)).value
                
                if not ciclo_val and not nombre_val:
                    fin_datos = row - 1
                    break
            
            estructura['rango_datos'] = (inicio_datos, fin_datos)
            logger.info(f"Rango de datos: filas {inicio_datos} a {fin_datos}")
        
        return estructura
    
    def extraer_cursos_de_hoja(self, sheet_name: str) -> List[Dict]:
        """Extraer cursos de una hoja específica"""
        logger.info(f"Extrayendo cursos de hoja: {sheet_name}")
        
        estructura = self.detectar_estructura_cursos(sheet_name)
        if not estructura['fila_headers']:
            logger.error(f"No se pudo detectar estructura en hoja {sheet_name}")
            return []
        
        sheet = self.workbook[sheet_name]
        cursos = []
        
        inicio, fin = estructura['rango_datos']
        columnas = estructura['columnas']
        
        for row in range(inicio, fin + 1):
            try:
                # Extraer datos básicos
                ciclo_raw = sheet.cell(row=row, column=columnas.get('ciclo', 1)).value
                nombre_raw = sheet.cell(row=row, column=columnas.get('nombre', 2)).value
                modalidad_raw = sheet.cell(row=row, column=columnas.get('modalidad', 3)).value
                
                # Saltar filas vacías
                if not ciclo_raw or not nombre_raw:
                    continue
                
                # Procesar ciclo
                ciclo = self.procesar_ciclo(ciclo_raw)
                if not ciclo:
                    continue
                
                # Procesar nombre
                nombre = str(nombre_raw).strip()
                if not nombre or len(nombre) < 3:
                    continue
                
                # Procesar modalidad
                modalidad = self.procesar_modalidad(modalidad_raw)
                
                # Extraer números de alumnos
                alumnos_teoria = self.extraer_numero(sheet.cell(row=row, column=columnas.get('alumnos_teoria', 4)).value)
                alumnos_practica = self.extraer_numero(sheet.cell(row=row, column=columnas.get('alumnos_practica', 5)).value)
                alumnos_laboratorio = self.extraer_numero(sheet.cell(row=row, column=columnas.get('alumnos_laboratorio', 6)).value)
                
                # Extraer números de grupos
                grupos_teoria = self.extraer_numero(sheet.cell(row=row, column=columnas.get('grupos_teoria', 7)).value)
                grupos_practica = self.extraer_numero(sheet.cell(row=row, column=columnas.get('grupos_practica', 8)).value)
                grupos_laboratorio = self.extraer_numero(sheet.cell(row=row, column=columnas.get('grupos_laboratorio', 9)).value)
                
                # Crear código único para el curso
                codigo = self.generar_codigo_curso(ciclo, nombre)
                
                # Determinar si requiere laboratorio/práctica
                requiere_laboratorio = alumnos_laboratorio > 0 or grupos_laboratorio > 0
                requiere_practica = alumnos_practica > 0 or grupos_practica > 0
                
                curso = {
                    'codigo': codigo,
                    'nombre': nombre,
                    'ciclo': ciclo,
                    'modalidad': modalidad,
                    'alumnos_teoria': alumnos_teoria,
                    'alumnos_practica': alumnos_practica,
                    'alumnos_laboratorio': alumnos_laboratorio,
                    'grupos_teoria': grupos_teoria,
                    'grupos_practica': grupos_practica,
                    'grupos_laboratorio': grupos_laboratorio,
                    'requiere_laboratorio': requiere_laboratorio,
                    'requiere_practica': requiere_practica,
                    'creditos': self.calcular_creditos(grupos_teoria, grupos_practica, grupos_laboratorio),
                    'fila_excel': row
                }
                
                cursos.append(curso)
                logger.debug(f"Curso extraído: {codigo} - {nombre} (Ciclo {ciclo})")
                
            except Exception as e:
                logger.error(f"Error procesando fila {row}: {e}")
                continue
        
        logger.info(f"Extraídos {len(cursos)} cursos de la hoja {sheet_name}")
        return cursos
    
    def procesar_ciclo(self, ciclo_raw) -> Optional[int]:
        """Procesar valor de ciclo"""
        if not ciclo_raw:
            return None
        
        ciclo_str = str(ciclo_raw).strip().upper()
        
        # Si ya es número
        if ciclo_str.isdigit():
            return int(ciclo_str)
        
        # Si está en formato C1, C2, etc.
        if ciclo_str in self.mapeo_ciclos:
            return self.mapeo_ciclos[ciclo_str]
        
        # Buscar patrón C + número
        match = re.search(r'C(\d+)', ciclo_str)
        if match:
            return int(match.group(1))
        
        logger.warning(f"No se pudo procesar ciclo: {ciclo_raw}")
        return None
    
    def procesar_modalidad(self, modalidad_raw) -> str:
        """Procesar modalidad"""
        if not modalidad_raw:
            return 'presencial'  # Por defecto
        
        modalidad_str = str(modalidad_raw).strip().upper()
        
        if modalidad_str in self.mapeo_modalidades:
            return self.mapeo_modalidades[modalidad_str]
        
        # Buscar patrones
        if 'NO' in modalidad_str or 'NPR' in modalidad_str:
            return 'no_presencial'
        
        return 'presencial'  # Por defecto
    
    def extraer_numero(self, valor) -> int:
        """Extraer número de una celda"""
        if valor is None:
            return 0
        
        if isinstance(valor, (int, float)):
            return max(0, int(valor))
        
        if isinstance(valor, str):
            # Extraer números de strings
            numeros = re.findall(r'\d+', valor)
            if numeros:
                return int(numeros[0])
        
        return 0
    
    def generar_codigo_curso(self, ciclo: int, nombre: str) -> str:
        """Generar código único para el curso"""
        # Tomar primeras letras del nombre
        palabras = nombre.upper().split()
        iniciales = ''.join([palabra[0] for palabra in palabras if palabra])[:4]
        
        # Formato: C{ciclo}_{iniciales}
        return f"C{ciclo}_{iniciales}"
    
    def calcular_creditos(self, teoria: int, practica: int, laboratorio: int) -> int:
        """Calcular créditos aproximados"""
        # Fórmula aproximada: teoría + práctica/2 + laboratorio/2
        return teoria + (practica // 2) + (laboratorio // 2)
    
    def extraer_todos_cursos(self) -> List[Dict]:
        """Extraer cursos de todas las hojas relevantes"""
        if not self.abrir_excel():
            return []
        
        # Primero buscar hoja específica de cursos
        hoja_cursos = self.buscar_hoja_cursos()
        if hoja_cursos:
            self.cursos_extraidos = self.extraer_cursos_de_hoja(hoja_cursos)
        
        # Si no encontramos cursos, intentar con otras hojas
        if not self.cursos_extraidos:
            logger.info("Buscando cursos en todas las hojas...")
            for sheet_name in self.workbook.sheetnames:
                if sheet_name != hoja_cursos:  # Evitar duplicados
                    cursos_hoja = self.extraer_cursos_de_hoja(sheet_name)
                    if cursos_hoja:
                        self.cursos_extraidos.extend(cursos_hoja)
                        break
        
        self.workbook.close()
        
        # Eliminar duplicados por código
        cursos_unicos = {}
        for curso in self.cursos_extraidos:
            codigo = curso['codigo']
            if codigo not in cursos_unicos:
                cursos_unicos[codigo] = curso
        
        self.cursos_extraidos = list(cursos_unicos.values())
        
        logger.info(f"TOTAL CURSOS EXTRAÍDOS: {len(self.cursos_extraidos)}")
        return self.cursos_extraidos
    
    def mostrar_resumen(self):
        """Mostrar resumen de cursos extraídos"""
        if not self.cursos_extraidos:
            logger.warning("No hay cursos extraídos")
            return
        
        print("\n" + "="*60)
        print("RESUMEN DE CURSOS EXTRAÍDOS")
        print("="*60)
        
        # Agrupar por ciclo
        por_ciclo = {}
        for curso in self.cursos_extraidos:
            ciclo = curso['ciclo']
            if ciclo not in por_ciclo:
                por_ciclo[ciclo] = []
            por_ciclo[ciclo].append(curso)
        
        for ciclo in sorted(por_ciclo.keys()):
            cursos_ciclo = por_ciclo[ciclo]
            print(f"\n📚 CICLO {ciclo}: {len(cursos_ciclo)} cursos")
            
            for curso in cursos_ciclo:
                print(f"  • {curso['codigo']}: {curso['nombre']}")
                print(f"    Modalidad: {curso['modalidad']}")
                print(f"    Alumnos: T={curso['alumnos_teoria']}, P={curso['alumnos_practica']}, L={curso['alumnos_laboratorio']}")
                print(f"    Grupos: T={curso['grupos_teoria']}, P={curso['grupos_practica']}, L={curso['grupos_laboratorio']}")
                if curso['requiere_laboratorio']:
                    print(f"    🔬 Requiere laboratorio")
                if curso['requiere_practica']:
                    print(f"    🛠️ Requiere práctica")
                print()

def main():
    """Función principal de prueba"""
    excel_path = "../inputs/Horario_Docentes(2025-20).xlsx"
    
    extractor = ExtractorCursosV1(excel_path)
    cursos = extractor.extraer_todos_cursos()
    
    if cursos:
        extractor.mostrar_resumen()
        
        # Guardar en JSON para revisión
        import json
        with open('cursos_extraidos.json', 'w', encoding='utf-8') as f:
            json.dump(cursos, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ Cursos guardados en cursos_extraidos.json")
    else:
        print("❌ No se pudieron extraer cursos")

if __name__ == "__main__":
    main()