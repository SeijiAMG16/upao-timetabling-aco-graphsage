#!/usr/bin/env python3
"""
EXTRACTOR DE CURSOS V2 - UPAO TIMETABLING  
=============================================

Extractor especializado para el archivo Libro1.xlsx que contiene
la información estructurada de cursos de ISIA con:
- Ciclo (C1, C2, C3, C4)
- Asignatura
- Modalidad (PRS/NPR)
- Horas de Teoría, Práctica y Laboratorio
- Créditos

Autor: Sistema UPAO
Fecha: 2025-10-10
"""

import openpyxl
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ExtractorCursosV2:
    """Extractor de cursos desde Libro1.xlsx - Información estructurada ISIA"""
    
    def __init__(self, excel_path: str = "../inputs/Libro1.xlsx"):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.cursos_extraidos = []
        
        # Mapeo de ciclos
        self.mapeo_ciclos = {
            'C1': 1, 'C2': 2, 'C3': 3, 'C4': 4, 'C5': 5,
            'C6': 6, 'C7': 7, 'C8': 8, 'C9': 9, 'C10': 10
        }
        
        # Mapeo de modalidades
        self.mapeo_modalidades = {
            'PRS': 'presencial',
            'NPR': 'no_presencial'
        }
        
        # Estructura conocida del Excel
        self.columnas = {
            'ciclo': 3,        # Columna C
            'asignatura': 6,   # Columna F
            'creditos': 7,     # Columna G
            'ht': 8,           # Columna H (Horas Teoría)
            'hp': 9,           # Columna I (Horas Práctica)
            'hl': 10,          # Columna J (Horas Laboratorio)
            'modalidad': 14    # Columna N
        }
    
    def abrir_excel(self):
        """Abrir el archivo Excel"""
        try:
            logger.info(f"Abriendo Excel: {self.excel_path}")
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            logger.info(f"Excel abierto correctamente. Hojas: {self.workbook.sheetnames}")
            return True
        except Exception as e:
            logger.error(f"Error al abrir Excel: {e}")
            return False
    
    def extraer_cursos_hoja(self, sheet_name: str) -> List[Dict]:
        """Extraer cursos de una hoja específica"""
        logger.info(f"Extrayendo cursos de hoja: {sheet_name}")
        
        sheet = self.workbook[sheet_name]
        cursos = []
        
        # Comenzar desde la fila 2 (la 1 tiene headers)
        max_row = sheet.max_row
        
        for row in range(2, max_row + 1):
            try:
                # Extraer datos básicos
                ciclo_raw = sheet.cell(row=row, column=self.columnas['ciclo']).value
                asignatura_raw = sheet.cell(row=row, column=self.columnas['asignatura']).value
                modalidad_raw = sheet.cell(row=row, column=self.columnas['modalidad']).value
                
                # Saltar filas vacías o de totales
                if not ciclo_raw or not asignatura_raw or ciclo_raw in [None, '']:
                    continue
                
                # Saltar filas que no son de cursos (como totales)
                if not isinstance(ciclo_raw, str) or not ciclo_raw.startswith('C'):
                    continue
                
                # Procesar ciclo
                ciclo = self.procesar_ciclo(ciclo_raw)
                if not ciclo:
                    continue
                
                # Procesar nombre de asignatura
                nombre = str(asignatura_raw).strip()
                if not nombre or len(nombre) < 3:
                    continue
                
                # Procesar modalidad
                modalidad = self.procesar_modalidad(modalidad_raw)
                
                # Extraer horas y convertir a números de alumnos y grupos
                ht = self.extraer_numero(sheet.cell(row=row, column=self.columnas['ht']).value)
                hp = self.extraer_numero(sheet.cell(row=row, column=self.columnas['hp']).value)
                hl = self.extraer_numero(sheet.cell(row=row, column=self.columnas['hl']).value)
                creditos = self.extraer_numero(sheet.cell(row=row, column=self.columnas['creditos']).value)
                
                # Calcular grupos basado en horas (estimación)
                grupos_teoria = 1 if ht > 0 else 0
                grupos_practica = 1 if hp > 0 else 0
                grupos_laboratorio = 1 if hl > 0 else 0
                
                # Calcular alumnos estimados por grupo (valores típicos)
                alumnos_por_grupo_teoria = 40
                alumnos_por_grupo_practica = 20  
                alumnos_por_grupo_laboratorio = 15
                
                alumnos_teoria = grupos_teoria * alumnos_por_grupo_teoria if grupos_teoria > 0 else 0
                alumnos_practica = grupos_practica * alumnos_por_grupo_practica if grupos_practica > 0 else 0
                alumnos_laboratorio = grupos_laboratorio * alumnos_por_grupo_laboratorio if grupos_laboratorio > 0 else 0
                
                # Crear código único para el curso
                codigo = self.generar_codigo_curso(ciclo, nombre)
                
                # Determinar si requiere laboratorio/práctica
                requiere_laboratorio = hl > 0
                requiere_practica = hp > 0
                
                curso = {
                    'codigo': codigo,
                    'nombre': nombre,
                    'ciclo': ciclo,
                    'modalidad': modalidad,
                    'creditos': creditos,
                    'horas_teoria': ht,
                    'horas_practica': hp,
                    'horas_laboratorio': hl,
                    'alumnos_teoria': alumnos_teoria,
                    'alumnos_practica': alumnos_practica,
                    'alumnos_laboratorio': alumnos_laboratorio,
                    'grupos_teoria': grupos_teoria,
                    'grupos_practica': grupos_practica,
                    'grupos_laboratorio': grupos_laboratorio,
                    'requiere_laboratorio': requiere_laboratorio,
                    'requiere_practica': requiere_practica,
                    'fila_excel': row
                }
                
                cursos.append(curso)
                logger.debug(f"Curso extraído: {codigo} - {nombre} (Ciclo {ciclo})")
                
            except Exception as e:
                logger.error(f"Error procesando fila {row}: {e}")
                continue
        
        logger.info(f"Extraídos {len(cursos)} cursos de la hoja {sheet_name}")
        return cursos
    
    def procesar_ciclo(self, ciclo_raw) -> Optional[int]:
        """Procesar valor de ciclo"""
        if not ciclo_raw:
            return None
        
        ciclo_str = str(ciclo_raw).strip().upper()
        
        # Si ya es número
        if ciclo_str.isdigit():
            return int(ciclo_str)
        
        # Si está en formato C1, C2, etc.
        if ciclo_str in self.mapeo_ciclos:
            return self.mapeo_ciclos[ciclo_str]
        
        # Buscar patrón C + número
        match = re.search(r'C(\d+)', ciclo_str)
        if match:
            return int(match.group(1))
        
        logger.warning(f"No se pudo procesar ciclo: {ciclo_raw}")
        return None
    
    def procesar_modalidad(self, modalidad_raw) -> str:
        """Procesar modalidad"""
        if not modalidad_raw:
            return 'presencial'  # Por defecto
        
        modalidad_str = str(modalidad_raw).strip().upper()
        
        if modalidad_str in self.mapeo_modalidades:
            return self.mapeo_modalidades[modalidad_str]
        
        # Buscar patrones
        if 'NO' in modalidad_str or 'NPR' in modalidad_str:
            return 'no_presencial'
        
        return 'presencial'  # Por defecto
    
    def extraer_numero(self, valor) -> int:
        """Extraer número de una celda"""
        if valor is None:
            return 0
        
        if isinstance(valor, (int, float)):
            return max(0, int(valor))
        
        if isinstance(valor, str):
            # Extraer números de strings
            numeros = re.findall(r'\d+', valor)
            if numeros:
                return int(numeros[0])
        
        return 0
    
    def generar_codigo_curso(self, ciclo: int, nombre: str) -> str:
        """Generar código único para el curso"""
        # Limpiar nombre y tomar iniciales
        nombre_limpio = re.sub(r'[^A-Za-z\s]', '', nombre)
        palabras = nombre_limpio.upper().split()
        
        # Tomar las primeras 2-3 letras de las primeras 2-3 palabras
        iniciales = []
        for palabra in palabras[:3]:
            if len(palabra) >= 2:
                iniciales.append(palabra[:2])
            else:
                iniciales.append(palabra)
        
        codigo_base = ''.join(iniciales)[:6]  # Máximo 6 caracteres
        
        # Formato: C{ciclo}_{codigo}
        return f"C{ciclo}_{codigo_base}"
    
    def extraer_todos_cursos(self) -> List[Dict]:
        """Extraer cursos de todas las hojas"""
        if not self.abrir_excel():
            return []
        
        for sheet_name in self.workbook.sheetnames:
            cursos_hoja = self.extraer_cursos_hoja(sheet_name)
            self.cursos_extraidos.extend(cursos_hoja)
        
        self.workbook.close()
        
        # Eliminar duplicados por código
        cursos_unicos = {}
        for curso in self.cursos_extraidos:
            codigo = curso['codigo']
            if codigo not in cursos_unicos:
                cursos_unicos[codigo] = curso
        
        self.cursos_extraidos = list(cursos_unicos.values())
        
        logger.info(f"TOTAL CURSOS EXTRAÍDOS: {len(self.cursos_extraidos)}")
        return self.cursos_extraidos
    
    def mostrar_resumen(self):
        """Mostrar resumen de cursos extraídos"""
        if not self.cursos_extraidos:
            logger.warning("No hay cursos extraídos")
            return
        
        print("\n" + "="*80)
        print("RESUMEN DE CURSOS EXTRAÍDOS - ISIA")
        print("="*80)
        
        # Agrupar por ciclo
        por_ciclo = {}
        for curso in self.cursos_extraidos:
            ciclo = curso['ciclo']
            if ciclo not in por_ciclo:
                por_ciclo[ciclo] = []
            por_ciclo[ciclo].append(curso)
        
        total_creditos = 0
        
        for ciclo in sorted(por_ciclo.keys()):
            cursos_ciclo = por_ciclo[ciclo]
            creditos_ciclo = sum(curso['creditos'] for curso in cursos_ciclo)
            total_creditos += creditos_ciclo
            
            print(f"\n📚 CICLO {ciclo}: {len(cursos_ciclo)} cursos, {creditos_ciclo} créditos")
            print("-" * 60)
            
            for curso in cursos_ciclo:
                indicadores = []
                if curso['requiere_laboratorio']:
                    indicadores.append("🔬 LAB")
                if curso['requiere_practica']:
                    indicadores.append("🛠️ PRAC")
                indicadores_str = " " + " ".join(indicadores) if indicadores else ""
                
                print(f"  • {curso['codigo']}: {curso['nombre']}")
                print(f"    💳 {curso['creditos']} créditos | 📊 {curso['modalidad']}{indicadores_str}")
                print(f"    ⏰ T:{curso['horas_teoria']}h P:{curso['horas_practica']}h L:{curso['horas_laboratorio']}h")
                print(f"    👥 T:{curso['alumnos_teoria']} P:{curso['alumnos_practica']} L:{curso['alumnos_laboratorio']} alumnos")
                print()
        
        print(f"\n🎯 RESUMEN GENERAL:")
        print(f"   📖 Total cursos: {len(self.cursos_extraidos)}")
        print(f"   💳 Total créditos: {total_creditos}")
        print(f"   🎓 Ciclos: {len(por_ciclo)}")
        
        # Estadísticas por modalidad
        presenciales = sum(1 for curso in self.cursos_extraidos if curso['modalidad'] == 'presencial')
        no_presenciales = sum(1 for curso in self.cursos_extraidos if curso['modalidad'] == 'no_presencial')
        
        print(f"   🏢 Presenciales: {presenciales}")
        print(f"   💻 No presenciales: {no_presenciales}")
        
        # Estadísticas de laboratorios
        con_lab = sum(1 for curso in self.cursos_extraidos if curso['requiere_laboratorio'])
        con_prac = sum(1 for curso in self.cursos_extraidos if curso['requiere_practica'])
        
        print(f"   🔬 Con laboratorio: {con_lab}")
        print(f"   🛠️ Con práctica: {con_prac}")

def main():
    """Función principal de prueba"""
    extractor = ExtractorCursosV2()
    cursos = extractor.extraer_todos_cursos()
    
    if cursos:
        extractor.mostrar_resumen()
        
        # Guardar en JSON para revisión
        import json
        with open('cursos_isia_extraidos.json', 'w', encoding='utf-8') as f:
            json.dump(cursos, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ Cursos guardados en cursos_isia_extraidos.json")
    else:
        print("❌ No se pudieron extraer cursos")

if __name__ == "__main__":
    main()