#!/usr/bin/env python3
"""
Extractor EXACTO de nombres de profesores desde cada hoja del Excel
LEE ESPECÍFICAMENTE LA CELDA QUE DICE "Docente: [NOMBRE]"
"""
import pandas as pd
from openpyxl import load_workbook
import re

EXCEL_PATH = 'inputs/Horario_Docentes(2025-20).xlsx'

def extraer_nombres_exactos():
    print("=== EXTRACCIÓN EXACTA DE NOMBRES DE PROFESORES ===")
    
    wb = load_workbook(EXCEL_PATH)
    nombres_encontrados = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[HOJA] {sheet_name}")
        
        nombre_profesor = None
        
        # Buscar la celda que contiene "Docente:"
        for row in range(1, min(30, ws.max_row + 1)):  # Buscar en las primeras 30 filas
            for col in range(1, min(15, ws.max_column + 1)):  # Buscar en las primeras 15 columnas
                cell = ws.cell(row, col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    
                    # Buscar patrón "Docente:" seguido del nombre
                    if 'docente:' in cell_text.lower():
                        # Extraer el nombre después de "Docente:"
                        match = re.search(r'docente:\s*(.+)', cell_text, re.IGNORECASE)
                        if match:
                            nombre_profesor = match.group(1).strip()
                            print(f"  ✅ ENCONTRADO en celda {chr(64+col)}{row}: 'Docente: {nombre_profesor}'")
                            break
                    
                    # También buscar si hay una celda que solo contenga el nombre (después de encontrar "Docente:")
                    elif cell_text.lower() == 'docente:':
                        # Buscar en celdas adyacentes
                        for adj_col in range(col, min(col+5, ws.max_column + 1)):
                            adj_cell = ws.cell(row, adj_col)
                            if adj_cell.value and str(adj_cell.value).strip() != 'Docente:':
                                candidate = str(adj_cell.value).strip()
                                if len(candidate) > 5 and not candidate.isdigit():  # Validar que sea un nombre
                                    nombre_profesor = candidate
                                    print(f"  ✅ ENCONTRADO en celda adyacente {chr(64+adj_col)}{row}: '{nombre_profesor}'")
                                    break
                        if nombre_profesor:
                            break
            
            if nombre_profesor:
                break
        
        if nombre_profesor:
            nombres_encontrados[sheet_name] = nombre_profesor
            print(f"  📝 NOMBRE FINAL: '{nombre_profesor}'")
        else:
            print(f"  ❌ NO SE ENCONTRÓ nombre de profesor")
            # Buscar cualquier texto que pueda ser un nombre en la hoja
            print(f"  🔍 Buscando cualquier texto que pueda ser nombre...")
            for row in range(15, min(25, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    if cell.value:
                        text = str(cell.value).strip()
                        if len(text) > 8 and not text.isdigit() and 'hrs' not in text.lower():
                            print(f"    Posible nombre en {chr(64+col)}{row}: '{text}'")
    
    print(f"\n{'='*80}")
    print(f"RESUMEN DE NOMBRES ENCONTRADOS:")
    print(f"Total hojas: {len(wb.sheetnames)}")
    print(f"Nombres encontrados: {len(nombres_encontrados)}")
    print(f"Hojas sin nombre: {len(wb.sheetnames) - len(nombres_encontrados)}")
    
    print(f"\nLISTA COMPLETA DE NOMBRES:")
    for i, (hoja, nombre) in enumerate(nombres_encontrados.items(), 1):
        print(f"  {i:2d}. {hoja:15} -> '{nombre}'")
    
    return nombres_encontrados

if __name__ == "__main__":
    nombres = extraer_nombres_exactos()
    
    print(f"\n{'='*80}")
    print("CREANDO MAPEO PARA EL CÓDIGO:")
    print("MAPEO_EXCEL_REAL = {")
    for hoja, nombre in nombres.items():
        print(f"    '{hoja}': '{nombre}',")
    print("}")