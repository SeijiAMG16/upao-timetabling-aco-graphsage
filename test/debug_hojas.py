#!/usr/bin/env python3
"""
Debug específico del procesamiento de hojas
"""
import sys
sys.path.append('backend')

from openpyxl import load_workbook
import re

EXCEL_PATH = 'inputs/Horario_Docentes(2025-20).xlsx'

def simular_procesamiento_debug():
    print("=== DEBUG DEL PROCESAMIENTO DE HOJAS ===")
    
    wb = load_workbook(EXCEL_PATH)
    print(f"Total hojas en Excel: {len(wb.sheetnames)}")
    
    hojas_procesadas = 0
    hojas_saltadas = 0
    problemas = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[HOJA] {sheet_name}")
        
        # 1. Buscar fila de encabezado
        fila_encabezado = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and 'LUNES' in str(cell_value).upper():
                    fila_encabezado = row_idx
                    break
            if fila_encabezado:
                break
        
        if not fila_encabezado:
            print(f"  ❌ SALTAR: No se encontró encabezado con 'LUNES'")
            hojas_saltadas += 1
            problemas.append(f"{sheet_name}: Sin encabezado")
            continue
        
        # 2. Mapear columnas a días
        DIAS_MAPEO = {
            'LUNES': 'LUNES',
            'MARTES': 'MARTES',
            'MIÉRCOLES': 'MIERCOLES',
            'MIERCOLES': 'MIERCOLES',
            'JUEVES': 'JUEVES',
            'VIERNES': 'VIERNES',
            'SÁBADO': 'SABADO',
            'SABADO': 'SABADO'
        }
        
        columnas_dias = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(fila_encabezado, col).value
            if header:
                dia = str(header).strip().upper()
                if dia in DIAS_MAPEO:
                    columnas_dias[col] = DIAS_MAPEO[dia]
        
        if not columnas_dias:
            print(f"  ❌ SALTAR: No se encontraron columnas de días válidas")
            hojas_saltadas += 1
            problemas.append(f"{sheet_name}: Sin columnas de días")
            continue
        
        # 3. Identificar profesor (simplificado)
        profesor_encontrado = False
        for row_idx in range(fila_encabezado + 1, min(fila_encabezado + 20, ws.max_row + 1)):
            for col in columnas_dias.keys():
                cell = ws.cell(row_idx, col)
                if cell.value:
                    contenido = str(cell.value).strip()
                    # Simple check - si tiene texto que no sea solo números o espacios
                    if re.search(r'[A-Za-z]', contenido) and len(contenido) > 3:
                        profesor_encontrado = True
                        break
            if profesor_encontrado:
                break
        
        if not profesor_encontrado:
            print(f"  ❌ SALTAR: No se pudo identificar profesor en las celdas")
            hojas_saltadas += 1
            problemas.append(f"{sheet_name}: Sin identificación de profesor")
            continue
        
        print(f"  ✅ PROCESAR: Encabezado en fila {fila_encabezado}, {len(columnas_dias)} días, profesor identificado")
        hojas_procesadas += 1
    
    print(f"\n{'='*60}")
    print(f"RESUMEN:")
    print(f"  Hojas procesadas: {hojas_procesadas}")
    print(f"  Hojas saltadas: {hojas_saltadas}")
    print(f"  Total: {hojas_procesadas + hojas_saltadas}")
    
    if problemas:
        print(f"\nPROBLEMAS ENCONTRADOS:")
        for problema in problemas:
            print(f"  - {problema}")

if __name__ == "__main__":
    simular_procesamiento_debug()